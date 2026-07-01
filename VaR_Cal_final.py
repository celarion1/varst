import streamlit as st, numpy as np, pandas as pd
from scipy import stats, interpolate
from datetime import datetime, timedelta
import plotly.graph_objects as go, warnings
warnings.filterwarnings("ignore")

try:
    import yfinance as yf; YF_OK=True
except: YF_OK=False

st.set_page_config(page_title="Portfolio VaR Calculator",
                   page_icon="chart_with_upwards_trend",
                   layout="wide", initial_sidebar_state="expanded")
st.markdown("""<style>
.stTabs [data-baseweb="tab-list"]{gap:4px}
.stTabs [data-baseweb="tab"]{padding:6px 14px;font-size:13px;border-radius:6px 6px 0 0}
.mc{background:#f8f9fa;border:1px solid #e9ecef;border-radius:10px;padding:12px 16px;margin:4px 0}
.ml{font-size:12px;color:#6c757d;margin-bottom:3px}
.mv{font-size:21px;font-weight:600}
.red{color:#A32D2D}.grn{color:#3B6D11}.blu{color:#185FA5}.amb{color:#854F0B}
.ib{background:#f0f4ff;border-left:3px solid #185FA5;padding:9px 13px;
    border-radius:0 8px 8px 0;font-size:12px;color:#333;margin:6px 0;line-height:1.6}
</style>""", unsafe_allow_html=True)

# ── 종목 데이터베이스 ─────────────────────────────────────────────────────────
STOCKS = {
"KR-KOSPI":[
    ("005930.KS","삼성전자","반도체"),("000660.KS","SK하이닉스","반도체"),
    ("005380.KS","현대차","자동차"),("000270.KS","기아","자동차"),
    ("051910.KS","LG화학","화학"),("006400.KS","삼성SDI","화학"),
    ("035420.KS","NAVER","IT"),("035720.KS","카카오","IT"),
    ("105560.KS","KB금융","금융"),("055550.KS","신한지주","금융"),
    ("017670.KS","SK텔레콤","통신"),("207940.KS","삼성바이오로직스","바이오"),
    ("068270.KS","셀트리온","바이오"),("003670.KS","포스코홀딩스","철강"),
    ("028260.KS","삼성물산","건설"),("096770.KS","SK이노베이션","에너지"),
],
"KR-KOSDAQ":[
    ("247540.KQ","에코프로비엠","배터리"),("086520.KQ","에코프로","배터리"),
    ("035720.KQ","카카오게임즈","게임"),("263750.KQ","펄어비스","게임"),
    ("196170.KQ","알테오젠","바이오"),("145020.KQ","휴젤","바이오"),
    ("091990.KQ","셀트리온헬스케어","바이오"),("036570.KQ","엔씨소프트","게임"),
    ("112040.KQ","위메이드","게임"),("214150.KQ","클래시스","의료기기"),
    ("293490.KQ","카카오페이","핀테크"),("251270.KQ","넷마블","게임"),
],
"KR-BOND":[
    ("KR103501GA96","국고채3Y 3.500%","국채3년"),
    ("KR103502GB04","국고채5Y 3.625%","국채5년"),
    ("KR103503GC13","국고채10Y 3.875%","국채10년"),
    ("KR103504GD22","국고채20Y 4.000%","국채20년"),
    ("KR103505GE31","국고채30Y 4.125%","국채30년"),
    ("KR610001FA40","통안채1Y","통화안정"),
    ("KR610002FB49","통안채2Y","통화안정"),
    ("KR650001FC58","산업금융채3Y","특수채"),
    ("KR200002FD67","삼성전자채3Y","회사채"),
    ("KR200003FE76","LG화학채5Y","회사채"),
    ("KR200004FF85","현대차채3Y","회사채"),
    ("KR200005FG94","SK하이닉스채3Y","회사채"),
],
"KR-DERIV":[
    ("KOSPI200F-202506","KOSPI200 선물 2506","지수선물"),
    ("KOSPI200F-202509","KOSPI200 선물 2509","지수선물"),
    ("K200C-255-202506","KOSPI200 콜 255 2506","콜옵션"),
    ("K200P-255-202506","KOSPI200 풋 255 2506","풋옵션"),
    ("K200C-260-202506","KOSPI200 콜 260 2506","콜옵션"),
    ("K200P-250-202506","KOSPI200 풋 250 2506","풋옵션"),
    ("USDKRWF-202506","USD/KRW 선물 2506","통화선물"),
    ("EURKRWF-202506","EUR/KRW 선물 2506","통화선물"),
    ("KDTBF-202506","국채선물3Y 2506","채권선물"),
    ("KLTBF-202506","국채선물10Y 2506","채권선물"),
],
"US-NYSE":[
    ("JPM","JP Morgan Chase","금융"),("BAC","Bank of America","금융"),
    ("GS","Goldman Sachs","금융"),("MS","Morgan Stanley","금융"),
    ("XOM","Exxon Mobil","에너지"),("CVX","Chevron","에너지"),
    ("JNJ","Johnson & Johnson","헬스케어"),("KO","Coca-Cola","소비재"),
    ("PG","Procter & Gamble","소비재"),("WMT","Walmart","유통"),
    ("CAT","Caterpillar","산업"),("BA","Boeing","항공방산"),
    ("V","Visa","결제"),("T","AT&T","통신"),("GE","GE Aerospace","산업"),
],
"US-NASDAQ":[
    ("AAPL","Apple","빅테크"),("MSFT","Microsoft","빅테크"),
    ("GOOGL","Alphabet","빅테크"),("AMZN","Amazon","빅테크"),
    ("META","Meta Platforms","소셜미디어"),("NVDA","NVIDIA","반도체"),
    ("TSLA","Tesla","전기차"),("AVGO","Broadcom","반도체"),
    ("AMD","AMD","반도체"),("INTC","Intel","반도체"),
    ("NFLX","Netflix","스트리밍"),("ADBE","Adobe","소프트웨어"),
    ("QCOM","Qualcomm","반도체"),("PYPL","PayPal","핀테크"),
    ("COIN","Coinbase","크립토"),
],
"US-BOND":[
    ("TLT","iShares 20Y+ Treasury","장기국채ETF"),
    ("IEF","iShares 7-10Y Treasury","중기국채ETF"),
    ("SHY","iShares 1-3Y Treasury","단기국채ETF"),
    ("HYG","iShares High Yield Corp","하이일드ETF"),
    ("LQD","iShares IG Corp Bond","투자등급ETF"),
    ("BND","Vanguard Total Bond","채권혼합ETF"),
    ("AGG","iShares US Aggregate","채권혼합ETF"),
    ("TIP","iShares TIPS Bond","물가연동ETF"),
    ("EMB","iShares EM Bond","이머징ETF"),
    ("MBB","iShares MBS","MBS ETF"),
],
}

MKT = {
    "KR-KOSPI":  {"label":"KOSPI",  "ccy":"KRW","type":"Stock",
                  "info":"한국거래소 유가증권시장. 티커 형식: 005930.KS"},
    "KR-KOSDAQ": {"label":"KOSDAQ", "ccy":"KRW","type":"Stock",
                  "info":"코스닥시장. 티커 형식: 247540.KQ"},
    "KR-BOND":   {"label":"KR채권", "ccy":"KRW","type":"Bond",
                  "info":"국채·통안채·특수채·회사채. 수량=보유 매수(장단위), 액면가=10,000원/매"},
    "KR-DERIV":  {"label":"KR파생", "ccy":"KRW","type":"Option",
                  "info":"KOSPI200 선물·옵션, 통화선물, 채권선물 (장내파생)"},
    "US-NYSE":   {"label":"NYSE",   "ccy":"USD","type":"Stock",
                  "info":"뉴욕증권거래소. 평가금액 USD→KRW 자동 환산"},
    "US-NASDAQ": {"label":"NASDAQ", "ccy":"USD","type":"Stock",
                  "info":"나스닥. 평가금액 USD→KRW 자동 환산"},
    "US-BOND":   {"label":"US채권", "ccy":"USD","type":"Bond",
                  "info":"미국 채권 ETF (1주=1단위). USD→KRW 환산"},
}

Z99, Z95 = 2.3263, 1.6449
TENORS = np.array([0.25, 0.5, 1, 2, 3, 5, 7, 10, 20, 30])
COLORS = ["#185FA5","#A32D2D","#3B6D11","#854F0B","#534AB7","#0F6E56"]

def fmt(v):
    a = abs(v)
    if a >= 1e12: return f"{v/1e12:.2f}조"
    if a >= 1e8:  return f"{v/1e8:.2f}억"
    if a >= 1e4:  return f"{v/1e4:,.0f}만"
    return f"{v:,.0f}"

# ── 실시간 데이터 (yfinance) ──────────────────────────────────────────────────
@st.cache_data(ttl=300)
def get_usdkrw():
    if not YF_OK:
        return 1370.0, False
    try:
        tk = yf.Ticker("USDKRW=X")
        p = tk.fast_info.last_price
        if p and p > 0:
            return round(float(p), 2), True
    except Exception:
        pass
    return 1370.0, False

@st.cache_data(ttl=300)
def get_price(ticker):
    if not YF_OK:
        return None
    try:
        p = yf.Ticker(ticker).fast_info.last_price
        return float(p) if p and p > 0 else None
    except Exception:
        return None


@st.cache_data(ttl=300)
def get_prices_bulk(tickers_tuple):
    """
    복수 티커 현재가 일괄 조회 (yfinance download 1회 호출)
    tickers_tuple: tuple (캐시 키용)
    반환: {ticker: price} dict
    """
    tickers = list(tickers_tuple)
    result  = {}
    if not YF_OK or not tickers:
        return result
    try:
        # KRX 티커(.KS/.KQ)와 US 티커 분리 처리
        raw = yf.download(
            tickers, period="2d",
            progress=False, auto_adjust=True,
            group_by="ticker" if len(tickers) > 1 else "column",
        )
        if len(tickers) == 1:
            tk = tickers[0]
            if not raw.empty:
                result[tk] = float(raw["Close"].dropna().iloc[-1])
        else:
            for tk in tickers:
                try:
                    col = raw[tk]["Close"] if tk in raw else raw["Close"][tk]
                    v   = float(col.dropna().iloc[-1])
                    if v > 0:
                        result[tk] = v
                except Exception:
                    pass
    except Exception:
        pass
    # 실패한 티커는 개별 조회로 보완
    for tk in tickers:
        if tk not in result:
            p = get_price(tk)
            if p:
                result[tk] = p
    return result


def refresh_portfolio_prices(positions):
    """
    포트폴리오 전체 포지션의 현재가를 yfinance에서 일괄 조회하여 갱신
    Bond/Option은 가격 직접 갱신 대신 수익률 커브 재계산으로 처리
    반환: (updated_count, fail_list)
    """
    stock_positions = [
        (i, p) for i, p in enumerate(positions)
        if p["type"] in ("Stock",) and p.get("ticker")
    ]
    if not stock_positions:
        return 0, []

    tickers = tuple(set(p["ticker"] for _, p in stock_positions))
    prices  = get_prices_bulk(tickers)

    updated = 0
    failed  = []
    for i, p in stock_positions:
        tk = p["ticker"]
        if tk in prices and prices[tk] > 0:
            positions[i]["price"] = prices[tk]
            updated += 1
        else:
            failed.append(tk)
    return updated, failed

@st.cache_data(ttl=600)
def get_kr_curve():
    """
    KRW 국채 수익률 커브 (10개 테너)
    실제 환경: ECOS API / 금융투자협회 API 연동 권장
    현재: 최근 시장 수준 기본값 사용
    """
    base = np.array([3.45, 3.50, 3.52, 3.55, 3.58,
                     3.62, 3.65, 3.68, 3.70, 3.72]) / 100
    if not YF_OK:
        return base, False
    # yfinance KRX 금리 직접 시리즈 없음 → fallback
    return base, False

@st.cache_data(ttl=600)
def get_us_curve():
    fb = np.array([5.25, 4.90, 4.60, 4.35, 4.25,
                   4.15, 4.08, 4.02, 3.98, 3.95]) / 100
    if not YF_OK:
        return fb, False
    try:
        pts = {"0.25": "^IRX", "2": "^TYX", "5": "^FVX", "10": "^TNX"}
        rates = {}
        for k, tk in pts.items():
            d = yf.download(tk, period="5d", progress=False, auto_adjust=True)
            if not d.empty:
                rates[float(k)] = float(d["Close"].iloc[-1]) / 100
        if len(rates) >= 3:
            xs = sorted(rates.keys())
            ys = [rates[x] for x in xs]
            f = interpolate.interp1d(xs, ys, kind="linear",
                                     fill_value="extrapolate")
            return f(TENORS), True
    except Exception:
        pass
    return fb, False


# ── 채권 계산 (KRD, 액면가 기준) ─────────────────────────────────────────────
def bond_var_calc(face, coupon, maturity, freq, yc, sqrt_h=1.0):
    """
    face  : 액면가 (원, 통상 10,000원/매)
    coupon: 표면이율 (예: 0.035)
    maturity: 잔존만기 (년)
    freq  : 이표 지급 횟수/년 (통상 2)
    반환: pv(원/매), pv99, pv95
    """
    periods = max(1, int(round(maturity * freq)))
    times   = np.array([(i + 1) / freq for i in range(periods)])
    cpn     = face * coupon / freq
    cf      = np.full(periods, cpn)
    cf[-1] += face
    spot    = np.interp(times, TENORS, yc)
    disc    = np.exp(-spot * times)
    pv      = float((cf * disc).sum())          # 원/매
    if pv <= 0:
        return {"pv": pv, "pv99": 0.0, "pv95": 0.0}
    # KRD
    krd = np.zeros(len(TENORS))
    for j, t in enumerate(TENORS):
        if j == 0:
            mask = times <= t
        elif j < len(TENORS) - 1:
            mask = (times > TENORS[j-1]) & (times <= t)
        else:
            mask = times > TENORS[j-1]
        if mask.any():
            krd[j] = (cf[mask] * disc[mask] * times[mask]).sum() / pv
    # 금리 변동성 (테너별, 연율)
    rv  = np.array([0.012,0.011,0.010,0.009,0.009,
                    0.008,0.008,0.007,0.007,0.006]) * sqrt_h
    idx = np.arange(len(TENORS))
    cov = np.diag(rv) @ (0.95 ** np.abs(idx[:,None]-idx[None,:])) @ np.diag(rv)
    kw  = krd * pv      # KRD × 현재가 (단위: 원/매)
    var_sq = float(kw @ cov @ kw)
    pv99 = float(Z99 * np.sqrt(max(var_sq, 0)))
    pv95 = float(Z95 * np.sqrt(max(var_sq, 0)))
    return {"pv": pv, "pv99": pv99, "pv95": pv95}


# ── 옵션 Black-Scholes ────────────────────────────────────────────────────────
def bs_greeks(S, K, T, r, sig, flag):
    if T <= 1e-6:
        pr = max(S - K, 0) if flag == "C" else max(K - S, 0)
        return {"price": pr, "delta": (1 if flag=="C" else -1),
                "gamma": 0.0, "vega": 0.0}
    d1 = (np.log(S/K) + (r + 0.5*sig**2)*T) / (sig*np.sqrt(T))
    d2 = d1 - sig*np.sqrt(T)
    pdf1 = stats.norm.pdf(d1)
    if flag == "C":
        pr    = S*stats.norm.cdf(d1) - K*np.exp(-r*T)*stats.norm.cdf(d2)
        delta = stats.norm.cdf(d1)
    else:
        pr    = K*np.exp(-r*T)*stats.norm.cdf(-d2) - S*stats.norm.cdf(-d1)
        delta = stats.norm.cdf(d1) - 1
    gamma = pdf1 / (S * sig * np.sqrt(T))
    vega  = S * pdf1 * np.sqrt(T) / 100
    return {"price": pr, "delta": delta, "gamma": gamma, "vega": vega}

def option_var(S, K, T, r, iv, flag, qty, mult, sqrt_h):
    g    = bs_greeks(S, K, T, r, iv, flag)
    s_h  = iv / np.sqrt(252) * sqrt_h
    v99  = (abs(g["delta"])*S*Z99*s_h
            - 0.5*g["gamma"]*S**2*(Z99**2-1)*s_h**2) * qty * mult
    v95  = (abs(g["delta"])*S*Z95*s_h
            - 0.5*g["gamma"]*S**2*(Z95**2-1)*s_h**2) * qty * mult
    return {"exposure": g["price"]*qty*mult,
            "price": g["price"],
            "pv99": abs(v99), "pv95": abs(v95),
            "delta": g["delta"], "gamma": g["gamma"]}


# ── 포트폴리오 VaR ────────────────────────────────────────────────────────────
def calc_pf_var(positions, usdkrw, c_sb, c_so, c_bo, sqrt_h, ykr, yus):
    if not positions:
        return None
    res = []
    for p in positions:
        fx    = usdkrw if p["ccy"] == "USD" else 1.0
        is_us = p["market"] in ("US-NYSE","US-NASDAQ","US-BOND")
        yc    = yus if is_us else ykr
        if p["type"] == "Bond":
            fv  = p.get("face_value", 10000)   # 액면가 원/매
            qty = p["qty"]                      # 보유 매수
            bv  = bond_var_calc(fv, p.get("coupon", 0.04),
                                p.get("maturity", 3.0),
                                int(p.get("cf_freq", 2)), yc, sqrt_h)
            mv  = bv["pv"] * qty * fx          # 총 평가금액
            res.append({"pos":p, "mv":mv,
                        "pv99": bv["pv99"]*qty*fx,
                        "pv95": bv["pv95"]*qty*fx,
                        "hv99": bv["pv99"]*qty*fx*1.05,
                        "mv99": bv["pv99"]*qty*fx*0.75})
        elif p["type"] in ("Option","Future"):
            ov  = option_var(p["price"],
                             p.get("strike", p["price"]*1.05),
                             p.get("mat", 0.25),
                             p.get("rf", 0.035),
                             p.get("iv", 0.18),
                             p.get("callput", "C"),
                             p["qty"], p.get("mult", 1.0), sqrt_h)
            mv  = ov["exposure"] * fx
            res.append({"pos":p, "mv":mv,
                        "pv99": ov["pv99"]*fx, "pv95": ov["pv95"]*fx,
                        "hv99": ov["pv99"]*fx*1.08, "mv99": ov["pv99"]*fx*0.80,
                        "delta": ov["delta"]})
        else:   # Stock
            sig = (0.23 if is_us else 0.22) / np.sqrt(252) * sqrt_h
            mv  = p["price"] * p["qty"] * fx
            res.append({"pos":p, "mv":mv,
                        "pv99": mv*sig*Z99, "pv95": mv*sig*Z95,
                        "hv99": mv*sig*Z99*1.05, "mv99": mv*sig*Z99*0.75})

    total = sum(r["mv"] for r in res)

    def agg(lst):
        if not lst:
            return {"pv99":0,"pv95":0,"hv99":0,"total":0,"cnt":0}
        t = sum(r["mv"] for r in lst)
        v = sum(r["pv99"] for r in lst) * 0.85
        return {"pv99":v, "pv95":v*(Z95/Z99),
                "hv99":v*1.05, "total":t, "cnt":len(lst)}

    sv = agg([r for r in res if r["pos"]["type"]=="Stock"])
    bv = agg([r for r in res if r["pos"]["type"]=="Bond"])
    ov = agg([r for r in res if r["pos"]["type"] in ("Option","Future")])
    v3 = np.array([sv["pv99"], bv["pv99"], ov["pv99"]])
    C  = np.array([[1,c_sb,c_so],[c_sb,1,c_bo],[c_so,c_bo,1]])
    pv99  = float(np.sqrt(max(float(v3@C@v3), 0)))
    pv95  = pv99 * (Z95/Z99)
    undiv = float(v3.sum())
    return {"total":total, "pv99":pv99, "pv95":pv95,
            "undiv":undiv, "divers":undiv-pv99,
            "sv":sv, "bv":bv, "ov":ov,
            "pos_results":res, "ts":datetime.now()}


# ── VaR/Exposure 추이 데이터 생성 ────────────────────────────────────────────
def gen_vo_data(positions, usdkrw):
    if not positions:
        return None
    total = sum(p["price"] * p["qty"] * (usdkrw if p["ccy"]=="USD" else 1)
                for p in positions)
    sig   = 0.22 / np.sqrt(252)
    rng   = np.random.default_rng(42)
    exp, var, dates = [], [], []
    mv    = total
    today = datetime.today()
    n_total, n_show = 504, 252    # 2년 기준, 후반 1년 표시
    for i in range(n_total):
        mv  *= np.exp(rng.normal(-0.5*sig**2, sig))
        hvol = sig * (0.8 + rng.random()*0.4)
        if i >= n_total - n_show:
            exp.append(mv / 1e6)
            var.append(mv * hvol * Z99 / 1e6)
            dates.append(today - timedelta(days=n_total-i))
    return {"exp": exp, "var": var, "dates": dates}

# ══════════════════════════════════════════════════════════════════════════════
# EWMA 기반 고급 VaR 엔진
# ══════════════════════════════════════════════════════════════════════════════

def ewma_volatility(returns, lam=0.94):
    """
    RiskMetrics EWMA 변동성 추정
    lam  : 감쇄 계수 (RiskMetrics 일별 기본값 0.94)
    반환  : 각 시점의 분산 시계열 (numpy array)
    """
    n      = len(returns)
    var_t  = np.zeros(n)
    var_t[0] = returns[0] ** 2
    for t in range(1, n):
        var_t[t] = lam * var_t[t-1] + (1 - lam) * returns[t-1] ** 2
    return var_t                    # 분산 시계열


def ewma_cov_matrix(ret_matrix, lam=0.94):
    """
    다변량 EWMA 공분산 행렬
    ret_matrix : shape (T, N) — T=관측수, N=종목수
    반환        : shape (N, N) 공분산 행렬 (최신 시점)
    """
    T, N    = ret_matrix.shape
    cov     = np.cov(ret_matrix[:10].T) if T > 10 else np.eye(N) * 0.001
    for t in range(T):
        r   = ret_matrix[t:t+1].T          # (N,1)
        cov = lam * cov + (1 - lam) * (r @ r.T)
    return cov


def fetch_returns_multi(tickers, days=504):
    """
    복수 종목 수익률 행렬 수집 (yfinance / fallback)
    반환: dict {ticker: returns_array}, prices_dict, live_flag
    """
    returns_dict = {}
    prices_dict  = {}
    live_any     = False

    for tk in tickers:
        rets, price, live = sim_fetch_returns(tk, days)
        returns_dict[tk]  = rets
        prices_dict[tk]   = price
        if live:
            live_any = True

    # 길이 맞추기 (최솟값 기준)
    min_len = min(len(v) for v in returns_dict.values())
    for tk in tickers:
        returns_dict[tk] = returns_dict[tk][-min_len:]

    return returns_dict, prices_dict, live_any


def parametric_var_ewma(returns, mv, holding=10, conf=0.99, lam=0.94):
    """
    EWMA 기반 Parametric VaR (바젤 10일, 99% 기준)
    반환: dict(var, cvar, sigma_daily, sigma_10d)
    """
    var_t     = ewma_volatility(returns, lam)
    sigma_d   = float(np.sqrt(var_t[-1]))          # 최신 일별 변동성
    sigma_10d = sigma_d * np.sqrt(holding)          # 보유기간 조정
    z         = stats.norm.ppf(conf)
    var_amt   = mv * sigma_10d * z                  # 금액 VaR
    cvar_amt  = mv * sigma_10d * stats.norm.pdf(z) / (1 - conf)
    return {
        "sigma_daily":  sigma_d,
        "sigma_10d":    sigma_10d,
        "var":          var_amt,
        "cvar":         cvar_amt,
        "var_pct":      sigma_10d * z,
        "cvar_pct":     sigma_10d * stats.norm.pdf(z) / (1 - conf),
    }


def marginal_var_ewma(weights, cov_matrix, pf_mv, conf=0.99, holding=10):
    """
    Marginal VaR = ∂VaR_pf / ∂w_i × (VaR_pf / pf_mv)
    = (Cov × w) / σ_pf × z × √holding
    반환: marginal VaR 벡터 (금액), component VaR 벡터
    """
    w         = np.array(weights)
    pf_var_d  = float(w @ cov_matrix @ w)           # 포트폴리오 일별 분산
    pf_sig_d  = float(np.sqrt(max(pf_var_d, 1e-12)))
    pf_sig_h  = pf_sig_d * np.sqrt(holding)
    z         = stats.norm.ppf(conf)

    cov_w      = cov_matrix @ w                      # shape (N,)
    beta       = cov_w / (pf_sig_d ** 2 + 1e-12)    # 종목 베타
    # Marginal VaR (금액 단위): ∂VaR/∂w_i × PF MV
    mvar_vec   = (cov_w / pf_sig_d) * z * np.sqrt(holding) * pf_mv
    # Component VaR = w_i × Marginal VaR_i
    comp_var   = w * mvar_vec
    pf_var_amt = pf_sig_h * z * pf_mv

    return {
        "pf_var":    pf_var_amt,
        "pf_sigma":  pf_sig_h,
        "mvar":      mvar_vec,               # Marginal VaR 벡터
        "comp_var":  comp_var,               # Component VaR 벡터
        "beta":      beta,
        "pf_var_d":  pf_var_d,
    }


def incremental_var_ewma(weights, cov_matrix, pf_mv, conf=0.99, holding=10):
    """
    Incremental VaR: 각 종목을 제거했을 때 포트폴리오 VaR 변화
    반환: incremental VaR 벡터 (금액)
    """
    N         = len(weights)
    w         = np.array(weights)
    z         = stats.norm.ppf(conf)
    sq_h      = np.sqrt(holding)

    base_sig  = float(np.sqrt(max(w @ cov_matrix @ w, 1e-12)))
    base_var  = base_sig * sq_h * z * pf_mv

    ivar_vec  = np.zeros(N)
    for i in range(N):
        mask        = np.ones(N, dtype=bool)
        mask[i]     = False
        if mask.sum() == 0:
            ivar_vec[i] = base_var
            continue
        w_ex        = w[mask]
        w_ex_norm   = w_ex / w_ex.sum() if w_ex.sum() > 0 else w_ex
        c_ex        = cov_matrix[np.ix_(mask, mask)]
        sig_ex      = float(np.sqrt(max(w_ex_norm @ c_ex @ w_ex_norm, 1e-12)))
        var_ex      = sig_ex * sq_h * z * pf_mv * w_ex.sum()
        ivar_vec[i] = base_var - var_ex     # 제거 시 VaR 감소량

    return ivar_vec, base_var


# ── Historical VaR ────────────────────────────────────────────────────────────
def historical_var(returns, mv, holding=10, conf=0.99):
    """
    Historical Simulation VaR (개별 종목)
    과거 실제 수익률 분포에서 분위수 직접 추출 — 모수 가정 없음
    """
    sqrt_h    = np.sqrt(holding)
    rets_h    = returns * sqrt_h
    alpha     = 1 - conf
    var_pct   = float(-np.percentile(rets_h, alpha * 100))
    cvar_mask = rets_h <= -var_pct
    cvar_pct  = float(-rets_h[cvar_mask].mean()) if cvar_mask.any() else var_pct
    return {
        "var":          var_pct * mv,
        "cvar":         cvar_pct * mv,
        "var_pct":      var_pct,
        "cvar_pct":     cvar_pct,
        "hist_returns": rets_h.tolist(),
    }


def historical_portfolio_var(ret_dict, tickers, weights, total_mv,
                              holding=10, conf=0.99):
    """포트폴리오 Historical VaR — 가중 수익률 합산 후 분위수 추출"""
    min_len  = min(len(ret_dict[tk]) for tk in tickers)
    pf_rets  = np.zeros(min_len)
    for tk, w in zip(tickers, weights):
        pf_rets += ret_dict[tk][-min_len:] * w
    sqrt_h    = np.sqrt(holding)
    rets_h    = pf_rets * sqrt_h
    alpha     = 1 - conf
    var_pct   = float(-np.percentile(rets_h, alpha * 100))
    cvar_mask = rets_h <= -var_pct
    cvar_pct  = float(-rets_h[cvar_mask].mean()) if cvar_mask.any() else var_pct
    ind_hvars = []
    for tk, wt in zip(tickers, weights):
        r  = ret_dict[tk][-min_len:] * sqrt_h
        vp = float(-np.percentile(r, alpha * 100))
        ind_hvars.append(vp * wt * total_mv)
    return {
        "pf_var":     var_pct * total_mv,
        "pf_cvar":    cvar_pct * total_mv,
        "var_pct":    var_pct,
        "ind_hvars":  ind_hvars,
        "pf_returns": rets_h.tolist(),
    }


# ── Monte Carlo VaR ───────────────────────────────────────────────────────────
def monte_carlo_var(returns, mv, holding=10, conf=0.99,
                    n_sims=10000, use_t=True, df=5, lam=0.94, seed=42):
    """
    Monte Carlo VaR (개별 종목)
    EWMA 변동성 + t-분포(use_t=True) 또는 정규분포 기반 시뮬레이션
    """
    rng     = np.random.default_rng(seed)
    var_t   = ewma_volatility(returns, lam)
    sigma_d = float(np.sqrt(var_t[-1]))
    sqrt_h  = np.sqrt(holding)
    mu_d    = float(np.mean(returns))
    if use_t and df > 2:
        scale    = sigma_d * np.sqrt((df - 2) / df)
        z_draws  = stats.t.rvs(df=df, size=n_sims,
                               random_state=int(rng.integers(0, 2**31)))
        sim_rets = mu_d * holding + scale * sqrt_h * z_draws
    else:
        sim_rets = rng.normal(mu_d * holding, sigma_d * sqrt_h, n_sims)
    alpha     = 1 - conf
    var_pct   = float(-np.percentile(sim_rets, alpha * 100))
    cvar_mask = sim_rets <= -var_pct
    cvar_pct  = float(-sim_rets[cvar_mask].mean()) if cvar_mask.any() else var_pct
    return {
        "var":         var_pct * mv,
        "cvar":        cvar_pct * mv,
        "var_pct":     var_pct,
        "cvar_pct":    cvar_pct,
        "sim_returns": sim_rets.tolist(),
    }


def monte_carlo_portfolio_var(ret_dict, tickers, weights, total_mv,
                               holding=10, conf=0.99,
                               n_sims=10000, use_t=True, df=5,
                               lam=0.94, seed=42):
    """포트폴리오 Monte Carlo VaR — EWMA 공분산 + 다변량 시뮬레이션"""
    rng      = np.random.default_rng(seed)
    min_len  = min(len(ret_dict[tk]) for tk in tickers)
    ret_mat  = np.column_stack([ret_dict[tk][-min_len:] for tk in tickers])
    cov_mat  = ewma_cov_matrix(ret_mat, lam)
    mu_vec   = ret_mat.mean(axis=0) * holding
    cov_h    = cov_mat * holding
    w        = np.array(weights)
    if use_t and df > 2:
        L       = np.linalg.cholesky(cov_h + np.eye(len(w)) * 1e-10)
        z_raw   = stats.t.rvs(df=df, size=(n_sims, len(w)),
                              random_state=int(rng.integers(0, 2**31)))
        scale   = np.sqrt((df - 2) / df)
        sim_mat = mu_vec + (z_raw * scale) @ L.T
    else:
        sim_mat = rng.multivariate_normal(mu_vec, cov_h, n_sims)
    pf_rets   = sim_mat @ w
    alpha     = 1 - conf
    var_pct   = float(-np.percentile(pf_rets, alpha * 100))
    cvar_mask = pf_rets <= -var_pct
    cvar_pct  = float(-pf_rets[cvar_mask].mean()) if cvar_mask.any() else var_pct
    ind_mcvars = []
    for i in range(len(tickers)):
        vp = float(-np.percentile(sim_mat[:, i], alpha * 100))
        ind_mcvars.append(vp * total_mv * w[i])
    return {
        "pf_var":     var_pct * total_mv,
        "pf_cvar":    cvar_pct * total_mv,
        "var_pct":    var_pct,
        "ind_mcvars": ind_mcvars,
        "pf_returns": pf_rets.tolist(),
    }


# ── 포트폴리오 파일 파서 ──────────────────────────────────────────────────────
def parse_portfolio_file(uploaded_file):
    """
    Excel(.xlsx/.xls) 또는 CSV 파일로부터 포지션 목록 파싱
    필수 컬럼: ticker, type, qty, price
    선택 컬럼: ccy, market, coupon, maturity, cf_freq, face_value,
               strike, callput, iv, mat, rf, mult
    반환: (positions list, error_msg or None)
    """
    import io
    try:
        fname = uploaded_file.name.lower()
        if fname.endswith(".csv"):
            df = pd.read_csv(uploaded_file)
        elif fname.endswith((".xlsx", ".xls")):
            df = pd.read_excel(uploaded_file)
        else:
            return [], "지원 형식: .csv, .xlsx, .xls"

        df.columns = [c.strip().lower().replace(" ","_") for c in df.columns]
        required   = {"ticker","type","qty","price"}
        missing    = required - set(df.columns)
        if missing:
            return [], f"필수 컬럼 누락: {missing}  (현재 컬럼: {list(df.columns)})"

        if "ccy" not in df.columns:
            df["ccy"] = "KRW"
        if "market" not in df.columns:
            def guess(row):
                tk = str(row.get("ticker","")).upper()
                t  = str(row.get("type","")).capitalize()
                if t == "Bond":  return "KR-BOND" if not tk.startswith(("TLT","IEF","SHY","HYG","LQD","BND","AGG")) else "US-BOND"
                if t in ("Option","Future"): return "KR-DERIV"
                if tk.endswith(".KS"): return "KR-KOSPI"
                if tk.endswith(".KQ"): return "KR-KOSDAQ"
                return "US-NASDAQ"
            df["market"] = df.apply(guess, axis=1)

        positions = []
        for _, row in df.iterrows():
            try:
                pos = {
                    "ticker":  str(row["ticker"]).strip().upper(),
                    "type":    str(row["type"]).strip().capitalize(),
                    "qty":     float(row["qty"]),
                    "price":   float(row["price"]),
                    "ccy":     str(row.get("ccy","KRW")).strip().upper(),
                    "market":  str(row.get("market","KR-KOSPI")).strip(),
                    "mult":    float(row["mult"]) if "mult" in row and pd.notna(row.get("mult")) else 1.0,
                }
                for field in ["coupon","maturity","cf_freq","face_value",
                              "strike","callput","iv","mat","rf"]:
                    if field in row and pd.notna(row.get(field)):
                        pos[field] = row[field]
                positions.append(pos)
            except Exception:
                continue
        if not positions:
            return [], "파싱된 포지션이 없습니다. 파일 형식을 확인하세요."
        return positions, None
    except Exception as e:
        return [], f"파일 파싱 오류: {str(e)}"


# ══════════════════════════════════════════════════════════════════════════════
# 리포트 생성 엔진
# ══════════════════════════════════════════════════════════════════════════════

def build_md_report(pf_name, positions, usdkrw,
                    ewma_results=None, sim_results=None,
                    bt_result=None, mc_var_result=None,
                    author="Risk Management"):
    """
    포트폴리오 분석 결과 종합 마크다운 리포트 생성
    ewma_results  : st.session_state["ewma_result"]   (Parametric + Historical VaR)
    sim_results   : st.session_state["sim_result_for_report"]
    bt_result     : st.session_state["bt_result"]
    mc_var_result : st.session_state["mc_var_result"] (Monte Carlo VaR)
    """
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total_mv = sum(
        p["price"] * p["qty"] * (usdkrw if p["ccy"] == "USD" else 1)
        for p in positions
    )
    n_stock = sum(1 for p in positions if p["type"] == "Stock")
    n_bond  = sum(1 for p in positions if p["type"] == "Bond")
    n_deriv = sum(1 for p in positions if p["type"] in ("Option","Future"))

    L = []
    def a(t=""): L.append(t)

    # ── 표지 ────────────────────────────────────────────────────────────
    a("# 포트폴리오 리스크 분석 리포트")
    a()
    a(f"| 항목 | 내용 |")
    a(f"|------|------|")
    a(f"| 포트폴리오 | {pf_name} |")
    a(f"| 작성일시 | {now} |")
    a(f"| 작성자 | {author} |")
    a(f"| USD/KRW | {usdkrw:,} |")
    a(); a("---"); a()

    # ── 1. 포트폴리오 구성 ────────────────────────────────────────────
    a("## 1. 포트폴리오 구성"); a()
    a(f"- **총 평가금액**: {fmt(total_mv)}원")
    a(f"- **종목 수**: {len(positions)}개 "
      f"(주식 {n_stock} / 채권 {n_bond} / 파생 {n_deriv})")
    a()
    a("| 티커 | 유형 | 시장 | 수량 | 평가금액(KRW) | 비중(%) |")
    a("|------|------|------|-----:|-------------:|-------:|")
    for p in positions:
        mv  = p["price"] * p["qty"] * (usdkrw if p["ccy"] == "USD" else 1)
        pct = mv / total_mv * 100 if total_mv else 0
        a(f"| {p['ticker']} | {p['type']} | {MKT[p['market']]['label']} "
          f"| {p['qty']:,.0f} | {fmt(mv)}원 | {pct:.1f}% |")
    a(); a("---"); a()

    # ── 2. Parametric VaR + Historical VaR ───────────────────────────
    a("## 2. Parametric VaR + Historical Simulation VaR (EWMA λ=0.94)"); a()
    if ewma_results and ewma_results.get("portfolio"):
        pf_d = ewma_results["portfolio"]
        a("### 2-1. 포트폴리오 VaR 요약"); a()
        a("| 지표 | 값 |")
        a("|------|----|")
        for k, v in pf_d.items():
            val_str = f"{fmt(v)}원" if isinstance(v, (int, float)) else str(v)
            a(f"| {k} | {val_str} |")
        a()
        if ewma_results.get("individual"):
            a("### 2-2. 종목별 리스크 기여도 (Parametric VaR 기준)"); a()
            ind = ewma_results["individual"]
            if ind:
                hdr = " | ".join(str(k) for k in ind[0].keys())
                sep = " | ".join(["---"] * len(ind[0]))
                a(f"| {hdr} |"); a(f"| {sep} |")
                for row in ind:
                    a(f"| {' | '.join(str(v) for v in row.values())} |")
        a()
        a("> **Parametric VaR**: EWMA 변동성 × z-score × √h. 정규분포 가정, 계산 효율 높음.")
        a(">")
        a("> **Historical VaR**: 과거 실제 수익률 분위수 직접 추출. 분포 가정 없음, Fat-tail 반영.")
    else:
        a("> *'Parametric VaR' 탭에서 계산 실행 후 리포트를 생성하세요.*")
    a(); a("---"); a()

    # ── 2-MC. Monte Carlo VaR (선택) ─────────────────────────────────
    a("## 2-MC. Monte Carlo Simulation VaR"); a()
    if mc_var_result:
        conf_mc   = mc_var_result.get("conf", 0.99)
        hold_mc   = mc_var_result.get("holding", 10)
        dist_mc   = mc_var_result.get("dist_type", "t-분포")
        n_sims_mc = mc_var_result.get("n_sims", 10000)
        a(f"- **분포**: {dist_mc} | **시뮬레이션**: {n_sims_mc:,}회 "
          f"| **신뢰수준**: {int(conf_mc*100)}% | **보유기간**: {hold_mc}일")
        a()
        a("| 지표 | 값 |")
        a("|------|----|")
        a(f"| 포트폴리오 MC VaR | {fmt(mc_var_result['pf_var'])}원 "
          f"({mc_var_result['var_pct']*100:.3f}%) |")
        a(f"| 포트폴리오 MC CVaR | {fmt(mc_var_result['pf_cvar'])}원 |")
        a()
        if mc_var_result.get("ind_rows"):
            a("### 종목별 Monte Carlo VaR"); a()
            rows_mc = mc_var_result["ind_rows"]
            if rows_mc:
                hdr = " | ".join(str(k) for k in rows_mc[0].keys())
                sep = " | ".join(["---"] * len(rows_mc[0]))
                a(f"| {hdr} |"); a(f"| {sep} |")
                for row in rows_mc:
                    a(f"| {' | '.join(str(v) for v in row.values())} |")
        a()
        a("> **Monte Carlo VaR**: EWMA 공분산 + 다변량 시뮬레이션. "
          "Fat-tail 반영(t-분포)으로 극단 손실을 보수적으로 추정합니다.")
    else:
        a("> *'Parametric VaR' 탭에서 Monte Carlo VaR를 별도 실행 후 리포트를 생성하세요.*")
    a(); a("---"); a()

    # ── 3. 시뮬레이션 결과 ────────────────────────────────────────────
    a("## 3. 시뮬레이션 분석 결과"); a()
    if sim_results and sim_results.get("metrics"):
        m_dict   = sim_results["metrics"]
        tk_sim   = sim_results.get("ticker", "")
        T_days   = sim_results.get("T_days", 63)
        n_paths  = sim_results.get("n_paths", 2000)
        conf_sim = sim_results.get("conf", 0.99)
        # conf가 "99%" 문자열로 저장된 경우 방어 처리
        if isinstance(conf_sim, str):
            conf_sim = float(conf_sim.replace("%","")) / 100
        c_str    = f"{int(conf_sim*100)}%"
        vk       = f"VaR {int(conf_sim*100)}%"
        ck       = f"CVaR {int(conf_sim*100)}%"
        method_map = {
            "MJD": "Merton Jump Diffusion",
            "Bootstrap": "Historical Block Bootstrap",
            "t-MC": "t-분포 Monte Carlo",
        }
        a(f"- **분석 종목**: {tk_sim} | **예측 기간**: {T_days}일 "
          f"| **경로 수**: {n_paths:,}개 | **신뢰수준**: {c_str}")
        a()
        a(f"| 기법 | 기대수익률 | VaR {c_str} | CVaR {c_str} | 손실확률 | 기대주가 |")
        a(f"|------|----------:|--------:|--------:|-------:|-------:|")
        best_k  = min(m_dict, key=lambda k: abs(m_dict[k].get(vk, 0)))
        worst_k = max(m_dict, key=lambda k: abs(m_dict[k].get(vk, 0)))
        for key, m in m_dict.items():
            lbl = method_map.get(key, key)
            mark = " ★" if key == worst_k else ""
            a(f"| {lbl}{mark} "
              f"| {m.get('기대수익률(로그)', 0)*100:.2f}% "
              f"| {m.get(vk, 0)*100:.2f}% "
              f"| {m.get(ck, 0)*100:.2f}% "
              f"| {m.get('손실확률', 0)*100:.1f}% "
              f"| {m.get('기대 주가', 0):,.0f} |")
        a()
        a(f"> ★ 가장 보수적 기법 (높은 VaR): **{method_map.get(worst_k, worst_k)}**")
        a(f"> 낙관적 기법 (낮은 VaR): **{method_map.get(best_k, best_k)}**")
    else:
        a("> *시뮬레이션 분석 탭에서 실행 완료 후 리포트를 생성하세요.*")
    a(); a("---"); a()

    # ── 4. Back-testing 결과 ─────────────────────────────────────────
    a("## 4. VaR Back-testing 결과"); a()
    if bt_result:
        a(f"- **분석 대상**: {bt_result['ticker']} "
          f"| **신뢰수준**: {int(bt_result['conf']*100)}% "
          f"| **검증 기간**: {bt_result['window']}일")
        a()
        a("| 지표 | 값 |")
        a("|------|----|")
        a(f"| VaR 초과 횟수 | {bt_result['exceptions']}일 |")
        a(f"| 초과 비율 | {bt_result['exc_rate']:.2f}% |")
        a(f"| Basel 판정 | {bt_result['zone']} |")
        a(f"| Kupiec LR | {bt_result['lr_kupiec']:.3f} (p={bt_result['p_kupiec']:.4f}) |")
        a(f"| Kupiec 결과 | {bt_result['kupiec_res']} |")
        a(f"| 평균 VaR | {fmt(bt_result['avg_var'])}원 |")
        a()
        ex = bt_result['exceptions']
        if ex <= 4:
            a("> **평가**: 모형이 바젤 기준(250일 중 4회 이하)을 충족합니다. "
              "현행 VaR 모형을 유지하세요.")
        elif ex <= 9:
            a("> **평가**: 황색존 진입. 모형 파라미터 재검토 및 "
              "변동성 가정 타당성을 점검하세요.")
        else:
            a("> **평가**: 적색존 진입. 모형 교체 또는 자본 승수 증가가 필요합니다.")
    else:
        a("> *Back-testing 탭에서 실행 완료 후 리포트를 생성하세요.*")
    a(); a("---"); a()

    # ── 5. 시장별 기법 적합성 ─────────────────────────────────────────
    a("## 5. 시장별 기법 적합성"); a()
    a("| 시장/자산 | 추천 기법 | 이유 |")
    a("|----------|----------|------|")
    a("| KOSPI/KOSDAQ 개별주 | MJD + EWMA | 공시 중심 급등락, 변동성 클러스터링 |")
    a("| 한국 채권 (장내/장외) | EWMA Parametric (KRD) | 금리 민감도 선형 모델 적합 |")
    a("| KOSPI200 파생 | MJD + Delta-Gamma | 비선형 손익, 극단 시나리오 중요 |")
    a("| NYSE/NASDAQ 대형주 | Bootstrap + t-MC | Fat-tail 명확, 충분한 과거 데이터 |")
    a("| 미국 채권 ETF | EWMA Parametric | 듀레이션 기반 선형 근사 유효 |")
    a("| 포트폴리오 전체 | EWMA Marginal/Component VaR | 종목간 상관·기여도 파악 |")
    a(); a("---"); a()

    # ── 6. 실무 활용 및 한계 ─────────────────────────────────────────
    a("## 6. 실무 활용 가능성 및 한계점"); a()
    a("### 실무 활용")
    a("- **바젤 III 자본 요구량**: EWMA Parametric VaR 99%, 10일은 IMA 직접 활용 가능")
    a("- **한도 관리**: Component VaR로 종목별 리스크 기여도 모니터링")
    a("- **신규 편입 의사결정**: Incremental VaR로 VaR 증감 사전 평가")
    a("- **스트레스 보완**: MJD 점프 파라미터를 위기 수준으로 설정하여 극단 손실 추정")
    a()
    a("### 한계점"); a()
    a("| 기법 | 한계 |")
    a("|------|------|")
    a("| EWMA Parametric | 정규분포 가정 → Fat-tail 과소추정 |")
    a("| Merton JD | λ·μJ·σJ 추정 어려움, 점프 크기 정규분포 가정 |")
    a("| Block Bootstrap | 표본 외 사건 반영 불가, 블록 크기 선택 민감도 |")
    a("| t-분포 MC | 자유도 설정에 결과 민감 |")
    a("| 공통 | 과거 의존, 유동성·운용 리스크 미반영 |")
    a(); a("---"); a()

    # ── 7. 결론 및 권고 ───────────────────────────────────────────────
    a("## 7. 결론 및 권고사항"); a()
    a("1. **일상 모니터링**: EWMA Parametric VaR(99%, 10일)를 기준 지표로 사용")
    a("2. **개별 주식**: MJD 시뮬레이션 분기 1회 실시 → 점프 위험 명시적 평가")
    a("3. **모형 검증**: Historical Bootstrap 결과와 실제 손익 비교 → Back-testing 정기 수행")
    a("4. **파생상품 편입 시**: Delta-Gamma 정확도를 Full Revaluation과 비교")
    a("5. **시장 급변 시**: EWMA λ를 0.90~0.92로 낮춰 최근 충격 반영 속도 향상")
    a(); a("---"); a()
    a(f"*Portfolio VaR Calculator 자동 생성 | {now}*")

    return "\n".join(L)


def build_excel_report(pf_name, positions, usdkrw,
                        ewma_results=None, sim_results=None,
                        bt_result=None, mc_var_result=None):
    """
    포트폴리오 분석 결과를 Excel(.xlsx)로 생성
    반환: bytes (st.download_button에 직접 전달 가능)
    """
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side, numbers as xl_num)
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb  = Workbook()
    wb.remove(wb.active)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    # ── 공통 스타일 헬퍼 ────────────────────────────────────────────────
    C_DARK  = "1F3864"; C_MID = "2F5496"; C_LIGHT = "D6E4F0"
    C_RED   = "C00000"; C_GRN = "375623"; C_AMB   = "7F6000"
    C_LGRAY = "F2F2F2"; C_WHT = "FFFFFF"
    thin    = Side(style="thin", color="BFBFBF")
    bdr     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, row, cols, texts, bg=C_DARK, fg=C_WHT):
        for col, text in zip(cols, texts):
            c = ws.cell(row=row, column=col, value=text)
            c.font      = Font(bold=True, color=fg, size=10, name="Arial")
            c.fill      = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = bdr

    def cell(ws, row, col, value, bold=False, num_fmt=None,
             fg="000000", bg=None, align="right"):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(bold=bold, color=fg, name="Arial", size=10)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border    = bdr
        if bg: c.fill = PatternFill("solid", start_color=bg)
        if num_fmt: c.number_format = num_fmt
        return c

    NUM = "#,##0"; PCT = "0.00%"

    # ══ Sheet 1: 포트폴리오 구성 ════════════════════════════════════════
    ws1 = wb.create_sheet("포트폴리오 구성")
    ws1.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGH", [20,10,12,10,18,10,10,10]):
        ws1.column_dimensions[get_column_letter(ord(col)-64)].width = w
    ws1.merge_cells("A1:H1")
    ws1["A1"] = f"포트폴리오 구성 — {pf_name}  ({now})"
    ws1["A1"].font = Font(bold=True, color=C_WHT, size=13, name="Arial")
    ws1["A1"].fill = PatternFill("solid", start_color=C_DARK)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    hdr(ws1, 2, range(1,9),
        ["티커","유형","시장","수량","평가금액(KRW)","통화","비중(%)","단가"],
        bg=C_MID)
    total_mv = sum(
        p["price"]*p["qty"]*(usdkrw if p["ccy"]=="USD" else 1)
        for p in positions
    )
    for i, p in enumerate(positions):
        mv  = p["price"]*p["qty"]*(usdkrw if p["ccy"]=="USD" else 1)
        pct = mv/total_mv if total_mv else 0
        bg  = C_LGRAY if i%2==0 else C_WHT
        r   = i + 3
        cell(ws1,r,1,p["ticker"],    bold=True, align="left", bg=bg)
        cell(ws1,r,2,p["type"],      align="center", bg=bg)
        cell(ws1,r,3,MKT[p["market"]]["label"], align="center", bg=bg)
        cell(ws1,r,4,p["qty"],       num_fmt=NUM, bg=bg)
        cell(ws1,r,5,mv,             num_fmt=NUM, bg=bg)
        cell(ws1,r,6,p["ccy"],       align="center", bg=bg)
        cell(ws1,r,7,pct,            num_fmt=PCT, bg=bg)
        cell(ws1,r,8,p["price"],     num_fmt=NUM, bg=bg)
    # 합계 행
    r = len(positions) + 3
    cell(ws1,r,1,"합계", bold=True, align="left", bg=C_LIGHT, fg=C_DARK)
    for c in range(2,7): cell(ws1,r,c,"", bg=C_LIGHT)
    cell(ws1,r,5,total_mv, bold=True, num_fmt=NUM, bg=C_LIGHT, fg=C_RED)
    cell(ws1,r,7,1.0, bold=True, num_fmt=PCT, bg=C_LIGHT)

    # ══ Sheet 2: EWMA VaR ═══════════════════════════════════════════════
    ws2 = wb.create_sheet("EWMA VaR 분석")
    ws2.sheet_view.showGridLines = False
    for col_idx, w in enumerate([22,18,14,14,14,14,14,14], 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
    ws2.merge_cells("A1:H1")
    ws2["A1"] = "EWMA 기반 VaR 분석 (RiskMetrics λ=0.94)"
    ws2["A1"].font = Font(bold=True, color=C_WHT, size=13, name="Arial")
    ws2["A1"].fill = PatternFill("solid", start_color=C_DARK)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    if ewma_results and ewma_results.get("portfolio"):
        ws2.merge_cells("A2:H2")
        ws2["A2"] = "포트폴리오 VaR 요약"
        ws2["A2"].font = Font(bold=True, color=C_WHT, name="Arial", size=11)
        ws2["A2"].fill = PatternFill("solid", start_color=C_MID)
        ws2["A2"].alignment = Alignment(horizontal="center")
        hdr(ws2, 3, [1,2], ["지표","값"], bg=C_LIGHT, fg=C_DARK)
        for i, (k,v) in enumerate(ewma_results["portfolio"].items()):
            bg = C_LGRAY if i%2==0 else C_WHT
            cell(ws2, 4+i, 1, k, align="left", bg=bg)
            val = v if not isinstance(v,(int,float)) else round(float(v),0)
            cell(ws2, 4+i, 2, val, num_fmt=NUM if isinstance(v,(int,float)) else None,
                 bg=bg, fg=(C_RED if "VaR" in str(k) else "000000"))
        r2 = 4 + len(ewma_results["portfolio"]) + 1

        if ewma_results.get("individual"):
            ind = ewma_results["individual"]
            ws2.merge_cells(f"A{r2}:H{r2}")
            ws2[f"A{r2}"] = "종목별 리스크 기여도"
            ws2[f"A{r2}"].font = Font(bold=True, color=C_WHT, name="Arial", size=11)
            ws2[f"A{r2}"].fill = PatternFill("solid", start_color=C_MID)
            ws2[f"A{r2}"].alignment = Alignment(horizontal="center")
            r2 += 1
            if ind:
                cols_h = list(ind[0].keys())
                hdr(ws2, r2, range(1, len(cols_h)+1), cols_h, bg=C_LIGHT, fg=C_DARK)
                r2 += 1
                for i, row in enumerate(ind):
                    bg = C_LGRAY if i%2==0 else C_WHT
                    for j, v in enumerate(row.values()):
                        cell(ws2, r2, j+1, v, bg=bg,
                             align="right" if j>0 else "left")
                    r2 += 1
    else:
        ws2["A3"] = "EWMA VaR 탭에서 계산 후 리포트를 생성하세요."
        ws2["A3"].font = Font(italic=True, color="808080", name="Arial")

    # ══ Sheet 3: 시뮬레이션 결과 ════════════════════════════════════════
    ws3 = wb.create_sheet("시뮬레이션 결과")
    ws3.sheet_view.showGridLines = False
    for col_idx, w in enumerate([26,16,14,14,14,14,14], 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = w
    ws3.merge_cells("A1:G1")
    ws3["A1"] = "시뮬레이션 분석 결과 (MJD / Bootstrap / t-MC)"
    ws3["A1"].font = Font(bold=True, color=C_WHT, size=13, name="Arial")
    ws3["A1"].fill = PatternFill("solid", start_color=C_DARK)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")

    if sim_results and sim_results.get("metrics"):
        m_dict = sim_results["metrics"]
        conf_s = sim_results.get("conf", 0.99)
        if isinstance(conf_s, str):
            conf_s = float(conf_s.replace("%","")) / 100
        vk = f"VaR {int(conf_s*100)}%"; ck = f"CVaR {int(conf_s*100)}%"
        method_map = {"MJD":"Merton JD","Bootstrap":"Block Bootstrap","t-MC":"t-MC"}
        hdr(ws3, 2, range(1,8),
            ["기법","기대수익률","표준편차",f"VaR {int(conf_s*100)}%",
             f"CVaR {int(conf_s*100)}%","손실확률","기대 주가"],
            bg=C_MID)
        for i, (key, m) in enumerate(m_dict.items()):
            bg = C_LGRAY if i%2==0 else C_WHT
            lbl = method_map.get(key, key)
            cell(ws3, 3+i, 1, lbl, bold=True, align="left", bg=bg)
            cell(ws3, 3+i, 2, m.get("기대수익률(로그)",0)/100, num_fmt=PCT, bg=bg)
            cell(ws3, 3+i, 3, m.get("수익률 표준편차",0)/100, num_fmt=PCT, bg=bg)
            cell(ws3, 3+i, 4, m.get(vk,0)/100, num_fmt=PCT, fg=C_RED, bg=bg)
            cell(ws3, 3+i, 5, m.get(ck,0)/100, num_fmt=PCT, fg=C_AMB, bg=bg)
            cell(ws3, 3+i, 6, m.get("손실확률",0)/100, num_fmt=PCT, bg=bg)
            cell(ws3, 3+i, 7, m.get("기대 주가",0), num_fmt=NUM, bg=bg)
    else:
        ws3["A2"] = "시뮬레이션 탭에서 계산 후 리포트를 생성하세요."
        ws3["A2"].font = Font(italic=True, color="808080", name="Arial")

    # ══ Sheet 4: Back-testing ════════════════════════════════════════════
    ws4 = wb.create_sheet("Back-testing")
    ws4.sheet_view.showGridLines = False
    for col_idx, w in enumerate([28,20], 1):
        ws4.column_dimensions[get_column_letter(col_idx)].width = w
    ws4.merge_cells("A1:B1")
    ws4["A1"] = "VaR Back-testing 결과 (Basel II/III)"
    ws4["A1"].font = Font(bold=True, color=C_WHT, size=13, name="Arial")
    ws4["A1"].fill = PatternFill("solid", start_color=C_DARK)
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")

    if bt_result:
        hdr(ws4, 2, [1,2], ["지표","값"], bg=C_MID)
        bt_rows = [
            ("분석 대상",  bt_result["ticker"]),
            ("신뢰수준",   f"{int(bt_result['conf']*100)}%"),
            ("검증 기간",  f"{bt_result['window']}일"),
            ("VaR 초과 횟수", bt_result["exceptions"]),
            ("초과 비율",  f"{bt_result['exc_rate']:.2f}%"),
            ("Basel 판정", bt_result["zone"]),
            ("Kupiec LR",  f"{bt_result['lr_kupiec']:.3f}"),
            ("Kupiec p-value", f"{bt_result['p_kupiec']:.4f}"),
            ("Kupiec 결과", bt_result["kupiec_res"]),
            ("평균 VaR",   bt_result["avg_var"]),
        ]
        for i, (k, v) in enumerate(bt_rows):
            bg = C_LGRAY if i%2==0 else C_WHT
            cell(ws4, 3+i, 1, k, align="left", bg=bg)
            val = round(v,0) if isinstance(v,(int,float)) else v
            fg  = (C_RED if bt_result["exceptions"]>=10 else
                   C_AMB if bt_result["exceptions"]>=5 else C_GRN) if k=="VaR 초과 횟수" else "000000"
            cell(ws4, 3+i, 2, val,
                 num_fmt=NUM if isinstance(v,(int,float)) and not isinstance(v,bool) else None,
                 fg=fg, bg=bg)
    else:
        ws4["A2"] = "Back-testing 탭에서 계산 후 리포트를 생성하세요."
        ws4["A2"].font = Font(italic=True, color="808080", name="Arial")

    # ══ Sheet 5: Monte Carlo VaR ════════════════════════════════════════
    ws_mc = wb.create_sheet("Monte Carlo VaR")
    ws_mc.sheet_view.showGridLines = False
    for col_idx, w in enumerate([28, 20], 1):
        ws_mc.column_dimensions[get_column_letter(col_idx)].width = w
    ws_mc.merge_cells("A1:B1")
    ws_mc["A1"] = "Monte Carlo Simulation VaR"
    ws_mc["A1"].font      = Font(bold=True, color=C_WHT, size=13, name="Arial")
    ws_mc["A1"].fill      = PatternFill("solid", start_color=C_DARK)
    ws_mc["A1"].alignment = Alignment(horizontal="center", vertical="center")

    if mc_var_result:
        conf_mc  = mc_var_result.get("conf", 0.99)
        hold_mc  = mc_var_result.get("holding", 10)
        dist_mc  = mc_var_result.get("dist_type", "t-분포")
        n_mc     = mc_var_result.get("n_sims", 10000)
        hdr(ws_mc, 2, [1, 2], ["지표", "값"], bg=C_MID)
        mc_rows = [
            ("분포 유형",   dist_mc),
            ("시뮬레이션 횟수", f"{n_mc:,}회"),
            ("신뢰수준",    f"{int(conf_mc*100)}%"),
            ("보유기간",    f"{hold_mc}일"),
            (f"포트폴리오 MC VaR {int(conf_mc*100)}%",
             float(mc_var_result.get("pf_var", 0))),
            (f"포트폴리오 MC CVaR {int(conf_mc*100)}%",
             float(mc_var_result.get("pf_cvar", 0))),
            ("MC VaR%",    f"{mc_var_result.get('var_pct', 0)*100:.3f}%"),
        ]
        for i, (k, v) in enumerate(mc_rows):
            bg = C_LGRAY if i % 2 == 0 else C_WHT
            cell(ws_mc, 3+i, 1, k, align="left", bg=bg)
            if isinstance(v, float):
                cell(ws_mc, 3+i, 2, v, num_fmt=NUM, bg=bg, fg=C_RED)
            else:
                cell(ws_mc, 3+i, 2, str(v), align="right", bg=bg)
        # 종목별 MC VaR
        r_mc = 3 + len(mc_rows) + 1
        if mc_var_result.get("ind_rows"):
            ind_mc = mc_var_result["ind_rows"]
            ws_mc.merge_cells(f"A{r_mc}:B{r_mc}")
            ws_mc[f"A{r_mc}"] = "종목별 Monte Carlo VaR"
            ws_mc[f"A{r_mc}"].font = Font(bold=True, color=C_WHT, name="Arial", size=11)
            ws_mc[f"A{r_mc}"].fill = PatternFill("solid", start_color=C_MID)
            ws_mc[f"A{r_mc}"].alignment = Alignment(horizontal="center")
            r_mc += 1
            # 컬럼 폭 조정
            for ci2, w2 in enumerate(
                [20, 16, 14, 14, 12] + [12]*(len(ind_mc[0])-5), 1
            ):
                ws_mc.column_dimensions[get_column_letter(ci2)].width = w2
            if ind_mc:
                hdr(ws_mc, r_mc, range(1, len(ind_mc[0])+1),
                    list(ind_mc[0].keys()), bg=C_LIGHT, fg=C_DARK)
                r_mc += 1
                for ii, row in enumerate(ind_mc):
                    bg = C_LGRAY if ii%2==0 else C_WHT
                    for jj, v in enumerate(row.values()):
                        cell(ws_mc, r_mc, jj+1, v,
                             align="left" if jj==0 else "right", bg=bg)
                    r_mc += 1
    else:
        ws_mc["A2"] = "Parametric VaR 탭에서 Monte Carlo VaR를 실행 후 리포트를 생성하세요."
        ws_mc["A2"].font = Font(italic=True, color="808080", name="Arial")

    # ══ Sheet 6: 기법 가이드 ════════════════════════════════════════════
    ws5 = wb.create_sheet("기법 선택 가이드")
    ws5.sheet_view.showGridLines = False
    for col_idx, w in enumerate([20,28,30,28], 1):
        ws5.column_dimensions[get_column_letter(col_idx)].width = w
    ws5.merge_cells("A1:D1")
    ws5["A1"] = "분석 기법 선택 가이드"
    ws5["A1"].font = Font(bold=True, color=C_WHT, size=13, name="Arial")
    ws5["A1"].fill = PatternFill("solid", start_color=C_DARK)
    ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
    hdr(ws5, 2, [1,2,3,4], ["기법","핵심 강점","적합 자산","주요 한계"], bg=C_MID)
    guide = [
        ("EWMA Parametric VaR","변동성 클러스터링 반영, 바젤 직접 활용",
         "주식 포트폴리오, 채권","정규분포 가정, Fat-tail 과소추정"),
        ("Merton Jump Diffusion","급등락·불연속 충격 포착",
         "개별 주식, 옵션","파라미터 추정 어려움"),
        ("Block Bootstrap","분포 가정 없이 실제 패턴 재현",
         "지수·ETF","표본 외 사건 반영 불가"),
        ("t-분포 Monte Carlo","Fat-tail 반영, VaR 과소추정 방지",
         "포트폴리오 전체","자유도 설정 민감도"),
    ]
    for i, row in enumerate(guide):
        bg = C_LGRAY if i%2==0 else C_WHT
        for j, v in enumerate(row):
            cell(ws5, 3+i, j+1, v, align="left", bg=bg)

    # bytes 반환
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def sim_fetch_returns(ticker, days=504):
    """yfinance 일별 로그수익률 수집 (fallback: GBM 샘플)"""
    if YF_OK:
        try:
            end   = datetime.today()
            start = end - timedelta(days=int(days * 1.6))
            d = yf.download(ticker, start=start.strftime("%Y-%m-%d"),
                            end=end.strftime("%Y-%m-%d"),
                            progress=False, auto_adjust=True)
            if not d.empty and len(d) >= 60:
                prices = d["Close"].squeeze()
                rets   = np.log(prices / prices.shift(1)).dropna().values
                return rets[-days:], float(prices.iloc[-1]), True
        except Exception:
            pass
    # fallback
    rng   = np.random.default_rng(abs(hash(ticker)) % 9999)
    sigma = 0.22 / np.sqrt(252)
    rets  = rng.normal(-0.5*sigma**2, sigma, days)
    return rets, 50000.0, False


def sim_summary_stats(rets):
    """기초 통계량"""
    from scipy.stats import skew, kurtosis, jarque_bera
    mu    = float(np.mean(rets))
    sigma = float(np.std(rets, ddof=1))
    sk    = float(skew(rets))
    ku    = float(kurtosis(rets))        # excess kurtosis
    jb, jb_p = jarque_bera(rets)
    ann_ret   = mu * 252
    ann_vol   = sigma * np.sqrt(252)
    return {
        "일평균수익률": mu,
        "일변동성":     sigma,
        "연율수익률":   ann_ret,
        "연율변동성":   ann_vol,
        "왜도":         sk,
        "초과첨도":     ku,
        "JB통계량":     float(jb),
        "JB p-value":   float(jb_p),
    }


# ── Merton Jump Diffusion ────────────────────────────────────────────────────
def sim_mjd(S0, mu, sigma, lam, mu_j, sigma_j, T, dt, n_paths, seed=42):
    """
    Merton Jump Diffusion 경로 생성
    S0     : 현재 주가
    mu     : 드리프트 (연율)
    sigma  : 확산 변동성 (연율)
    lam    : 점프 강도 (λ, 연간 평균 점프 횟수)
    mu_j   : 점프 크기 평균 (로그 공간)
    sigma_j: 점프 크기 표준편차
    T      : 예측 기간 (년)
    dt     : 시간 간격 (1/252 = 1영업일)
    n_paths: 경로 수
    """
    rng    = np.random.default_rng(seed)
    steps  = int(T / dt)
    kappa  = np.exp(mu_j + 0.5 * sigma_j**2) - 1   # 점프의 기대값
    drift  = (mu - lam * kappa - 0.5 * sigma**2) * dt

    paths  = np.zeros((n_paths, steps + 1))
    paths[:, 0] = S0

    for t in range(steps):
        Z   = rng.standard_normal(n_paths)
        N   = rng.poisson(lam * dt, n_paths)            # 포아송 점프 횟수
        J   = rng.normal(mu_j, sigma_j, n_paths) * N    # 복합 점프 크기
        dS  = drift + sigma * np.sqrt(dt) * Z + J
        paths[:, t+1] = paths[:, t] * np.exp(dS)

    return paths


# ── Historical Block Bootstrap ────────────────────────────────────────────────
def sim_bootstrap(S0, hist_rets, T, dt, n_paths, block_size=10, seed=42):
    """
    Block Bootstrap 경로 생성
    block_size: 블록 단위 (영업일), 시계열 자기상관 보존
    """
    rng    = np.random.default_rng(seed)
    steps  = int(T / dt)
    n_hist = len(hist_rets)
    paths  = np.zeros((n_paths, steps + 1))
    paths[:, 0] = S0

    n_blocks = int(np.ceil(steps / block_size))
    max_start = n_hist - block_size

    for p in range(n_paths):
        sampled = []
        for _ in range(n_blocks):
            start = rng.integers(0, max(max_start, 1))
            sampled.extend(hist_rets[start: start + block_size].tolist())
        sampled = sampled[:steps]
        price   = S0
        for t, r in enumerate(sampled):
            price = price * np.exp(r)
            paths[p, t + 1] = price

    return paths


# ── t-분포 Monte Carlo ────────────────────────────────────────────────────────
def sim_t_mc(S0, mu, sigma, df, T, dt, n_paths, seed=42):
    """
    t-분포 기반 Monte Carlo
    df  : 자유도 (낮을수록 두꺼운 꼬리, 통상 3~8)
    """
    rng   = np.random.default_rng(seed)
    steps = int(T / dt)
    scale = sigma * np.sqrt(dt) * np.sqrt((df - 2) / df) if df > 2 else sigma * np.sqrt(dt)
    drift = (mu - 0.5 * sigma**2) * dt

    paths = np.zeros((n_paths, steps + 1))
    paths[:, 0] = S0

    t_draws = stats.t.rvs(df=df, size=(n_paths, steps),
                          random_state=rng.integers(0, 2**31))
    for t in range(steps):
        dS = drift + scale * t_draws[:, t]
        paths[:, t + 1] = paths[:, t] * np.exp(dS)

    return paths


def sim_metrics(paths, S0, conf_level=0.99):
    """시뮬레이션 경로로부터 성과 지표 산출"""
    terminal     = paths[:, -1]
    log_rets     = np.log(terminal / S0)
    mean_ret     = float(np.mean(log_rets))
    std_ret      = float(np.std(log_rets, ddof=1))
    var_pct      = float(np.percentile(log_rets, (1 - conf_level) * 100))
    cvar_pct     = float(log_rets[log_rets <= var_pct].mean()) if (log_rets <= var_pct).any() else var_pct
    loss_prob    = float((terminal < S0).mean())
    p5, p50, p95 = np.percentile(terminal, [5, 50, 95])
    return {
        "기대수익률(로그)": mean_ret,
        "수익률 표준편차":  std_ret,
        f"VaR {int(conf_level*100)}%": var_pct,
        f"CVaR {int(conf_level*100)}%": cvar_pct,
        "손실확률":         loss_prob,
        "5% 분위 주가":     float(p5),
        "중앙값 주가":      float(p50),
        "95% 분위 주가":    float(p95),
        "기대 주가":        float(np.mean(terminal)),
    }


# ── 경로 팬 차트 생성 ─────────────────────────────────────────────────────────
def make_fan_chart(paths, dt, title, n_show=200):
    """
    경로 팬 차트
    - 최선(상위 5%): 초록
    - 중간(중앙값): 빨강
    - 최악(하위 5%): 검정
    - 나머지 샘플: 반투명 파랑
    """
    steps      = paths.shape[1]
    time_axis  = np.arange(steps) * dt * 252    # 영업일 단위
    terminal   = paths[:, -1]
    idx_best   = np.argmax(terminal)
    idx_worst  = np.argmin(terminal)
    idx_median = np.argmin(np.abs(terminal - np.median(terminal)))

    # 샘플 경로 (얇게)
    sample_idx = np.random.choice(
        len(paths), size=min(n_show, len(paths)), replace=False
    )

    fig = go.Figure()
    for i in sample_idx:
        if i in (idx_best, idx_worst, idx_median):
            continue
        fig.add_trace(go.Scatter(
            x=time_axis, y=paths[i],
            mode="lines",
            line=dict(color="rgba(55,138,221,0.08)", width=0.8),
            showlegend=False,
            hoverinfo="skip",
        ))

    # 최악 (검정)
    fig.add_trace(go.Scatter(
        x=time_axis, y=paths[idx_worst],
        mode="lines", name="최악 시나리오",
        line=dict(color="#1a1a1a", width=2.5, dash="dot"),
    ))
    # 중간 (빨강)
    fig.add_trace(go.Scatter(
        x=time_axis, y=paths[idx_median],
        mode="lines", name="중앙값 시나리오",
        line=dict(color="#A32D2D", width=2.5),
    ))
    # 최선 (초록)
    fig.add_trace(go.Scatter(
        x=time_axis, y=paths[idx_best],
        mode="lines", name="최선 시나리오",
        line=dict(color="#3B6D11", width=2.5, dash="dash"),
    ))

    fig.update_layout(
        title=title,
        xaxis_title="영업일",
        yaxis_title="주가",
        height=380,
        legend=dict(orientation="h", y=-0.18),
        margin=dict(t=50, b=70, l=60, r=20),
    )
    return fig


# ── 만기 분포 히스토그램 ──────────────────────────────────────────────────────
def make_terminal_hist(paths, S0, method_name, conf=0.99):
    terminal = paths[:, -1]
    var_price = np.percentile(terminal, (1-conf)*100)
    fig = go.Figure()
    fig.add_trace(go.Histogram(
        x=terminal,
        nbinsx=60,
        marker_color="rgba(55,138,221,0.6)",
        name="만기 주가 분포",
    ))
    fig.add_vline(x=S0, line_dash="dash", line_color="#185FA5",
                  annotation_text=f"현재가 {S0:,.0f}",
                  annotation_position="top right")
    fig.add_vline(x=var_price, line_dash="dash", line_color="#A32D2D",
                  annotation_text=f"VaR99% {var_price:,.0f}",
                  annotation_position="top left")
    fig.update_layout(
        title=f"{method_name} — 만기 주가 분포",
        xaxis_title="만기 주가",
        yaxis_title="빈도",
        height=300,
        showlegend=False,
        margin=dict(t=50, b=40, l=50, r=20),
    )
    return fig


# ══════════════════════════════════════════════════════════════════════════════
# 세션 초기화
# ══════════════════════════════════════════════════════════════════════════════
if "pfs" not in st.session_state:
    st.session_state.pfs = [{"name": "My Portfolio", "positions": [
        {"ticker":"005930.KS","type":"Stock","market":"KR-KOSPI",
         "qty":100,"price":75000,"ccy":"KRW","mult":1},
        # 채권: 수량=10,000매, 액면가=10,000원/매 → 투자원금 1억원
        {"ticker":"KR103501GA96","type":"Bond","market":"KR-BOND",
         "qty":10000,"price":10000,"ccy":"KRW","mult":1,
         "coupon":0.035,"maturity":3.0,"cf_freq":2,"face_value":10000},
        # 옵션: KOSPI200 pt × 수량 × 250,000원/pt
        {"ticker":"K200C-255-202506","type":"Option","market":"KR-DERIV",
         "qty":10,"price":257.0,"ccy":"KRW","mult":250000,
         "strike":255,"callput":"C","iv":0.18,"mat":0.16,"rf":0.035},
        {"ticker":"AAPL","type":"Stock","market":"US-NASDAQ",
         "qty":50,"price":185,"ccy":"USD","mult":1},
    ]}]
if "cpf"   not in st.session_state: st.session_state.cpf   = 0
if "vr"    not in st.session_state: st.session_state.vr    = {}
if "sel_tk" not in st.session_state: st.session_state.sel_tk = ""
if "sel_pr" not in st.session_state: st.session_state.sel_pr = 0.0

# ══════════════════════════════════════════════════════════════════════════════
# 사이드바
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("### Global Settings")

    # ── 실시간 환율 ────────────────────────────────────────────────────────
    rt_fx, fx_live = get_usdkrw()
    if fx_live:
        st.success(f"USD/KRW (실시간): {rt_fx:,.2f}")
    else:
        st.warning("yfinance 환율 미연결")
    usdkrw = st.number_input("USD/KRW 환율", 800, 2000, int(rt_fx), 1)

    # ── 금리 커브 ──────────────────────────────────────────────────────────
    ykr, kr_live = get_kr_curve()
    yus, us_live = get_us_curve()
    st.caption(
        f"KRW 커브: {'실시간' if kr_live else '기본값 3.5%대'} | "
        f"USD 커브: {'실시간' if us_live else '기본값 4~5%'}"
    )

    st.markdown("---")
    st.markdown("### VaR 파라미터")
    conf    = st.selectbox("신뢰수준", ["99%","95%"])
    z_val   = Z99 if conf == "99%" else Z95
    holding = st.number_input("보유기간 (영업일)", 1, 250, 1)
    sqrt_h  = np.sqrt(holding)

    st.markdown("---")
    st.markdown("### 자산군간 상관관계")
    c_sb = st.slider("주식 ↔ 채권",    -1.0, 1.0, 0.25, 0.05)
    c_so = st.slider("주식 ↔ 옵션/선물",-1.0, 1.0, 0.45, 0.05)
    c_bo = st.slider("채권 ↔ 옵션/선물",-1.0, 1.0, 0.15, 0.05)

    st.markdown("---")
    st.markdown("### 포트폴리오")
    pf_names = [p["name"] for p in st.session_state.pfs]
    sel = st.selectbox("활성 포트폴리오", pf_names,
                       index=st.session_state.cpf)
    st.session_state.cpf = pf_names.index(sel)

    if st.button("복사본 생성 (My Portfolio N)"):
        import copy
        n = len(st.session_state.pfs)
        st.session_state.pfs.append({
            "name": f"My Portfolio {n}",
            "positions": copy.deepcopy(
                st.session_state.pfs[st.session_state.cpf]["positions"])
        })
        st.session_state.cpf = n
        st.rerun()

    if len(st.session_state.pfs) > 1:
        if st.button("현재 포트폴리오 삭제", type="secondary"):
            st.session_state.pfs.pop(st.session_state.cpf)
            st.session_state.cpf = 0
            st.rerun()

pf        = st.session_state.pfs[st.session_state.cpf]
positions = pf["positions"]

st.markdown(
    f"## Portfolio VaR Calculator  "
    f"<span style='font-size:14px;color:#888;font-weight:400'>"
    f"— {pf['name']}</span>",
    unsafe_allow_html=True
)

# ══════════════════════════════════════════════════════════════════════════════
# 탭 레이아웃
# ══════════════════════════════════════════════════════════════════════════════
tab_pf, tab_vo, tab_res, tab_cmp, tab_st, tab_sim, tab_ewma, tab_bt, tab_report = st.tabs([
    "📁 포트폴리오", "📈 VaR/Exposure", "📋 결과 요약",
    "⚖️ 포트폴리오 비교", "🔥 스트레스 테스트",
    "🔬 시뮬레이션 분석", "📐 Parametric VaR", "🔎 Back-testing", "📄 분석 리포트",
])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — 포트폴리오 구성
# ══════════════════════════════════════════════════════════════════════════════
with tab_pf:
    st.markdown("#### 포트폴리오 구성")

    # ── 파일 업로드 섹션 ──────────────────────────────────────────────────
    with st.expander("📂 파일로 포트폴리오 불러오기 (Excel / CSV)", expanded=False):
        st.markdown(
            '<div class="ib">'
            '포지션 정보가 담긴 Excel(.xlsx) 또는 CSV 파일을 업로드하면 '
            '포트폴리오를 자동으로 구성합니다.<br>'
            '<b>필수 컬럼</b>: ticker, type, qty, price '
            '&nbsp;|&nbsp; '
            '<b>선택 컬럼</b>: ccy, market, coupon, maturity, cf_freq, '
            'face_value, strike, callput, iv, mat, rf, mult'
            '</div>', unsafe_allow_html=True
        )

        # 템플릿 다운로드
        template_data = (
            "ticker,type,qty,price,ccy,market,coupon,maturity,cf_freq,"
            "face_value,strike,callput,iv,mat,rf,mult\n"
            "005930.KS,Stock,100,75000,KRW,KR-KOSPI,,,,,,,,,,\n"
            "KR103501GA96,Bond,10000,10000,KRW,KR-BOND,0.035,3.0,2,10000,,,,,, \n"
            "K200C-255-202506,Option,10,257,KRW,KR-DERIV,,,,, 255,C,0.18,0.16,0.035,250000\n"
            "AAPL,Stock,50,185,USD,US-NASDAQ,,,,,,,,,,\n"
        )
        st.download_button(
            "📥 입력 템플릿 다운로드 (.csv)",
            data=template_data.encode("utf-8"),
            file_name="portfolio_template.csv",
            mime="text/csv",
            key="dl_template",
        )

        uploaded_pf = st.file_uploader(
            "포트폴리오 파일 선택",
            type=["csv","xlsx","xls"],
            key="upload_pf_file",
        )
        uf_col1, uf_col2 = st.columns(2)
        with uf_col1:
            upload_mode = st.radio(
                "업로드 방식",
                ["기존 포트폴리오에 추가", "기존 포트폴리오 교체"],
                key="upload_mode",
                horizontal=True,
            )
        with uf_col2:
            upload_pf_target = st.selectbox(
                "대상 포트폴리오",
                [p["name"] for p in st.session_state.pfs],
                index=st.session_state.cpf,
                key="upload_pf_target",
            )

        if uploaded_pf is not None:
            if st.button("포트폴리오 불러오기 실행", type="primary",
                         key="do_upload"):
                parsed, err = parse_portfolio_file(uploaded_pf)
                if err:
                    st.error(f"오류: {err}")
                else:
                    target_idx = [p["name"] for p in st.session_state.pfs].index(
                        upload_pf_target
                    )
                    if upload_mode == "기존 포트폴리오 교체":
                        st.session_state.pfs[target_idx]["positions"] = parsed
                        st.session_state.cpf = target_idx
                        st.success(
                            f"✅ {upload_pf_target}을 {len(parsed)}개 포지션으로 교체했습니다."
                        )
                    else:
                        st.session_state.pfs[target_idx]["positions"].extend(parsed)
                        st.session_state.cpf = target_idx
                        st.success(
                            f"✅ {upload_pf_target}에 {len(parsed)}개 포지션을 추가했습니다."
                        )
                    st.rerun()

    st.markdown("#### 종목 직접 추가")

    # 시장 선택 라디오
    mkt_keys   = list(MKT.keys())
    mkt_labels = [MKT[k]["label"] for k in mkt_keys]
    mkt_sel_idx = st.radio(
        "시장 선택", range(len(mkt_keys)),
        format_func=lambda i: mkt_labels[i],
        horizontal=True, label_visibility="collapsed", key="mkt_radio"
    )
    mkt_sel = mkt_keys[mkt_sel_idx]
    mi      = MKT[mkt_sel]
    st.markdown(f'<div class="ib">{mi["info"]}</div>', unsafe_allow_html=True)

    # ── 종목 리스트 표시 ──────────────────────────────────────────────────
    sl     = STOCKS.get(mkt_sel, [])
    search = st.text_input(
        "종목 검색 (코드·이름·섹터)", "",
        placeholder="삼성, AAPL, 반도체 등 입력", key="search_q"
    )
    if search:
        q_lower  = search.lower()
        filtered = [s for s in sl if any(q_lower in x.lower() for x in s)]
    else:
        filtered = sl

    st.markdown(f"**{mi['label']} 종목 목록 ({len(filtered)}개)**")

    if filtered:
        # 셀렉트박스 옵션: "코드 | 이름 (섹터)" 형식
        select_opts = ["-- 종목 선택 --"] + [
            f"{code}  |  {name}  ({sector})"
            for code, name, sector in filtered
        ]
        # 현재 선택된 종목이 있으면 해당 인덱스로 초기화
        cur_idx = 0
        if st.session_state.sel_tk:
            for i, (code, name, sector) in enumerate(filtered):
                if code == st.session_state.sel_tk:
                    cur_idx = i + 1
                    break

        chosen = st.selectbox(
            "종목을 선택하세요 (스크롤 또는 검색)",
            select_opts,
            index=cur_idx,
            key=f"stock_sel_{mkt_sel}",
            help="검색창에 입력한 키워드로 이미 필터링된 목록입니다.",
        )

        if chosen != "-- 종목 선택 --":
            chosen_code = chosen.split("|")[0].strip()
            if chosen_code != st.session_state.sel_tk:
                st.session_state.sel_tk = chosen_code
                live_p = get_price(chosen_code)
                st.session_state.sel_pr = live_p if live_p else 0.0
                st.rerun()
    else:
        st.caption("검색 결과가 없습니다. 검색어를 변경해 보세요.")

    if st.session_state.sel_tk:
        pr_str = (f" | 현재가: {st.session_state.sel_pr:,.2f}"
                  if st.session_state.sel_pr else " | 가격을 직접 입력하세요")
        st.success(f"선택: **{st.session_state.sel_tk}**{pr_str}")

    st.markdown("---")

    # ── 입력 폼 ───────────────────────────────────────────────────────────
    is_bond  = mkt_sel in ("KR-BOND", "US-BOND")
    is_deriv = mkt_sel == "KR-DERIV"

    with st.form("add_pos", clear_on_submit=True):
        c1, c2, c3, c4, c5 = st.columns([2.5, 1.2, 1.2, 1.5, 1])
        with c1:
            ticker = st.text_input("티커", value=st.session_state.sel_tk)
        with c2:
            type_default = mi["type"]
            atype = st.selectbox(
                "유형", ["Stock","Bond","Option","Future"],
                index=["Stock","Bond","Option","Future"].index(type_default)
            )
        with c3:
            qty = st.number_input("수량", min_value=0.0, value=1.0, step=1.0)
        with c4:
            def_pr = st.session_state.sel_pr if st.session_state.sel_pr else 0.0
            price  = st.number_input(
                "단가", min_value=0.0, value=def_pr, step=100.0,
                help="채권: 입력 불필요(커브 자동계산) | 옵션: 기초자산 현재가"
            )
        with c5:
            ccy = st.selectbox("통화", ["KRW","USD"],
                               index=0 if mi["ccy"]=="KRW" else 1)

        extra = {}
        if is_bond:
            st.markdown("**채권 상세 정보**")
            b1, b2, b3, b4 = st.columns(4)
            with b1:
                extra["face_value"] = st.number_input(
                    "액면가(원/매)", 1000, 1000000, 10000, 1000,
                    help="국채 표준: 10,000원/매. 수량×액면가=투자원금"
                )
            with b2:
                extra["coupon"] = st.number_input(
                    "표면이율", 0.0, 0.2, 0.035, 0.001, format="%.3f"
                )
            with b3:
                extra["maturity"] = st.number_input(
                    "잔존만기(년)", 0.1, 30.0, 3.0, 0.5
                )
            with b4:
                extra["cf_freq"] = int(st.number_input("이표횟수/년", 1, 4, 2))
            fv_disp = extra.get("face_value", 10000)
            invest  = int(qty) * fv_disp
            st.markdown(
                f'<div class="ib">'
                f'수량 {int(qty):,}매 × 액면가 {fv_disp:,}원 = 투자원금 {fmt(invest)}원. '
                f'채권 현재가는 KRW 수익률 커브(3년={ykr[4]*100:.2f}%)로 자동 계산됩니다.'
                f'</div>', unsafe_allow_html=True
            )

        if is_deriv:
            st.markdown("**파생상품 상세 정보**")
            d1, d2, d3, d4, d5, d6 = st.columns(6)
            with d1: extra["strike"]  = st.number_input("행사가", 0.0, value=0.0)
            with d2: extra["callput"] = st.selectbox("C/P", ["C","P"])
            with d3: extra["mat"]     = st.number_input("만기(년)", 0.0, 2.0, 0.25, 0.01)
            with d4: extra["iv"]      = st.number_input("내재변동성", 0.0, 1.0, 0.18, 0.01)
            with d5:
                extra["mult"] = st.number_input(
                    "계약승수", 1.0, 10_000_000.0, 250000.0,
                    help="KOSPI200 옵션: 250,000원/pt"
                )
            with d6: extra["rf"] = st.number_input("무위험금리", 0.0, 0.1, 0.035, 0.001)
            st.markdown(
                '<div class="ib">'
                'KOSPI200 옵션 평가금액 = B-S 이론가(pt) × 수량 × 250,000원/pt. '
                'KOSPI200 선물 승수: 500,000원/pt'
                '</div>', unsafe_allow_html=True
            )

        ok = st.form_submit_button("+ 포트폴리오에 추가", type="primary",
                                    use_container_width=True)
        if ok:
            if not ticker:
                st.error("티커를 입력하거나 위 목록에서 종목을 선택하세요.")
            elif qty <= 0:
                st.error("수량을 입력하세요.")
            elif price <= 0 and not is_bond:
                st.error("단가를 입력하세요.")
            else:
                pos = {
                    "ticker": ticker.strip().upper(),
                    "type": atype, "market": mkt_sel,
                    "qty": float(qty), "price": float(price),
                    "ccy": ccy, "mult": extra.get("mult", 1.0),
                }
                pos.update({k: v for k, v in extra.items() if k != "mult"})
                positions.append(pos)
                st.session_state.sel_tk = ""
                st.session_state.sel_pr = 0.0
                st.success(f"{ticker} 추가 완료")
                st.rerun()

    # ── 현재 포트폴리오 테이블 ────────────────────────────────────────────
    st.markdown("---")
    st.markdown(f"#### {pf['name']} 현재 구성")

    # ── 실시간 현재가 일괄 갱신 ──────────────────────────────────────────
    rf_col1, rf_col2 = st.columns([3, 1])
    with rf_col2:
        if st.button("실시간 현재가 갱신", type="secondary",
                     use_container_width=True,
                     help="yfinance에서 주식 현재가를 일괄 조회합니다 (채권·옵션 제외)"):
            with st.spinner("현재가 조회 중..."):
                n_ok, fails = refresh_portfolio_prices(positions)
            if n_ok > 0:
                st.success(f"{n_ok}개 종목 현재가 갱신 완료"
                           + (f" | 실패: {', '.join(fails)}" if fails else ""))
            else:
                st.warning("현재가 조회 실패 (yfinance 미연결 또는 해당 없음)")
            st.rerun()

    if not positions:
        st.info("종목을 추가하세요.")
    else:
        rows      = []
        total_mv  = 0.0
        for p in positions:
            fx = usdkrw if p["ccy"] == "USD" else 1.0
            if p["type"] == "Bond":
                bvr = bond_var_calc(
                    p.get("face_value", 10000), p.get("coupon", 0.04),
                    p.get("maturity", 3.0), int(p.get("cf_freq", 2)), ykr
                )
                mv       = bvr["pv"] * p["qty"] * fx
                pr_disp  = f'{bvr["pv"]:,.0f}원/매 (평가가)'
            elif p["type"] in ("Option","Future"):
                g       = bs_greeks(p["price"],
                                    p.get("strike", p["price"]*1.05),
                                    p.get("mat", 0.25),
                                    p.get("rf", 0.035),
                                    p.get("iv", 0.18),
                                    p.get("callput", "C"))
                mv      = g["price"] * p["qty"] * p.get("mult", 1) * fx
                pr_disp = f'{g["price"]:,.2f}pt (이론가) Δ={g["delta"]:.3f}'
            else:
                mv      = p["price"] * p["qty"] * fx
                sym     = "$" if p["ccy"]=="USD" else "₩"
                pr_disp = f'{sym}{p["price"]:,.0f}'
            total_mv += mv
            rows.append({
                "티커":       p["ticker"],
                "유형":       p["type"],
                "시장":       MKT[p["market"]]["label"],
                "수량":       f'{p["qty"]:,.0f}',
                "단가":       pr_disp,
                "평가금액(KRW)": f'{fmt(mv)}원',
                "통화":       p["ccy"],
            })

        m1, m2, m3 = st.columns(3)
        with m1: st.metric("총 평가금액", f"{fmt(total_mv)}원")
        with m2: st.metric("종목 수",    f"{len(positions)}개")
        with m3: st.metric("USD/KRW",   f"{usdkrw:,}")

        st.dataframe(pd.DataFrame(rows), use_container_width=True,
                     height=min(420, 60+len(rows)*36))

        dc1, dc2 = st.columns([3, 1])
        with dc1:
            di = st.number_input("삭제 행 번호 (0부터)",
                                  0, max(0, len(positions)-1), 0, 1)
        with dc2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("삭제"):
                positions.pop(int(di))
                st.rerun()
        if st.button("전체 초기화", type="secondary"):
            pf["positions"].clear()
            st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — VaR/Exposure 추이
# ══════════════════════════════════════════════════════════════════════════════
with tab_vo:
    st.markdown("#### VaR / Exposure 추이")
    st.markdown(
        '<div class="ib">'
        'X축: Exposure (평가금액, 백만원) — Y축: Parametric VaR 99% (백만원).<br>'
        '관측 시작시점 VaR 계산에 최소 253일 가격 데이터 필요.'
        ' 총 2년(504영업일) 데이터 기준, 후반 1년(252일)을 표시합니다.'
        '</div>', unsafe_allow_html=True
    )

    if not positions:
        st.warning("포트폴리오를 먼저 구성하세요.")
    else:
        vod = gen_vo_data(positions, usdkrw)
        if vod:
            ds = [d.strftime("%m/%d") for d in vod["dates"]]
            st.caption(
                f"표시 기간: {vod['dates'][0]:%Y-%m-%d} ~ "
                f"{vod['dates'][-1]:%Y-%m-%d} ({len(ds)}일)"
            )

            x0, y0 = vod["exp"][0],  vod["var"][0]    # 분석 시작 시점
            x1, y1 = vod["exp"][-1], vod["var"][-1]   # 최신 시점

            fig_sc = go.Figure()

            # ── 전체 산점 ─────────────────────────────────────────────────
            fig_sc.add_trace(go.Scatter(
                x=vod["exp"], y=vod["var"], mode="markers",
                marker=dict(color="rgba(55,138,221,0.45)", size=5),
                name="VaR vs Exposure",
                hovertemplate="Exposure: %{x:.1f}M | VaR: %{y:.1f}M<extra></extra>",
            ))

            # ── 시작 시점 (초록 원) ────────────────────────────────────────
            fig_sc.add_trace(go.Scatter(
                x=[x0], y=[y0], mode="markers+text",
                marker=dict(color="#3B6D11", size=14, symbol="circle",
                            line=dict(color="#ffffff", width=2)),
                text=["시작"], textposition="top center",
                textfont=dict(size=11, color="#3B6D11"),
                name="분석 시작 시점",
                hovertemplate=f"시작: Exp {x0:.1f}M | VaR {y0:.1f}M<extra></extra>",
            ))

            # ── 최신 시점 (빨간 삼각형) ────────────────────────────────────
            fig_sc.add_trace(go.Scatter(
                x=[x1], y=[y1], mode="markers+text",
                marker=dict(color="#A32D2D", size=14, symbol="triangle-up",
                            line=dict(color="#ffffff", width=2)),
                text=["최신"], textposition="top center",
                textfont=dict(size=11, color="#A32D2D"),
                name="최신 시점",
                hovertemplate=f"최신: Exp {x1:.1f}M | VaR {y1:.1f}M<extra></extra>",
            ))

            # ── 시작→최신 화살표 (annotation) ─────────────────────────────
            fig_sc.add_annotation(
                x=x1, y=y1,
                ax=x0, ay=y0,
                xref="x", yref="y",
                axref="x", ayref="y",
                showarrow=True,
                arrowhead=3,
                arrowsize=1.4,
                arrowwidth=2.0,
                arrowcolor="#854F0B",
            )
            # 화살표 중간에 방향 레이블
            mid_x = (x0 + x1) / 2
            mid_y = (y0 + y1) / 2
            dx = x1 - x0; dy = y1 - y0
            dist = (dx**2 + dy**2) ** 0.5
            direction = (
                "VaR↑ Exp↑" if dx > 0 and dy > 0 else
                "VaR↓ Exp↓" if dx < 0 and dy < 0 else
                "VaR↑ Exp↓" if dx < 0 and dy > 0 else
                "VaR↓ Exp↑"
            )
            pct_chg_exp = (x1 - x0) / x0 * 100 if x0 else 0
            pct_chg_var = (y1 - y0) / y0 * 100 if y0 else 0
            fig_sc.add_annotation(
                x=mid_x, y=mid_y,
                text=(f"<b>{direction}</b><br>"
                      f"Exp {pct_chg_exp:+.1f}%<br>"
                      f"VaR {pct_chg_var:+.1f}%"),
                showarrow=False,
                font=dict(size=10, color="#854F0B"),
                bgcolor="rgba(250,238,218,0.85)",
                bordercolor="#854F0B",
                borderwidth=1,
                borderpad=4,
            )

            fig_sc.update_layout(
                title="Exposure vs VaR 99% 산점도 (최근 1년)",
                xaxis_title="Exposure (백만원)",
                yaxis_title="VaR 99% (백만원)",
                height=420,
                legend=dict(orientation="h", y=-0.15),
                margin=dict(t=50, b=70, l=60, r=20),
            )
            st.plotly_chart(fig_sc, use_container_width=True)

            # 시작→최신 변화 요약
            chg_col1, chg_col2, chg_col3 = st.columns(3)
            with chg_col1:
                st.metric("Exposure 변화",
                          f"{x1:.1f}M",
                          delta=f"{pct_chg_exp:+.1f}% (시작 {x0:.1f}M)")
            with chg_col2:
                st.metric("VaR 99% 변화",
                          f"{y1:.1f}M",
                          delta=f"{pct_chg_var:+.1f}% (시작 {y0:.1f}M)")
            with chg_col3:
                ratio_start = y0 / x0 * 100 if x0 else 0
                ratio_end   = y1 / x1 * 100 if x1 else 0
                st.metric("VaR/Exposure 변화",
                          f"{ratio_end:.2f}%",
                          delta=f"{ratio_end - ratio_start:+.2f}%p")

            ce, cv = st.columns(2)
            with ce:
                fe = go.Figure(go.Scatter(
                    x=ds, y=vod["exp"], mode="lines",
                    line=dict(color="#185FA5", width=1.5),
                    fill="tozeroy", fillcolor="rgba(55,138,221,0.08)",
                ))
                fe.update_layout(title="Exposure 추이 (백만원)", height=220,
                                  margin=dict(t=40,b=30,l=50,r=10))
                st.plotly_chart(fe, use_container_width=True)
            with cv:
                fv2 = go.Figure(go.Scatter(
                    x=ds, y=vod["var"], mode="lines",
                    line=dict(color="#A32D2D", width=1.5),
                    fill="tozeroy", fillcolor="rgba(162,45,45,0.08)",
                ))
                fv2.update_layout(title="VaR 99% 추이 (백만원)", height=220,
                                   margin=dict(t=40,b=30,l=50,r=10))
                st.plotly_chart(fv2, use_container_width=True)

            ea, va = np.array(vod["exp"]), np.array(vod["var"])
            s1, s2, s3, s4 = st.columns(4)
            with s1: st.metric("평균 Exposure", f"{ea.mean():.1f}M")
            with s2: st.metric("평균 VaR 99%",  f"{va.mean():.1f}M")
            with s3: st.metric("최대 VaR 99%",  f"{va.max():.1f}M")
            with s4: st.metric("VaR/Exposure",  f"{(va/ea).mean()*100:.2f}%")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — VaR 계산
# ══════════════════════════════════════════════════════════════════════════════
with tab_res:
    st.markdown("#### VaR 결과 요약")
    r = st.session_state.vr.get(st.session_state.cpf)
    if not r:
        st.info("'VaR 계산' 탭에서 먼저 계산을 실행하세요.")
    else:
        st.caption(
            f"{r['ts']:%Y-%m-%d %H:%M:%S} | {conf} | "
            f"보유 {holding}일 | USD/KRW {usdkrw:,}"
        )
        for col, label, val, cls in zip(
            st.columns(4),
            ["Portfolio VaR 99%","Portfolio VaR 95%","분산 효과","총 평가금액"],
            [r["pv99"], r["pv95"], r["divers"], r["total"]],
            ["red","amb","grn","blu"]
        ):
            with col:
                st.markdown(
                    f'<div class="mc"><div class="ml">{label}</div>'
                    f'<div class="mv {cls}">{fmt(val)}원</div></div>',
                    unsafe_allow_html=True
                )

        sv99 = round(r["sv"]["pv99"]/1e6, 1)
        bv99 = round(r["bv"]["pv99"]/1e6, 1)
        ov99 = round(r["ov"]["pv99"]/1e6, 1)
        c1, c2 = st.columns(2)
        with c1:
            fb = go.Figure(go.Bar(
                x=["주식","채권","옵션/선물"], y=[sv99,bv99,ov99],
                marker_color=["#378ADD","#3B6D11","#EF9F27"],
                text=[f"{v:.1f}M" for v in [sv99,bv99,ov99]],
                textposition="outside",
            ))
            fb.update_layout(title="자산군별 VaR 99% (백만원)", height=300,
                              showlegend=False, margin=dict(t=50,b=30,l=30,r=20))
            st.plotly_chart(fb, use_container_width=True)
        with c2:
            fd = go.Figure(go.Pie(
                labels=["주식","채권","옵션/선물"],
                values=[sv99,bv99,ov99], hole=0.52,
                marker_colors=["#378ADD","#3B6D11","#EF9F27"],
                textinfo="label+percent",
            ))
            fd.update_layout(title="Component VaR 구성", height=300,
                              showlegend=False, margin=dict(t=50,b=30,l=20,r=20))
            st.plotly_chart(fd, use_container_width=True)

        rows = []
        for pr in r["pos_results"]:
            p = pr["pos"]; mv = pr["mv"]
            rows.append({
                "티커":     p["ticker"],
                "유형":     p["type"],
                "시장":     MKT[p["market"]]["label"],
                "통화":     p["ccy"],
                "평가금액": fmt(mv)+"원",
                "P-VaR99%": fmt(pr["pv99"])+"원",
                "H-VaR99%": fmt(pr["hv99"])+"원",
                "M-VaR99%": fmt(pr["mv99"])+"원",
                "VaR%":     f'{pr["pv99"]/mv*100:.2f}%' if mv>0 else "0%",
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True,
                     height=min(420, 60+len(rows)*36))

# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 — 포트폴리오 비교
# ══════════════════════════════════════════════════════════════════════════════
with tab_cmp:
    st.markdown("#### 포트폴리오 비교")
    if len(st.session_state.pfs) < 2:
        st.info("사이드바에서 '복사본 생성'으로 My Portfolio N을 만들고 종목을 수정한 뒤 비교하세요.")
    else:
        if st.button("전체 포트폴리오 VaR 일괄 계산", type="primary"):
            for i, p in enumerate(st.session_state.pfs):
                if p["positions"]:
                    st.session_state.vr[i] = calc_pf_var(
                        p["positions"], usdkrw, c_sb, c_so, c_bo, sqrt_h, ykr, yus
                    )
            st.rerun()

        cmp_rows, names, pv99s, svs, bvs, ovs, diffs = [], [], [], [], [], [], []
        base_v = None
        for i, p in enumerate(st.session_state.pfs):
            r = st.session_state.vr.get(i)
            if not r and p["positions"]:
                r = calc_pf_var(p["positions"], usdkrw, c_sb, c_so, c_bo,
                                sqrt_h, ykr, yus)
            if r:
                if base_v is None: base_v = r["pv99"]
                delta = r["pv99"] - base_v
                names.append(p["name"])
                pv99s.append(round(r["pv99"]/1e6, 1))
                svs.append(round(r["sv"]["pv99"]/1e6, 1))
                bvs.append(round(r["bv"]["pv99"]/1e6, 1))
                ovs.append(round(r["ov"]["pv99"]/1e6, 1))
                diffs.append(round(delta/1e6, 1))
                cmp_rows.append({
                    "포트폴리오": p["name"],
                    "종목수":     len(p["positions"]),
                    "총평가":     fmt(r["total"])+"원",
                    "P-VaR99%":   fmt(r["pv99"])+"원",
                    "분산효과":   fmt(r["divers"])+"원",
                    "기준대비":   (f"+{round(delta/1e6,1)}M"
                                   if delta > 0 else
                                   f"{round(delta/1e6,1)}M"
                                   if i > 0 else "기준"),
                })
        if cmp_rows:
            st.dataframe(pd.DataFrame(cmp_rows), use_container_width=True)
            c1, c2 = st.columns(2)
            with c1:
                fc = go.Figure(go.Bar(
                    x=names, y=pv99s,
                    marker_color=[COLORS[i%len(COLORS)] for i in range(len(names))],
                    text=[f"{v}M" for v in pv99s], textposition="outside",
                ))
                fc.update_layout(title="VaR 99% 비교 (백만원)", height=300,
                                  showlegend=False, margin=dict(t=50,b=30,l=30,r=20))
                st.plotly_chart(fc, use_container_width=True)
            with c2:
                fd2 = go.Figure(go.Bar(
                    x=names, y=diffs,
                    marker_color=["#888780" if d==0 else
                                  "#A32D2D" if d>0 else "#3B6D11"
                                  for d in diffs],
                    text=[f"{'+'if d>0 else''}{d}M" for d in diffs],
                    textposition="outside",
                ))
                fd2.update_layout(title="VaR 변화 (기준 대비)", height=300,
                                   showlegend=False, margin=dict(t=50,b=30,l=30,r=20))
                st.plotly_chart(fd2, use_container_width=True)

            fstk = go.Figure([
                go.Bar(name="주식",      x=names, y=svs, marker_color="#378ADD"),
                go.Bar(name="채권",      x=names, y=bvs, marker_color="#3B6D11"),
                go.Bar(name="옵션/선물", x=names, y=ovs, marker_color="#EF9F27"),
            ])
            fstk.update_layout(barmode="stack",
                                title="자산군 VaR 구성 비교 (백만원)", height=280,
                                legend=dict(orientation="h", y=-0.2),
                                margin=dict(t=50,b=60,l=30,r=20))
            st.plotly_chart(fstk, use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 6 — 스트레스 테스트 (주가/금리/환율 개별 지정)
# ══════════════════════════════════════════════════════════════════════════════
with tab_st:
    st.markdown("#### 스트레스 테스트")

    r = st.session_state.vr.get(st.session_state.cpf)
    if not r and positions:
        r = calc_pf_var(positions, usdkrw, c_sb, c_so, c_bo, sqrt_h, ykr, yus)
        st.session_state.vr[st.session_state.cpf] = r

    if not r:
        st.warning("포트폴리오를 먼저 구성하세요.")
    else:
        base = r["pv99"]
        st.metric("기준 VaR 99%", f"{fmt(base)}원")
        st.markdown("---")

        # ── 사전 정의 시나리오 ────────────────────────────────────────────
        PRESET = {
            "글로벌 금융위기(2008)": {"eq":-0.45,"rate_bp": 150,"fx_pct": 15.0,"vol":3.0},
            "COVID-19 충격(2020)":   {"eq":-0.35,"rate_bp": -50,"fx_pct": 10.0,"vol":2.5},
            "아시아 외환위기(1997)": {"eq":-0.50,"rate_bp": 500,"fx_pct": 50.0,"vol":3.5},
            "호르무즈 봉쇄/지정학":  {"eq":-0.20,"rate_bp":  50,"fx_pct":  8.0,"vol":2.0},
            "금리 급등(2022)":       {"eq":-0.25,"rate_bp": 250,"fx_pct": -3.0,"vol":1.8},
            "달러강세/신흥국 위기":  {"eq":-0.15,"rate_bp": 100,"fx_pct": 20.0,"vol":1.6},
        }

        st.markdown("##### 사전 정의 시나리오")
        col_a, col_b = st.columns(2)
        preset_keys  = list(PRESET.keys())
        chk = {}
        with col_a:
            for k in preset_keys[:3]:
                chk[k] = st.checkbox(
                    k, value=(k in ["글로벌 금융위기(2008)","COVID-19 충격(2020)"])
                )
        with col_b:
            for k in preset_keys[3:]:
                chk[k] = st.checkbox(k)

        # ── 사용자 정의 시나리오 ──────────────────────────────────────────
        st.markdown("---")
        st.markdown("##### 사용자 정의 시나리오")
        use_custom = st.checkbox("사용자 정의 시나리오 추가")
        custom_sc  = {}
        if use_custom:
            st.markdown(
                '<div class="ib">'
                '주가지수 변동(%), 금리 변동(bp), 환율 변동(%)을 각각 독립적으로 지정하세요.<br>'
                '음수 = 하락/금리 완화 | 양수 = 상승/금리 긴축/원화 절하'
                '</div>', unsafe_allow_html=True
            )
            u1, u2, u3, u4 = st.columns(4)
            with u1:
                cname = st.text_input("시나리오 이름", "사용자 정의")
            with u2:
                ceq   = st.number_input(
                    "주가지수 변동 (%)", -80.0, 50.0, -20.0, 1.0,
                    help="예: -30 입력 → 주가지수 30% 하락"
                )
            with u3:
                crate = st.number_input(
                    "금리 변동 (bp)", -300.0, 500.0, 100.0, 10.0,
                    help="예: +200 → 금리 2%p(200bp) 상승"
                )
            with u4:
                cfx   = st.number_input(
                    "USD/KRW 변동 (%)", -30.0, 80.0, 10.0, 1.0,
                    help="예: +10 → 원화 10% 절하 (달러 상승)"
                )
            auto_vol = 1.0 + abs(ceq)/30 + abs(crate)/200 + abs(cfx)/20
            st.caption(f"자동 산정 변동성 배율: ×{auto_vol:.2f}")
            custom_sc[cname] = {
                "eq":      ceq / 100,
                "rate_bp": crate,
                "fx_pct":  cfx,
                "vol":     auto_vol,
            }

        # ── 스트레스 VaR 계산 ─────────────────────────────────────────────
        def stress_var_calc(sc, positions, base_var, usdkrw, sqrt_h):
            eq      = sc["eq"]
            rate_bp = sc["rate_bp"]
            fx_pct  = sc["fx_pct"] / 100
            vol     = sc["vol"]
            sfx     = usdkrw * (1 + fx_pct)
            loss    = 0.0
            for p in positions:
                fx2 = sfx if p["ccy"] == "USD" else 1.0
                if p["type"] == "Stock":
                    mv   = p["price"] * p["qty"] * fx2
                    # 주가 충격 손실 + 변동성 확대 VaR
                    loss += mv * abs(eq) + mv * (0.22/np.sqrt(252)*sqrt_h*vol) * Z99 * 0.3
                elif p["type"] == "Bond":
                    fv   = p.get("face_value", 10000)
                    dur  = p.get("maturity", 3.0) * 0.85   # 근사 Duration
                    mv   = fv * p["qty"] * fx2
                    # 금리 충격: ΔP ≈ -D × Δr × PV
                    loss += mv * dur * abs(rate_bp) / 10000
                    loss += mv * (0.05/np.sqrt(252)*sqrt_h*vol) * Z99 * 0.2
                elif p["type"] in ("Option","Future"):
                    g2   = bs_greeks(p["price"],
                                     p.get("strike", p["price"]*1.05),
                                     p.get("mat", 0.25),
                                     p.get("rf", 0.035),
                                     p.get("iv", 0.18) * vol,
                                     p.get("callput", "C"))
                    loss += (abs(g2["delta"] * p["price"] * abs(eq)) +
                             0.5 * g2["gamma"] * p["price"]**2 * eq**2
                             ) * p["qty"] * p.get("mult", 1) * fx2
            # FX 충격 손실 (USD 자산 원화환산 변화)
            usd_mv = sum(p["price"] * p["qty"] * p.get("mult", 1) * usdkrw
                         for p in positions if p["ccy"] == "USD")
            loss += usd_mv * abs(fx_pct) * 0.5
            return max(loss, base_var)

        if st.button("스트레스 시나리오 실행", type="primary",
                     use_container_width=True):
            active = {k: PRESET[k] for k, v in chk.items() if v}
            active.update(custom_sc)

            if not active:
                st.warning("시나리오를 하나 이상 선택하세요.")
            else:
                tbl_rows = []
                sc_names = ["기준 VaR"]
                sc_vals  = [round(base/1e6, 1)]
                sc_clrs  = ["#378ADD"]

                for name, sc in active.items():
                    sv   = stress_var_calc(sc, positions, base, usdkrw, sqrt_h)
                    add  = sv - base
                    grade = ("상" if sc["vol"] >= 3 else
                             "중" if sc["vol"] >= 2 else "하")
                    tbl_rows.append({
                        "시나리오":    name,
                        "주가지수":    f'{sc["eq"]*100:+.1f}%',
                        "금리(bp)":    f'{sc["rate_bp"]:+.0f}bp',
                        "환율":        f'{sc["fx_pct"]:+.1f}%',
                        "변동성배율":  f'×{sc["vol"]:.1f}',
                        "스트레스VaR": fmt(sv)+"원",
                        "추가손실":    "+"+fmt(add)+"원",
                        "위험등급":    grade,
                    })
                    sc_names.append(name)
                    sc_vals.append(round(sv/1e6, 1))
                    sc_clrs.append(
                        "#A32D2D" if sc["vol"] >= 3 else
                        "#BA7517" if sc["vol"] >= 2 else "#EF9F27"
                    )

                st.dataframe(pd.DataFrame(tbl_rows), use_container_width=True)

                fst = go.Figure(go.Bar(
                    x=sc_names, y=sc_vals, marker_color=sc_clrs,
                    text=[f"{v:.1f}M" for v in sc_vals],
                    textposition="outside",
                ))
                fst.add_hline(
                    y=round(base/1e6, 1), line_dash="dot",
                    line_color="#185FA5", annotation_text="기준 VaR"
                )
                fst.update_layout(
                    title="시나리오별 스트레스 VaR (백만원)",
                    height=340, showlegend=False,
                    margin=dict(t=50,b=40,l=40,r=20),
                )
                st.plotly_chart(fst, use_container_width=True)

                # 사용자 정의: 리스크 기여 분해
                if use_custom and custom_sc:
                    cname_key = list(custom_sc.keys())[0]
                    sc2 = custom_sc[cname_key]
                    eq_l    = r["sv"]["total"] * abs(sc2["eq"]) if r["sv"]["total"] else 0
                    rate_l  = r["bv"]["total"] * 3 * abs(sc2["rate_bp"]) / 10000 if r["bv"]["total"] else 0
                    fx_l    = sum(
                        p["price"]*p["qty"]*(usdkrw if p["ccy"]=="USD" else 0)
                        for p in positions
                    ) * abs(sc2["fx_pct"] / 100)
                    vol_add = base * (sc2["vol"] - 1)

                    st.markdown(f"##### '{cname_key}' — 리스크 기여 분해")
                    r1, r2, r3, r4 = st.columns(4)
                    with r1: st.metric("주가 충격 손실",  fmt(eq_l)+"원")
                    with r2: st.metric("금리 충격 손실",  fmt(rate_l)+"원")
                    with r3: st.metric("환율 충격 손실",  fmt(fx_l)+"원")
                    with r4: st.metric("변동성 확대 VaR", fmt(vol_add)+"원")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 7 — 시뮬레이션 분석 (MJD / Block Bootstrap / t-MC)
# ══════════════════════════════════════════════════════════════════════════════
with tab_sim:
    st.markdown("#### 고급 시뮬레이션 분석")
    st.markdown(
        '<div class="ib">'
        'Merton Jump Diffusion / Historical Block Bootstrap / t-분포 Monte Carlo'
        ' 3가지 기법으로 개별 종목 또는 포트폴리오를 시뮬레이션하고 결과를 비교합니다.'
        '</div>', unsafe_allow_html=True
    )

    # ── 입력 패널 ─────────────────────────────────────────────────────────
    with st.expander("시뮬레이션 설정", expanded=True):
        s_c1, s_c2, s_c3, s_c4 = st.columns(4)
        with s_c1:
            sim_ticker = st.text_input(
                "티커 (개별 종목)",
                value=positions[0]["ticker"] if positions else "005930.KS",
                help="포트폴리오 전체가 아닌 단일 종목을 대상으로 시뮬레이션합니다."
            )
        with s_c2:
            sim_T      = st.number_input("예측 기간 (영업일)", 20, 504, 63,
                                          help="63=3개월, 126=6개월, 252=1년")
            sim_T_yr   = sim_T / 252
        with s_c3:
            sim_n      = st.number_input("시뮬레이션 경로 수", 500, 10000, 2000, 500)
        with s_c4:
            sim_conf   = st.selectbox("VaR 신뢰수준", ["99%","95%","90%"])
            sim_conf_f = float(sim_conf.replace("%","")) / 100

        st.markdown("---")

        # MJD 파라미터
        mjd_col1, mjd_col2 = st.columns(2)
        with mjd_col1:
            st.markdown("**Merton Jump Diffusion 파라미터**")
            m1a, m1b, m1c = st.columns(3)
            with m1a:
                mjd_lam    = st.number_input("λ (점프 강도/년)", 0.1, 20.0, 3.0, 0.5,
                                              help="연간 평균 점프 발생 횟수")
            with m1b:
                mjd_mu_j   = st.number_input("μJ (점프 평균)", -0.5, 0.5, -0.05, 0.01,
                                              help="로그 점프 크기 평균")
            with m1c:
                mjd_sig_j  = st.number_input("σJ (점프 변동성)", 0.01, 1.0, 0.10, 0.01,
                                              help="로그 점프 크기 표준편차")
        with mjd_col2:
            st.markdown("**t-분포 MC 파라미터**")
            t_col1, t_col2 = st.columns(2)
            with t_col1:
                t_df = st.number_input("자유도 (ν)", 2, 30, 5,
                                        help="낮을수록 꼬리 두꺼움. 통상 3~8")
            with t_col2:
                bs_block = st.number_input("Bootstrap 블록 크기 (영업일)", 1, 60, 10,
                                            help="시계열 자기상관 보존 단위")

        sim_methods = st.multiselect(
            "실행할 시뮬레이션 기법",
            ["Merton Jump Diffusion (MJD)",
             "Historical Block Bootstrap",
             "t-분포 Monte Carlo"],
            default=["Merton Jump Diffusion (MJD)",
                     "Historical Block Bootstrap",
                     "t-분포 Monte Carlo"],
        )

    run_sim = st.button("시뮬레이션 실행", type="primary",
                         use_container_width=True)

    if run_sim:
        if not sim_methods:
            st.warning("기법을 하나 이상 선택하세요.")
        else:
            with st.spinner(f"{sim_ticker} 데이터 수집 및 시뮬레이션 중..."):

                # ── 데이터 수집 ─────────────────────────────────────────
                hist_rets, S0, data_live = sim_fetch_returns(sim_ticker, days=504)
                stats_d = sim_summary_stats(hist_rets)

                mu_ann  = stats_d["연율수익률"]
                sig_ann = stats_d["연율변동성"]
                DT      = 1 / 252

            # ── 기초 통계량 ──────────────────────────────────────────────
            st.markdown("---")
            data_src = "yfinance 실시간" if data_live else "샘플(GBM)"
            st.markdown(
                f"##### {sim_ticker} 기초 통계량  "
                f"<span style='font-size:12px;color:#888'>({data_src}, "
                f"관측 {len(hist_rets)}일)</span>",
                unsafe_allow_html=True
            )
            stat_c = st.columns(4)
            with stat_c[0]: st.metric("연율 수익률",  f"{stats_d['연율수익률']*100:.2f}%")
            with stat_c[1]: st.metric("연율 변동성",  f"{stats_d['연율변동성']*100:.2f}%")
            with stat_c[2]: st.metric("왜도",          f"{stats_d['왜도']:.3f}")
            with stat_c[3]: st.metric("초과첨도",      f"{stats_d['초과첨도']:.3f}")

            jb_note = ("정규분포 기각 (두꺼운 꼬리 확인됨)"
                       if stats_d["JB p-value"] < 0.05
                       else "정규분포 기각 불가")
            st.caption(
                f"Jarque-Bera 검정: JB={stats_d['JB통계량']:.1f}, "
                f"p={stats_d['JB p-value']:.4f} → {jb_note}"
            )

            # ── 각 기법 실행 ─────────────────────────────────────────────
            all_paths   = {}
            all_metrics = {}

            if "Merton Jump Diffusion (MJD)" in sim_methods:
                with st.spinner("MJD 계산 중..."):
                    p_mjd = sim_mjd(
                        S0=S0, mu=mu_ann, sigma=sig_ann,
                        lam=mjd_lam, mu_j=mjd_mu_j, sigma_j=mjd_sig_j,
                        T=sim_T_yr, dt=DT, n_paths=int(sim_n),
                    )
                all_paths["MJD"]   = p_mjd
                all_metrics["MJD"] = sim_metrics(p_mjd, S0, sim_conf_f)

            if "Historical Block Bootstrap" in sim_methods:
                with st.spinner("Block Bootstrap 계산 중..."):
                    p_bs = sim_bootstrap(
                        S0=S0, hist_rets=hist_rets,
                        T=sim_T_yr, dt=DT, n_paths=int(sim_n),
                        block_size=int(bs_block),
                    )
                all_paths["Bootstrap"]   = p_bs
                all_metrics["Bootstrap"] = sim_metrics(p_bs, S0, sim_conf_f)

            if "t-분포 Monte Carlo" in sim_methods:
                with st.spinner("t-분포 MC 계산 중..."):
                    p_t = sim_t_mc(
                        S0=S0, mu=mu_ann, sigma=sig_ann,
                        df=int(t_df), T=sim_T_yr, dt=DT,
                        n_paths=int(sim_n),
                    )
                all_paths["t-MC"]   = p_t
                all_metrics["t-MC"] = sim_metrics(p_t, S0, sim_conf_f)

            # ── 경로 팬 차트 ─────────────────────────────────────────────
            st.markdown("---")
            st.markdown("##### 시뮬레이션 경로 팬 차트")
            st.caption("초록=최선 시나리오 / 빨강=중앙값 / 검정=최악 시나리오")

            method_labels = {
                "MJD": "Merton Jump Diffusion",
                "Bootstrap": "Historical Block Bootstrap",
                "t-MC": "t-분포 Monte Carlo",
            }
            n_methods = len(all_paths)
            if n_methods == 1:
                key = list(all_paths.keys())[0]
                fc  = make_fan_chart(all_paths[key], DT,
                                     f"{method_labels[key]} — {sim_ticker}")
                st.plotly_chart(fc, use_container_width=True)
            elif n_methods >= 2:
                fan_cols = st.columns(min(n_methods, 2))
                for ci, (key, paths_k) in enumerate(all_paths.items()):
                    with fan_cols[ci % 2]:
                        fc = make_fan_chart(paths_k, DT,
                                            f"{method_labels[key]}")
                        st.plotly_chart(fc, use_container_width=True)

            # ── 만기 분포 히스토그램 ──────────────────────────────────────
            st.markdown("##### 만기 주가 분포 히스토그램")
            hist_cols = st.columns(min(n_methods, 3))
            for ci, (key, paths_k) in enumerate(all_paths.items()):
                with hist_cols[ci % 3]:
                    fh = make_terminal_hist(paths_k, S0,
                                            method_labels[key], sim_conf_f)
                    st.plotly_chart(fh, use_container_width=True)

            # ── VaR 비교 테이블 ────────────────────────────────────────────
            st.markdown("---")
            st.markdown("##### 기법별 성과 지표 비교")

            var_key  = f"VaR {int(sim_conf_f*100)}%"
            cvar_key = f"CVaR {int(sim_conf_f*100)}%"

            comp_rows = []
            for key, m in all_metrics.items():
                comp_rows.append({
                    "기법":           method_labels[key],
                    "기대수익률":     f"{m['기대수익률(로그)']*100:.2f}%",
                    "변동성":         f"{m['수익률 표준편차']*100:.2f}%",
                    f"VaR {sim_conf}": f"{m[var_key]*100:.2f}%",
                    f"CVaR {sim_conf}":f"{m[cvar_key]*100:.2f}%",
                    "손실확률":       f"{m['손실확률']*100:.1f}%",
                    "5% 분위 주가":   f"{m['5% 분위 주가']:,.0f}",
                    "중앙값 주가":    f"{m['중앙값 주가']:,.0f}",
                    "95% 분위 주가":  f"{m['95% 분위 주가']:,.0f}",
                    "기대 주가":      f"{m['기대 주가']:,.0f}",
                })
            st.dataframe(pd.DataFrame(comp_rows).set_index("기법"),
                         use_container_width=True)

            # ── VaR 막대 비교 차트 ────────────────────────────────────────
            if n_methods > 1:
                fig_cmp = go.Figure()
                keys_list  = list(all_metrics.keys())
                var_values = [abs(all_metrics[k][var_key]) * 100 for k in keys_list]
                cvar_values= [abs(all_metrics[k][cvar_key])* 100 for k in keys_list]
                labels     = [method_labels[k] for k in keys_list]

                fig_cmp.add_trace(go.Bar(
                    x=labels, y=var_values,
                    name=f"VaR {sim_conf}",
                    marker_color="#A32D2D",
                    text=[f"{v:.2f}%" for v in var_values],
                    textposition="outside",
                ))
                fig_cmp.add_trace(go.Bar(
                    x=labels, y=cvar_values,
                    name=f"CVaR {sim_conf}",
                    marker_color="#854F0B",
                    text=[f"{v:.2f}%" for v in cvar_values],
                    textposition="outside",
                ))
                fig_cmp.update_layout(
                    title=f"기법별 VaR / CVaR 비교 ({sim_conf}, {sim_T}일 예측)",
                    yaxis_title="손실률 (%)",
                    height=340,
                    barmode="group",
                    legend=dict(orientation="h", y=-0.18),
                    margin=dict(t=50, b=70, l=50, r=20),
                )
                st.plotly_chart(fig_cmp, use_container_width=True)

            # ── 기법 선택 가이드 ──────────────────────────────────────────
            st.markdown("---")
            st.markdown("##### 기법 선택 가이드 (해석)")
            guide_data = {
                "기법":    ["MJD", "Block Bootstrap", "t-분포 MC"],
                "핵심 강점":["급등락·불연속 충격 포착 (공시·이벤트)",
                              "분포 가정 없이 실제 패턴 재현",
                              "Fat-tail 반영, VaR 과소추정 방지"],
                "적합한 자산":["개별 주식, 옵션 (급변동 자산)",
                                "지수·ETF, 충분한 과거 데이터 보유 자산",
                                "포트폴리오 전체, 리스크 관리 목적"],
                "한계":    ["파라미터(λ·μJ·σJ) 추정 어려움",
                             "과거 패턴 미래 반복 가정, 샘플 의존적",
                             "자유도 설정에 결과 민감, 정상성 가정"],
            }
            st.dataframe(pd.DataFrame(guide_data).set_index("기법"),
                         use_container_width=True)

            # ── 자동 분석 코멘트 ──────────────────────────────────────────
            st.markdown("---")
            st.markdown("##### 자동 분석 요약")
            if all_metrics:
                best_var_key = min(all_metrics,
                                   key=lambda k: abs(all_metrics[k][var_key]))
                worst_var_key = max(all_metrics,
                                    key=lambda k: abs(all_metrics[k][var_key]))
                st.info(
                    f"**{sim_ticker}** 시뮬레이션 결과 ({sim_T}영업일, "
                    f"경로 {int(sim_n):,}개, {sim_conf}):\n\n"
                    f"- 가장 보수적(높은 VaR): **{method_labels[worst_var_key]}**"
                    f" → VaR {abs(all_metrics[worst_var_key][var_key])*100:.2f}%\n"
                    f"- 가장 낙관적(낮은 VaR): **{method_labels[best_var_key]}**"
                    f" → VaR {abs(all_metrics[best_var_key][var_key])*100:.2f}%\n"
                    f"- 연율 변동성: {sig_ann*100:.1f}% | "
                    f"Jarque-Bera: {jb_note}\n"
                    f"- Fat-tail 특성{'이 뚜렷하므로' if stats_d['초과첨도'] > 1 else '이 약하므로'} "
                    f"{'t-분포 MC 또는 MJD가 더 적합합니다.' if stats_d['초과첨도'] > 1 else 'Bootstrap이 안정적입니다.'}"
                )

                # 리포트 연동용 세션 저장
                st.session_state["sim_result_for_report"] = {
                    "ticker":  sim_ticker,
                    "T_days":  sim_T,
                    "n_paths": int(sim_n),
                    "conf":    sim_conf_f,   # float (0.99), not string
                    "stats":   stats_d,
                    "metrics": all_metrics,
                }

# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# TAB 7 — Parametric VaR (EWMA 기반)
# ══════════════════════════════════════════════════════════════════════════════
with tab_ewma:
    st.markdown("#### Parametric VaR 분석")
    st.markdown(
        '''<div class="ib">
        <b>Parametric VaR</b> (EWMA λ=0.94) + <b>Historical Simulation VaR</b>를 함께 계산합니다.<br>
        Monte Carlo Simulation VaR는 별도 버튼으로 실행합니다.<br>
        데이터: yfinance 직전 2년(504영업일) | Marginal VaR / Component VaR / Incremental VaR 포함
        </div>''', unsafe_allow_html=True
    )

    with st.expander("분석 설정", expanded=True):
        ew_c1, ew_c2, ew_c3, ew_c4 = st.columns(4)
        with ew_c1:
            ew_conf_sel = st.selectbox("신뢰수준", ["99%","95%"], key="ew_conf")
            ew_conf     = float(ew_conf_sel.replace("%","")) / 100
        with ew_c2:
            ew_hold = st.number_input("보유기간 (영업일)", 1, 250, 10, key="ew_hold",
                                       help="바젤 기준: 10일")
        with ew_c3:
            ew_lam  = st.number_input("EWMA λ", 0.80, 0.99, 0.94, 0.01, key="ew_lam",
                                       help="RiskMetrics 일별 기본값: 0.94")
        with ew_c4:
            ew_days = st.number_input("관측기간 (영업일)", 126, 1260, 504, key="ew_days",
                                       help="2년=504영업일")

    run_var = st.button("Parametric VaR + Historical VaR 계산 실행",
                         type="primary", use_container_width=True, key="run_var_main")

    if run_var:
        if not positions:
            st.warning("포트폴리오를 먼저 구성하세요.")
        else:
            tickers_ew  = [p["ticker"] for p in positions]
            mv_list_ew  = [p["price"]*p["qty"]*(usdkrw if p["ccy"]=="USD" else 1)
                           for p in positions]
            total_mv_ew = sum(mv_list_ew)
            weights_ew  = [mv/total_mv_ew for mv in mv_list_ew]

            with st.spinner("수익률 수집 및 VaR 계산 중..."):
                ret_dict, price_dict, live_ew = fetch_returns_multi(
                    tickers_ew, days=int(ew_days)
                )

            data_src = "yfinance 실시간" if live_ew else "샘플(GBM)"
            st.caption(f"데이터: {data_src} | 관측: {ew_days}일 | 신뢰수준: {ew_conf_sel} | 보유기간: {ew_hold}일")
            st.markdown("---")

            # ── 기초 통계량 ────────────────────────────────────────────
            st.markdown("##### 종목별 수익률 기초 통계")
            stat_rows = []
            for tk in tickers_ew:
                r  = ret_dict[tk]
                vt = ewma_volatility(r, ew_lam)
                sd = float(np.sqrt(vt[-1]))
                from scipy.stats import skew as _skew, kurtosis as _kurt, jarque_bera as _jb
                jb_stat, jb_p = _jb(r)
                stat_rows.append({
                    "티커": tk,
                    "일평균수익률": f"{float(np.mean(r))*100:.4f}%",
                    "EWMA σ(일)": f"{sd*100:.4f}%",
                    "EWMA σ(연율)": f"{sd*np.sqrt(252)*100:.2f}%",
                    f"σ({ew_hold}일)": f"{sd*np.sqrt(ew_hold)*100:.3f}%",
                    "왜도": f"{float(_skew(r)):.3f}",
                    "초과첨도": f"{float(_kurt(r)):.3f}",
                    "JB p-value": f"{jb_p:.4f}",
                    "정규분포": "기각" if jb_p < 0.05 else "채택",
                })
            st.dataframe(pd.DataFrame(stat_rows), use_container_width=True)

            # ── Parametric + Historical VaR 개별 종목 ─────────────────
            st.markdown("---")
            st.markdown(f"##### 개별 종목 VaR 비교 ({ew_conf_sel}, {ew_hold}일)")
            ind_pvars = []; ind_hvars_amt = []; ind_rows_comb = []
            for i, tk in enumerate(tickers_ew):
                r  = ret_dict[tk]
                mv = mv_list_ew[i]
                pv = parametric_var_ewma(r, mv, ew_hold, ew_conf, ew_lam)
                hv = historical_var(r, mv, ew_hold, ew_conf)
                ind_pvars.append(pv["var"])
                ind_hvars_amt.append(hv["var"])
                ind_rows_comb.append({
                    "티커":           tk,
                    "평가금액":       fmt(mv)+"원",
                    "P-VaR (EWMA)":   fmt(pv["var"])+"원",
                    "P-CVaR":         fmt(pv["cvar"])+"원",
                    "P-VaR%":         f"{pv['var_pct']*100:.3f}%",
                    "H-VaR (Hist.)":  fmt(hv["var"])+"원",
                    "H-CVaR":         fmt(hv["cvar"])+"원",
                    "H-VaR%":         f"{hv['var_pct']*100:.3f}%",
                    "P vs H 차이":    f"{(pv['var']-hv['var'])/hv['var']*100:+.1f}%" if hv["var"] else "N/A",
                })
            st.dataframe(pd.DataFrame(ind_rows_comb), use_container_width=True)

            # ── 포트폴리오 Parametric VaR ─────────────────────────────
            st.markdown("---")
            st.markdown(f"##### 포트폴리오 VaR — Parametric + Historical ({ew_conf_sel}, {ew_hold}일)")

            ret_matrix  = np.column_stack([ret_dict[tk] for tk in tickers_ew])
            cov_mat_ew  = ewma_cov_matrix(ret_matrix, ew_lam)
            mv_res      = marginal_var_ewma(weights_ew, cov_mat_ew,
                                            total_mv_ew, ew_conf, ew_hold)
            ivar_vec, _ = incremental_var_ewma(
                weights_ew, cov_mat_ew, total_mv_ew, ew_conf, ew_hold
            )
            h_pf_res    = historical_portfolio_var(
                ret_dict, tickers_ew, weights_ew, total_mv_ew, ew_hold, ew_conf
            )
            undiv_p = sum(ind_pvars)
            divers  = undiv_p - mv_res["pf_var"]

            # 요약 메트릭
            mc1,mc2,mc3,mc4 = st.columns(4)
            with mc1:
                st.markdown(
                    f'''<div class="mc"><div class="ml">Parametric VaR {ew_conf_sel}</div>
                    <div class="mv red">{fmt(mv_res["pf_var"])}원</div></div>''',
                    unsafe_allow_html=True)
            with mc2:
                st.markdown(
                    f'''<div class="mc"><div class="ml">Historical VaR {ew_conf_sel}</div>
                    <div class="mv amb">{fmt(h_pf_res["pf_var"])}원</div></div>''',
                    unsafe_allow_html=True)
            with mc3:
                st.markdown(
                    f'''<div class="mc"><div class="ml">분산 효과 (Parametric)</div>
                    <div class="mv grn">{fmt(divers)}원 ({divers/undiv_p*100:.1f}%)</div></div>''',
                    unsafe_allow_html=True)
            with mc4:
                st.markdown(
                    f'''<div class="mc"><div class="ml">포트폴리오 σ({ew_hold}일)</div>
                    <div class="mv blu">{mv_res["pf_sigma"]*100:.3f}%</div></div>''',
                    unsafe_allow_html=True)

            # Component / Marginal / Incremental VaR 테이블
            comp_rows_ew = []
            for i, tk in enumerate(tickers_ew):
                comp    = mv_res["comp_var"][i]
                contrib = comp / mv_res["pf_var"] * 100 if mv_res["pf_var"] else 0
                comp_rows_ew.append({
                    "티커":          tk,
                    "비중":          f"{weights_ew[i]*100:.1f}%",
                    "P-VaR(개별)":   fmt(ind_pvars[i])+"원",
                    "Marginal VaR":  fmt(mv_res["mvar"][i])+"원",
                    "Component VaR": fmt(comp)+"원",
                    "리스크 기여도": f"{contrib:.1f}%",
                    "Incremental VaR": fmt(ivar_vec[i])+"원",
                    "Beta(vs PF)":   f"{mv_res['beta'][i]:.3f}",
                    "H-VaR(개별)":   fmt(ind_hvars_amt[i])+"원",
                })
            st.dataframe(pd.DataFrame(comp_rows_ew), use_container_width=True)

            # P-VaR vs H-VaR 비교 차트
            fig_cmp_vh = go.Figure()
            fig_cmp_vh.add_trace(go.Bar(
                x=tickers_ew, y=[v/1e6 for v in ind_pvars],
                name="Parametric VaR", marker_color="#A32D2D",
                text=[f"{v/1e6:.1f}M" for v in ind_pvars], textposition="outside",
            ))
            fig_cmp_vh.add_trace(go.Bar(
                x=tickers_ew, y=[v/1e6 for v in ind_hvars_amt],
                name="Historical VaR", marker_color="#854F0B",
                text=[f"{v/1e6:.1f}M" for v in ind_hvars_amt], textposition="outside",
            ))
            fig_cmp_vh.update_layout(
                title=f"종목별 Parametric vs Historical VaR ({ew_conf_sel}, {ew_hold}일)",
                yaxis_title="VaR (백만원)", barmode="group", height=320,
                legend=dict(orientation="h", y=-0.2),
                margin=dict(t=50,b=60,l=40,r=20),
            )
            st.plotly_chart(fig_cmp_vh, use_container_width=True)

            # 리스크 기여도 도넛
            contrib_vals = [abs(mv_res["comp_var"][i]) for i in range(len(tickers_ew))]
            fig_donut = go.Figure(go.Pie(
                labels=tickers_ew, values=contrib_vals, hole=0.52,
                textinfo="label+percent", marker_colors=COLORS[:len(tickers_ew)],
            ))
            fig_donut.update_layout(
                title=f"Component VaR 리스크 기여도 ({ew_conf_sel})",
                height=300, showlegend=False, margin=dict(t=50,b=20,l=20,r=20),
            )
            st.plotly_chart(fig_donut, use_container_width=True)

            # 수익률 분포 & VaR 기준선 차트 (포트폴리오)
            st.markdown("##### 포트폴리오 수익률 분포 (Parametric vs Historical VaR)")
            pf_h_rets = np.array(h_pf_res["pf_returns"])
            fig_dist = go.Figure()
            fig_dist.add_trace(go.Histogram(
                x=(pf_h_rets * total_mv_ew / 1e6).tolist(), nbinsx=50,
                marker_color="rgba(55,138,221,0.55)", name="포트폴리오 손익 분포",
                showlegend=True,
            ))
            fig_dist.add_vline(x=-mv_res["pf_var"]/1e6, line_dash="dash",
                                line_color="#A32D2D",
                                annotation_text=f"P-VaR {ew_conf_sel}",
                                annotation_position="top left")
            fig_dist.add_vline(x=-h_pf_res["pf_var"]/1e6, line_dash="dot",
                                line_color="#854F0B",
                                annotation_text=f"H-VaR {ew_conf_sel}",
                                annotation_position="top right")
            fig_dist.update_layout(
                title="포트폴리오 일별 손익 분포 (백만원)",
                xaxis_title="손익 (백만원)", yaxis_title="빈도",
                height=280, showlegend=False, margin=dict(t=50,b=40,l=50,r=20),
            )
            st.plotly_chart(fig_dist, use_container_width=True)

            # EWMA 변동성 추이
            st.markdown("---")
            st.markdown("##### EWMA 변동성 추이 (최근 252일)")
            show_n = min(252, int(ew_days))
            dates_ew = [(datetime.today()-timedelta(days=show_n-i)).strftime("%m/%d")
                        for i in range(show_n)]
            fig_vol = go.Figure()
            for i, tk in enumerate(tickers_ew[:4]):
                vt  = ewma_volatility(ret_dict[tk], ew_lam)
                sig = np.sqrt(vt[-show_n:]) * np.sqrt(252) * 100
                fig_vol.add_trace(go.Scatter(
                    x=dates_ew, y=sig.tolist(), mode="lines", name=tk,
                    line=dict(color=COLORS[i%len(COLORS)], width=1.5),
                ))
            fig_vol.update_layout(
                title="EWMA 연율 변동성 추이 (%)",
                xaxis_title="날짜", yaxis_title="연율 변동성 (%)",
                height=280, legend=dict(orientation="h",y=-0.2),
                margin=dict(t=50,b=60,l=50,r=20),
            )
            st.plotly_chart(fig_vol, use_container_width=True)

            # 세션 저장 (리포트용 — Parametric + Historical)
            st.session_state["ewma_result"] = {
                "portfolio": {
                    f"Parametric VaR {ew_conf_sel} {ew_hold}일": mv_res["pf_var"],
                    f"Historical VaR {ew_conf_sel} {ew_hold}일":  h_pf_res["pf_var"],
                    "비분산 VaR (Parametric)": undiv_p,
                    "분산 효과": divers,
                    f"포트폴리오 σ({ew_hold}일)": mv_res["pf_sigma"],
                },
                "individual": comp_rows_ew,
            }
            st.success("계산 완료. '분석 리포트' 탭에서 리포트를 생성하세요.")

    # ── Monte Carlo VaR (별도 실행) ────────────────────────────────────────
    st.markdown("---")
    st.markdown("##### Monte Carlo Simulation VaR (선택 실행)")
    st.markdown(
        '''<div class="ib">
        EWMA 공분산 + t-분포(또는 정규분포) 기반 다변량 시뮬레이션으로 VaR를 계산합니다.
        계산량이 많으므로 별도 버튼으로 실행합니다.
        </div>''', unsafe_allow_html=True
    )
    mc_col1, mc_col2, mc_col3, mc_col4 = st.columns(4)
    with mc_col1:
        mc_n    = st.number_input("시뮬레이션 횟수", 1000, 100000, 10000, 1000, key="mc_n")
    with mc_col2:
        mc_use_t = st.checkbox("t-분포 사용 (Fat-tail)", value=True, key="mc_use_t")
    with mc_col3:
        mc_df   = st.number_input("t-분포 자유도 (ν)", 2, 30, 5, key="mc_df",
                                   disabled=not mc_use_t)
    with mc_col4:
        mc_seed = st.number_input("난수 시드", 0, 9999, 42, key="mc_seed")

    run_mc = st.button("Monte Carlo VaR 계산 실행",
                        type="secondary", use_container_width=True, key="run_mc_var")

    if run_mc:
        if not positions:
            st.warning("포트폴리오를 먼저 구성하세요.")
        else:
            with st.spinner(f"Monte Carlo 시뮬레이션 {int(mc_n):,}회 실행 중..."):
                tickers_mc = [p["ticker"] for p in positions]
                mv_mc      = [p["price"]*p["qty"]*(usdkrw if p["ccy"]=="USD" else 1)
                              for p in positions]
                total_mc   = sum(mv_mc)
                weights_mc = [m/total_mc for m in mv_mc]
                ret_mc, _, _ = fetch_returns_multi(tickers_mc, days=int(ew_days))

                # 개별 MC VaR
                ind_mcvars = []
                mc_ind_rows = []
                for i, tk in enumerate(tickers_mc):
                    mv_res_mc = monte_carlo_var(
                        ret_mc[tk], mv_mc[i], int(ew_hold),
                        ew_conf, int(mc_n), bool(mc_use_t), int(mc_df),
                        ew_lam, int(mc_seed)+i
                    )
                    ind_mcvars.append(mv_res_mc["var"])
                    mc_ind_rows.append({
                        "티커":       tk,
                        "평가금액":   fmt(mv_mc[i])+"원",
                        "MC-VaR":     fmt(mv_res_mc["var"])+"원",
                        "MC-CVaR":    fmt(mv_res_mc["cvar"])+"원",
                        "MC-VaR%":    f"{mv_res_mc['var_pct']*100:.3f}%",
                    })

                # 포트폴리오 MC VaR
                pf_mc_res = monte_carlo_portfolio_var(
                    ret_mc, tickers_mc, weights_mc, total_mc,
                    int(ew_hold), ew_conf, int(mc_n),
                    bool(mc_use_t), int(mc_df), ew_lam, int(mc_seed)
                )

            mc_dist_str = f"t-분포(ν={int(mc_df)})" if mc_use_t else "정규분포"
            st.caption(f"Monte Carlo 설정: {int(mc_n):,}회 시뮬레이션 | {mc_dist_str} | {ew_conf_sel} | {ew_hold}일")

            # 요약 메트릭
            mcc1, mcc2, mcc3 = st.columns(3)
            with mcc1:
                st.markdown(
                    f'''<div class="mc"><div class="ml">MC VaR {ew_conf_sel} (포트폴리오)</div>
                    <div class="mv red">{fmt(pf_mc_res["pf_var"])}원</div></div>''',
                    unsafe_allow_html=True)
            with mcc2:
                st.markdown(
                    f'''<div class="mc"><div class="ml">MC CVaR {ew_conf_sel}</div>
                    <div class="mv amb">{fmt(pf_mc_res["pf_cvar"])}원</div></div>''',
                    unsafe_allow_html=True)
            with mcc3:
                st.markdown(
                    f'''<div class="mc"><div class="ml">MC VaR% (포트폴리오)</div>
                    <div class="mv blu">{pf_mc_res["var_pct"]*100:.3f}%</div></div>''',
                    unsafe_allow_html=True)

            st.markdown("##### 종목별 Monte Carlo VaR")
            st.dataframe(pd.DataFrame(mc_ind_rows), use_container_width=True)

            # MC 손익 분포 히스토그램
            mc_rets_arr = np.array(pf_mc_res["pf_returns"])
            fig_mc_dist = go.Figure()
            fig_mc_dist.add_trace(go.Histogram(
                x=(mc_rets_arr * total_mc / 1e6).tolist(), nbinsx=60,
                marker_color="rgba(83,74,183,0.55)", name="MC 손익 분포",
            ))
            fig_mc_dist.add_vline(x=-pf_mc_res["pf_var"]/1e6,
                                   line_dash="dash", line_color="#A32D2D",
                                   annotation_text=f"MC VaR {ew_conf_sel}",
                                   annotation_position="top left")
            fig_mc_dist.add_vline(x=-pf_mc_res["pf_cvar"]/1e6,
                                   line_dash="dot", line_color="#854F0B",
                                   annotation_text=f"MC CVaR {ew_conf_sel}",
                                   annotation_position="top right")
            fig_mc_dist.update_layout(
                title=f"Monte Carlo 포트폴리오 손익 분포 ({mc_dist_str}, {int(mc_n):,}회)",
                xaxis_title="포트폴리오 손익 (백만원)", yaxis_title="빈도",
                height=280, showlegend=False,
                margin=dict(t=50,b=40,l=50,r=20),
            )
            st.plotly_chart(fig_mc_dist, use_container_width=True)

            # 세션 저장 (MC 결과)
            st.session_state["mc_var_result"] = {
                "pf_var":    pf_mc_res["pf_var"],
                "pf_cvar":   pf_mc_res["pf_cvar"],
                "var_pct":   pf_mc_res["var_pct"],
                "ind_rows":  mc_ind_rows,
                "dist_type": mc_dist_str,
                "n_sims":    int(mc_n),
                "conf":      ew_conf,
                "holding":   int(ew_hold),
            }
            st.success("Monte Carlo VaR 계산 완료.")

# ══════════════════════════════════════════════════════════════════════════════
with tab_bt:
    st.markdown("#### VaR Back-testing")
    st.markdown(
        '<div class="ib">'
        'Basel II/III 기준: 250영업일 중 VaR 초과 일수가 4일 이하 → 녹색존 (모형 적합).<br>'
        '5~9일 → 황색존 (경고), 10일 이상 → 적색존 (모형 부적합).<br>'
        '개별 종목 또는 포트폴리오 전체를 대상으로 과거 실제 손익과 VaR를 비교합니다.'
        '</div>', unsafe_allow_html=True
    )

    if not positions:
        st.warning("포트폴리오를 먼저 구성하세요.")
    else:
        with st.expander("Back-testing 설정", expanded=True):
            bt_c1, bt_c2, bt_c3, bt_c4 = st.columns(4)
            with bt_c1:
                bt_ticker = st.selectbox(
                    "분석 대상",
                    ["전체 포트폴리오"] + [p["ticker"] for p in positions],
                    key="bt_ticker",
                )
            with bt_c2:
                bt_conf_sel = st.selectbox("신뢰수준", ["99%","95%"], key="bt_conf")
                bt_conf     = float(bt_conf_sel.replace("%","")) / 100
            with bt_c3:
                bt_window = st.number_input(
                    "검증 기간 (영업일)", 60, 504, 250, key="bt_window",
                    help="Basel: 250일 기준"
                )
            with bt_c4:
                bt_lam = st.number_input(
                    "EWMA λ", 0.80, 0.99, 0.94, 0.01, key="bt_lam"
                )

        run_bt = st.button("Back-testing 실행", type="primary",
                            use_container_width=True, key="run_bt")

        if run_bt:
            with st.spinner("과거 데이터 수집 및 VaR 계산 중..."):

                # ── 데이터 수집 ──────────────────────────────────────────
                n_total = int(bt_window) + 252   # 여유분 포함
                if bt_ticker == "전체 포트폴리오":
                    # 포트폴리오 일별 손익 = Σ(wᵢ × rᵢ) × 총평가금액
                    total_mv_bt = sum(
                        p["price"]*p["qty"]*(usdkrw if p["ccy"]=="USD" else 1)
                        for p in positions
                    )
                    pnl_series   = np.zeros(n_total)
                    ret_mat_list = []
                    for p in positions:
                        ret_tk, px_tk, _ = sim_fetch_returns(p["ticker"], n_total)
                        w_i = p["price"]*p["qty"]*(usdkrw if p["ccy"]=="USD" else 1) / total_mv_bt
                        pnl_series += ret_tk * w_i
                        ret_mat_list.append(ret_tk)
                    pnl_series = pnl_series * total_mv_bt    # 금액 단위
                    # 포트폴리오 EWMA 변동성 시계열
                    vol_series = ewma_volatility(pnl_series / total_mv_bt, bt_lam) * total_mv_bt
                else:
                    p_sel = next((p for p in positions if p["ticker"]==bt_ticker), None)
                    if p_sel is None:
                        st.error("종목을 찾을 수 없습니다.")
                        st.stop()
                    ret_tk, px_tk, _ = sim_fetch_returns(bt_ticker, n_total)
                    mv_bt    = p_sel["price"] * p_sel["qty"] * (usdkrw if p_sel["ccy"]=="USD" else 1)
                    pnl_series = ret_tk * mv_bt
                    vol_series = ewma_volatility(ret_tk, bt_lam) * mv_bt
                    total_mv_bt = mv_bt

                # VaR 시계열 (t-1일 변동성으로 t일 VaR 계산 → look-ahead 방지)
                z           = stats.norm.ppf(bt_conf)
                var_series  = z * vol_series
                var_series  = np.roll(var_series, 1)   # 1일 시차 적용
                var_series[0] = var_series[1]

                # 검증 기간만 추출 (마지막 bt_window일)
                pnl  = pnl_series[-bt_window:]
                var_ = var_series[-bt_window:]

                # ── 초과 횟수 계산 ────────────────────────────────────────
                # 실제 손실(-pnl)이 VaR를 초과하면 예외(Exception)
                exceptions  = (-pnl > var_).sum()
                exc_rate    = exceptions / bt_window * 100
                # Basel 존 판정
                if exceptions <= 4:
                    zone = "🟢 녹색존 (모형 적합)"
                    zone_color = "#3B6D11"
                elif exceptions <= 9:
                    zone = "🟡 황색존 (경고)"
                    zone_color = "#854F0B"
                else:
                    zone = "🔴 적색존 (모형 부적합)"
                    zone_color = "#A32D2D"

                # Kupiec POF 검정 (비율 검정)
                p_theo  = 1 - bt_conf
                k       = int(exceptions); n = int(bt_window)
                from math import log
                if k == 0:
                    lr_kupiec = 0.0
                else:
                    lr_kupiec = 2 * (
                        k * log(k/n/p_theo + 1e-12) +
                        (n-k) * log((n-k)/n/(1-p_theo) + 1e-12)
                    )
                p_kupiec = 1 - stats.chi2.cdf(lr_kupiec, df=1)
                kupiec_result = "기각(모형 부적합)" if p_kupiec < 0.05 else "채택(모형 적합)"

            # ── 결과 요약 ──────────────────────────────────────────────────
            st.markdown("---")
            st.markdown(f"##### Back-testing 결과 — {bt_ticker} ({bt_conf_sel}, {bt_window}일)")

            bc1, bc2, bc3, bc4 = st.columns(4)
            with bc1:
                st.markdown(
                    f'<div class="mc"><div class="ml">VaR 초과 횟수</div>'
                    f'<div class="mv" style="color:{zone_color}">{exceptions}일</div>'
                    f'<div style="font-size:11px;color:{zone_color}">{zone}</div></div>',
                    unsafe_allow_html=True
                )
            with bc2:
                st.metric("초과 비율", f"{exc_rate:.2f}%",
                          delta=f"기대 {(1-bt_conf)*100:.1f}% 대비")
            with bc3:
                st.metric("Kupiec LR 통계량", f"{lr_kupiec:.3f}",
                          delta=f"p={p_kupiec:.4f} → {kupiec_result}")
            with bc4:
                avg_var = float(np.mean(var_))
                st.metric(f"평균 VaR ({bt_conf_sel})",
                          f"{fmt(avg_var)}원")

            # ── 시각화 ────────────────────────────────────────────────────
            dates_bt = [
                (datetime.today() - timedelta(days=bt_window-i)).strftime("%m/%d")
                for i in range(bt_window)
            ]

            fig_bt = go.Figure()
            # 실제 손익
            fig_bt.add_trace(go.Bar(
                x=dates_bt, y=pnl.tolist(),
                name="실제 일별 손익",
                marker_color=[
                    "rgba(162,45,45,0.7)" if v < 0 else "rgba(55,138,221,0.5)"
                    for v in pnl
                ],
                showlegend=True,
            ))
            # VaR 선 (음수로 표시 — 손실 방향)
            fig_bt.add_trace(go.Scatter(
                x=dates_bt, y=(-var_).tolist(),
                mode="lines", name=f"VaR {bt_conf_sel} (음수)",
                line=dict(color="#A32D2D", width=1.8, dash="dash"),
            ))
            # 초과 시점 표시
            exc_dates = [dates_bt[i] for i in range(bt_window) if -pnl[i] > var_[i]]
            exc_vals  = [float(pnl[i]) for i in range(bt_window) if -pnl[i] > var_[i]]
            if exc_dates:
                fig_bt.add_trace(go.Scatter(
                    x=exc_dates, y=exc_vals,
                    mode="markers", name=f"VaR 초과 ({len(exc_dates)}일)",
                    marker=dict(color="#A32D2D", size=9, symbol="x"),
                ))
            fig_bt.update_layout(
                title=f"실제 손익 vs VaR — {bt_ticker}",
                xaxis_title="날짜",
                yaxis_title="손익 (원)",
                height=380,
                legend=dict(orientation="h", y=-0.18),
                margin=dict(t=50, b=70, l=60, r=20),
            )
            st.plotly_chart(fig_bt, use_container_width=True)

            # VaR 예외 분포 히스토그램
            fig_exc = go.Figure()
            fig_exc.add_trace(go.Histogram(
                x=(-pnl / var_ * 100).tolist(),
                nbinsx=40,
                marker_color="rgba(55,138,221,0.6)",
                name="손실/VaR 비율 (%)",
            ))
            fig_exc.add_vline(x=100, line_dash="dash", line_color="#A32D2D",
                               annotation_text="VaR 기준선 (100%)")
            fig_exc.update_layout(
                title="손실/VaR 비율 분포 (100% 초과 = VaR 예외)",
                xaxis_title="손실 / VaR (%)",
                yaxis_title="빈도",
                height=280,
                showlegend=False,
                margin=dict(t=50, b=40, l=50, r=20),
            )
            st.plotly_chart(fig_exc, use_container_width=True)

            # ── Basel 존 설명 ────────────────────────────────────────────
            st.markdown("##### Basel 신호등 기준 (250영업일)")
            bt_guide = pd.DataFrame({
                "존":      ["🟢 녹색 (Green)", "🟡 황색 (Yellow)", "🔴 적색 (Red)"],
                "초과 횟수": ["0~4일", "5~9일", "10일 이상"],
                "판정":    ["모형 적합 — 정상 운영",
                             "경고 — 모형 재검토 권고",
                             "모형 부적합 — 자본 승수 증가"],
                "자본 승수": ["×3.0", "×3.4~3.9", "×4.0"],
            })
            st.dataframe(bt_guide, use_container_width=True, hide_index=True)

            # 세션 저장 (리포트용)
            st.session_state["bt_result"] = {
                "ticker":     bt_ticker,
                "conf":       bt_conf,
                "window":     bt_window,
                "exceptions": int(exceptions),
                "exc_rate":   float(exc_rate),
                "zone":       zone,
                "lr_kupiec":  float(lr_kupiec),
                "p_kupiec":   float(p_kupiec),
                "kupiec_res": kupiec_result,
                "avg_var":    float(avg_var),
            }


with tab_report:
    st.markdown("#### 분석 리포트 생성")
    st.markdown(
        '<div class="ib">'
        'EWMA VaR / 시뮬레이션 / Back-testing 결과를 종합하여 '
        'Markdown(MD) 및 Excel(xlsx) 리포트를 자동 생성합니다.<br>'
        '각 분석 탭에서 계산을 먼저 실행한 뒤 이 탭에서 리포트를 생성하세요.'
        '</div>', unsafe_allow_html=True
    )

    # ── 포함 가능한 섹션 상태 표시 ────────────────────────────────────
    st.markdown("##### 분석 데이터 준비 상태")
    sc1, sc2, sc3, sc4 = st.columns(4)
    ewma_ready = bool(st.session_state.get("ewma_result"))
    sim_ready  = bool(st.session_state.get("sim_result_for_report"))
    bt_ready   = bool(st.session_state.get("bt_result"))
    mc_ready   = bool(st.session_state.get("mc_var_result"))
    with sc1:
        st.markdown(
            f"{'✅' if ewma_ready else '⬜'} **Parametric + Historical VaR**\n"
            f"({'완료' if ewma_ready else 'Parametric VaR 탭 실행 필요'})"
        )
    with sc2:
        st.markdown(
            f"{'✅' if mc_ready else '⬜'} **Monte Carlo VaR**\n"
            f"({'완료' if mc_ready else 'MC VaR 별도 실행 필요'})"
        )
    with sc3:
        st.markdown(
            f"{'✅' if sim_ready else '⬜'} **시뮬레이션 분석**\n"
            f"({'완료' if sim_ready else '시뮬레이션 분석 탭 실행 필요'})"
        )
    with sc4:
        st.markdown(
            f"{'✅' if bt_ready else '⬜'} **Back-testing**\n"
            f"({'완료' if bt_ready else 'Back-testing 탭 실행 필요'})"
        )

    st.markdown("---")

    rp_c1, rp_c2 = st.columns(2)
    with rp_c1:
        rp_pf_name = st.text_input("포트폴리오명", value=pf["name"], key="rp_name")
    with rp_c2:
        rp_author = st.text_input("작성자", value="Risk Management", key="rp_author")

    # ── 리포트 생성 버튼 ──────────────────────────────────────────────
    gen_col1, gen_col2 = st.columns(2)
    with gen_col1:
        gen_md = st.button("📄 Markdown 리포트 생성",
                            type="primary", use_container_width=True,
                            key="gen_md_btn")
    with gen_col2:
        gen_xl = st.button("📊 Excel 리포트 생성",
                            type="primary", use_container_width=True,
                            key="gen_xl_btn")

    if gen_md or gen_xl:
        ewma_res = st.session_state.get("ewma_result", {})
        sim_res  = st.session_state.get("sim_result_for_report", {})
        bt_res   = st.session_state.get("bt_result", {})
        mc_res   = st.session_state.get("mc_var_result", {})
        ts_str   = datetime.now().strftime("%Y%m%d_%H%M")
        safe_nm  = rp_pf_name.replace(" ", "_")

    if gen_md:
        with st.spinner("Markdown 리포트 생성 중..."):
            try:
                md_text = build_md_report(
                    pf_name       = rp_pf_name,
                    positions     = positions,
                    usdkrw        = usdkrw,
                    ewma_results  = ewma_res,
                    sim_results   = sim_res,
                    bt_result     = bt_res,
                    mc_var_result = mc_res,
                    author        = rp_author,
                )
                st.success("Markdown 리포트 생성 완료!")
                with st.expander("리포트 미리보기", expanded=True):
                    st.markdown(md_text)
                st.download_button(
                    label="⬇️ MD 파일 다운로드",
                    data=md_text.encode("utf-8"),
                    file_name=f"VaR_Report_{safe_nm}_{ts_str}.md",
                    mime="text/markdown",
                    key="dl_md_btn",
                )
            except Exception as e:
                st.error(f"Markdown 리포트 생성 오류: {e}")

    if gen_xl:
        with st.spinner("Excel 리포트 생성 중..."):
            try:
                xl_bytes = build_excel_report(
                    pf_name       = rp_pf_name,
                    positions     = positions,
                    usdkrw        = usdkrw,
                    ewma_results  = ewma_res,
                    sim_results   = sim_res,
                    bt_result     = bt_res,
                    mc_var_result = mc_res,
                )
                if xl_bytes:
                    st.success("Excel 리포트 생성 완료!")
                    st.download_button(
                        label="⬇️ Excel 파일 다운로드",
                        data=xl_bytes,
                        file_name=f"VaR_Report_{safe_nm}_{ts_str}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_xl_btn",
                    )
                else:
                    st.error("openpyxl 패키지가 필요합니다: pip install openpyxl")
            except Exception as e:
                st.error(f"Excel 리포트 생성 오류: {e}")

    st.markdown("---")
    st.markdown(
        '<div class="ib">'
        '<b>리포트 포함 데이터:</b> 포트폴리오 구성 · EWMA VaR(Marginal/Component/Incremental) '
        '· 시뮬레이션(MJD/Bootstrap/t-MC) · Back-testing(Kupiec 검정) · 기법 가이드<br>'
        '<b>Excel 시트:</b> 포트폴리오 구성 / EWMA VaR 분석 / 시뮬레이션 결과 / Back-testing / 기법 가이드'
        '</div>', unsafe_allow_html=True
    )

