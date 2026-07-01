import streamlit as st
import numpy as np
import pandas as pd
from scipy import stats
from datetime import datetime, timedelta
from math import log as mlog
import io
import warnings
warnings.filterwarnings("ignore")

try:
    import yfinance as yf; YF_OK = True
except ImportError:
    YF_OK = False

import plotly.graph_objects as go
from plotly.subplots import make_subplots

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter; OPX = True
except ImportError:
    OPX = False

# ── 상수 ──────────────────────────────────────────────────────────────────────
Z99, Z95 = 2.3263, 1.6449
COLORS = ["#185FA5", "#A32D2D", "#3B6D11", "#854F0B", "#534AB7", "#0F6E56"]

st.set_page_config(
    page_title="Portfolio VaR Calculator",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""<style>
[data-testid="stMetricValue"]{font-size:1.3rem}
.info-box{background:#f0f4ff;border-left:3px solid #185FA5;padding:9px 13px;
          border-radius:0 8px 8px 0;font-size:13px;color:#333;margin:6px 0;line-height:1.7}
</style>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# 핵심 수식 함수 (노트북 동일 로직)
# ══════════════════════════════════════════════════════════════════════════════
def ewma_vol(returns, lam=0.94):
    n = len(returns); v = np.zeros(n); v[0] = returns[0] ** 2
    for t in range(1, n):
        v[t] = lam * v[t - 1] + (1 - lam) * returns[t - 1] ** 2
    return np.sqrt(v)


def ewma_cov_mat(ret_matrix, lam=0.94):
    T, n = ret_matrix.shape
    w = np.array([(1 - lam) * lam ** (T - 1 - t) for t in range(T)])
    w /= w.sum()
    mu = ret_matrix.T @ w
    d = ret_matrix - mu
    return (d * w[:, None]).T @ d


def param_var(ret, mv, holding=10, conf=0.99, lam=0.94):
    sig = ewma_vol(ret, lam)[-1]
    z = stats.norm.ppf(conf)
    var = z * sig * np.sqrt(holding) * mv
    cvar = (stats.norm.pdf(z) / (1 - conf)) * sig * np.sqrt(holding) * mv
    return {"sig_d": sig, "var": var, "cvar": cvar, "var_pct": var / mv}


def hist_var(ret, mv, holding=10, conf=0.99):
    rets_h = ret * np.sqrt(holding)
    alpha = 1 - conf
    vp = float(-np.percentile(rets_h, alpha * 100))
    cm = rets_h <= -vp
    cp = float(-rets_h[cm].mean()) if cm.any() else vp
    return {"var": vp * mv, "cvar": cp * mv, "var_pct": vp}


def pf_var_ewma(weights, cov_mat, pf_mv, conf=0.99, holding=10):
    w = np.array(weights)
    pv = float(w @ cov_mat @ w)
    sig_p = np.sqrt(max(pv, 0))
    z = stats.norm.ppf(conf)
    pf_var = z * sig_p * np.sqrt(holding) * pf_mv
    mvar = z * np.sqrt(holding) * cov_mat @ w / sig_p if sig_p > 0 else np.zeros(len(w))
    comp = w * mvar * pf_mv
    beta = cov_mat @ w / pv if pv > 0 else np.zeros(len(w))
    return {"pf_var": pf_var, "pf_sigma": sig_p, "mvar": mvar, "comp": comp, "beta": beta}


def hist_pf_var(ret_dict, tickers, weights, total_mv, holding=10, conf=0.99):
    ml = min(len(ret_dict[tk]) for tk in tickers)
    pf = np.zeros(ml)
    for tk, w in zip(tickers, weights):
        pf += ret_dict[tk][-ml:] * w
    rets_h = pf * np.sqrt(holding)
    alpha = 1 - conf
    vp = float(-np.percentile(rets_h, alpha * 100))
    cm = rets_h <= -vp
    cp = float(-rets_h[cm].mean()) if cm.any() else vp
    return {"pf_var": vp * total_mv, "pf_cvar": cp * total_mv, "var_pct": vp}


def incr_var(weights, cov_mat, pf_mv, conf=0.99, holding=10):
    base = pf_var_ewma(weights, cov_mat, pf_mv, conf, holding)["pf_var"]
    w = np.array(weights)
    ivars = []
    for i in range(len(w)):
        mask = [j for j in range(len(w)) if j != i]
        if not mask:
            ivars.append(base)
            continue
        we = w[mask] / w[mask].sum()
        ce = cov_mat[np.ix_(mask, mask)]
        mve = pf_mv * w[mask].sum()
        ivars.append(base - pf_var_ewma(we, ce, mve, conf, holding)["pf_var"])
    return ivars


def mc_var_individual(ret, mv, holding=10, conf=0.99,
                      n=10000, use_t=True, df=5, lam=0.94, seed=42):
    rng = np.random.default_rng(seed)
    sig = ewma_vol(ret, lam)[-1]
    mu_d = float(np.mean(ret))
    if use_t and df > 2:
        scale = sig * np.sqrt((df - 2) / df)
        z = stats.t.rvs(df=df, size=n, random_state=int(rng.integers(0, 2 ** 31)))
        sim = mu_d * holding + scale * np.sqrt(holding) * z
    else:
        sim = rng.normal(mu_d * holding, sig * np.sqrt(holding), n)
    alpha = 1 - conf
    vp = float(-np.percentile(sim, alpha * 100))
    cm = sim <= -vp
    cp = float(-sim[cm].mean()) if cm.any() else vp
    return {"var": vp * mv, "cvar": cp * mv, "var_pct": vp, "sim": sim}


def mc_var_portfolio(ret_dict, tickers, weights, total_mv,
                     holding=10, conf=0.99, n=10000,
                     use_t=True, df=5, lam=0.94, seed=42):
    rng = np.random.default_rng(seed)
    ml = min(len(ret_dict[tk]) for tk in tickers)
    rm = np.column_stack([ret_dict[tk][-ml:] for tk in tickers])
    cov = ewma_cov_mat(rm, lam) * holding
    mu = rm.mean(axis=0) * holding
    w = np.array(weights)
    if use_t and df > 2:
        L = np.linalg.cholesky(cov + np.eye(len(w)) * 1e-10)
        scale = np.sqrt((df - 2) / df)
        z = stats.t.rvs(df=df, size=(n, len(w)),
                        random_state=int(rng.integers(0, 2 ** 31)))
        sim_mat = mu + (z * scale) @ L.T
    else:
        sim_mat = rng.multivariate_normal(mu, cov, n)
    pf_rets = sim_mat @ w
    alpha = 1 - conf
    vp = float(-np.percentile(pf_rets, alpha * 100))
    cm = pf_rets <= -vp
    cp = float(-pf_rets[cm].mean()) if cm.any() else vp
    ind_mc = [
        float(-np.percentile(sim_mat[:, i], alpha * 100)) * total_mv * w[i]
        for i in range(len(tickers))
    ]
    return {"pf_var": vp * total_mv, "pf_cvar": cp * total_mv,
            "var_pct": vp, "ind_mc": ind_mc, "pf_rets": pf_rets}


# ── 데이터 수집 (캐시) ─────────────────────────────────────────────────────────
@st.cache_data(ttl=300)
def get_usdkrw():
    if not YF_OK:
        return 1370.0, False
    try:
        p = yf.Ticker("USDKRW=X").fast_info.last_price
        if p and p > 0:
            return round(float(p), 2), True
    except Exception:
        pass
    return 1370.0, False


@st.cache_data(ttl=300)
def fetch_prices_bulk(tickers_tuple):
    result = {}
    if not YF_OK or not tickers_tuple:
        return result
    try:
        raw = yf.download(list(tickers_tuple), period="2d",
                          progress=False, auto_adjust=True)
        for tk in tickers_tuple:
            try:
                col = raw[tk]["Close"] if len(tickers_tuple) > 1 else raw["Close"]
                v = float(col.dropna().iloc[-1])
                if v > 0:
                    result[tk] = v
            except Exception:
                pass
    except Exception:
        pass
    for tk in tickers_tuple:
        if tk not in result and YF_OK:
            try:
                p = yf.Ticker(tk).fast_info.last_price
                if p and p > 0:
                    result[tk] = float(p)
            except Exception:
                pass
    return result


@st.cache_data(ttl=600)
def fetch_returns_cached(ticker, days=504):
    if YF_OK:
        try:
            end = datetime.today()
            start = end - timedelta(days=int(days * 1.6))
            d = yf.download(ticker,
                            start=start.strftime("%Y-%m-%d"),
                            end=end.strftime("%Y-%m-%d"),
                            progress=False, auto_adjust=True)
            if not d.empty and len(d) >= 60:
                px = d["Close"].squeeze()
                ret = np.log(px / px.shift(1)).dropna().values
                return ret[-days:], float(px.iloc[-1]), True
        except Exception:
            pass
    rng = np.random.default_rng(abs(hash(ticker)) % 9999)
    sigma = 0.22 / np.sqrt(252)
    return rng.normal(-0.5 * sigma ** 2, sigma, days), 50000.0, False


def fetch_returns_multi(tickers, days=504):
    ret_dict = {}; price_dict = {}; any_live = False
    for tk in tickers:
        r, px, live = fetch_returns_cached(tk, days)
        ret_dict[tk] = r; price_dict[tk] = px
        any_live = any_live or live
    return ret_dict, price_dict, any_live


# ── 포맷 헬퍼 ─────────────────────────────────────────────────────────────────
def fmt_n(v):
    a = abs(v)
    if a >= 1e12: return f"{v/1e12:.2f}조"
    if a >= 1e8:  return f"{v/1e8:.2f}억"
    if a >= 1e4:  return f"{v/1e4:,.0f}만"
    return f"{v:,.0f}"


# ── 포트폴리오 파일 파싱 ───────────────────────────────────────────────────────
def load_portfolio_file(file):
    name = file.name
    df = pd.read_csv(file) if name.endswith(".csv") else pd.read_excel(file)
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    if "ccy" not in df.columns:
        df["ccy"] = "KRW"
    if "market" not in df.columns:
        def guess(row):
            tk = str(row.get("ticker", "")).upper()
            t = str(row.get("type", "")).capitalize()
            if t == "Bond": return "KR-BOND"
            if t in ("Option", "Future"): return "KR-DERIV"
            if tk.endswith(".KS"): return "KR-KOSPI"
            if tk.endswith(".KQ"): return "KR-KOSDAQ"
            return "US-NASDAQ"
        df["market"] = df.apply(guess, axis=1)
    positions = []
    for _, row in df.iterrows():
        try:
            pos = {
                "ticker": str(row["ticker"]).strip().upper(),
                "type":   str(row["type"]).capitalize(),
                "qty":    float(row["qty"]),
                "price":  float(row["price"]),
                "ccy":    str(row.get("ccy", "KRW")).upper(),
                "market": str(row.get("market", "KR-KOSPI")),
                "mult":   float(row["mult"]) if "mult" in row and pd.notna(row.get("mult")) else 1.0,
            }
            for f in ["coupon", "maturity", "cf_freq", "face_value",
                      "strike", "callput", "iv", "mat", "rf"]:
                if f in row and pd.notna(row.get(f)):
                    pos[f] = row[f]
            positions.append(pos)
        except Exception:
            continue
    return positions


# ── 세션 상태 초기화 ───────────────────────────────────────────────────────────
DEFAULT_POSITIONS = [
    {"ticker": "005930.KS", "type": "Stock",  "qty": 100,   "price": 75000,
     "ccy": "KRW", "market": "KR-KOSPI", "mult": 1},
    {"ticker": "KR103501GA96", "type": "Bond", "qty": 10000, "price": 10000,
     "ccy": "KRW", "market": "KR-BOND",
     "coupon": 0.035, "maturity": 3.0, "cf_freq": 2, "face_value": 10000, "mult": 1},
    {"ticker": "K200C-255-202506", "type": "Option", "qty": 10, "price": 257.0,
     "ccy": "KRW", "market": "KR-DERIV",
     "strike": 255, "callput": "C", "iv": 0.18, "mat": 0.16, "rf": 0.035, "mult": 250000},
    {"ticker": "AAPL", "type": "Stock", "qty": 50, "price": 185,
     "ccy": "USD", "market": "US-NASDAQ", "mult": 1},
]

for key, default in [
    ("positions", DEFAULT_POSITIONS.copy()),
    ("returns",   {}),
    ("var_results", None),
]:
    if key not in st.session_state:
        st.session_state[key] = default


# ══════════════════════════════════════════════════════════════════════════════
# 사이드바 — 파라미터 설정
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.title("⚙️ 파라미터 설정")

    conf = st.selectbox(
        "신뢰수준", [0.99, 0.975, 0.95],
        format_func=lambda x: f"{x*100:.1f}%"
    )
    holding = st.number_input("보유기간 (영업일)", min_value=1, max_value=250, value=10)
    lam = st.slider("EWMA λ (감쇄계수)", 0.90, 0.99, 0.94, 0.01)
    obs_days = st.selectbox(
        "관측 기간", [252, 504, 756], index=1,
        format_func=lambda x: f"{x}일 ({x//252}년)"
    )

    st.divider()
    st.subheader("Monte Carlo")
    mc_n = st.number_input("시뮬레이션 횟수", min_value=1000, max_value=100000,
                           value=10000, step=1000)
    use_t = st.checkbox("t-분포 (Fat-tail 반영)", value=True)
    mc_df = st.number_input("자유도 ν", min_value=3, max_value=30, value=5) if use_t else 5
    dist_str = f"t-분포(ν={mc_df})" if use_t else "정규분포"

    st.divider()
    st.subheader("Back-testing")
    bt_window = st.selectbox("백테스팅 기간", [250, 500],
                              format_func=lambda x: f"{x}영업일")

    st.divider()
    usdkrw, fx_live = get_usdkrw()
    st.metric("USD/KRW", f"{usdkrw:,.2f}",
              delta="실시간" if fx_live else "기본값(1,370)")


# ══════════════════════════════════════════════════════════════════════════════
# 헤더
# ══════════════════════════════════════════════════════════════════════════════
st.title("📊 포트폴리오 VaR 분석")
st.caption("VaR_Cal_final.ipynb 기반 | Parametric · Historical · Monte Carlo | Back-testing | 리포트")

tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "① 포트폴리오 구성",
    "② 기초 통계량",
    "③ VaR 분석",
    "④ Back-testing",
    "⑤ 리포트 다운로드",
])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — 포트폴리오 구성
# ══════════════════════════════════════════════════════════════════════════════
with tab1:
    st.subheader("포트폴리오 구성")

    # 파일 업로드
    uploaded = st.file_uploader(
        "Excel / CSV 파일 업로드",
        type=["xlsx", "xls", "csv"],
        help="필수 컬럼: ticker, type, qty, price | 선택: ccy, market, coupon, maturity, cf_freq, face_value, strike, callput, iv, mat, rf, mult"
    )
    if uploaded:
        if st.button("파일 적용", type="primary"):
            try:
                st.session_state.positions = load_portfolio_file(uploaded)
                st.session_state.returns = {}
                st.session_state.var_results = None
                st.success(f"{len(st.session_state.positions)}개 포지션 로드 완료")
                st.rerun()
            except Exception as e:
                st.error(f"파일 오류: {e}")

    st.divider()

    # 종목 추가 폼
    with st.expander("➕ 종목 직접 추가"):
        c1, c2, c3, c4 = st.columns(4)
        a_ticker = c1.text_input("티커", placeholder="005930.KS")
        a_type   = c2.selectbox("유형", ["Stock", "Bond", "Option", "Future"])
        a_qty    = c3.number_input("수량", min_value=1, value=100)
        a_price  = c4.number_input("현재가", min_value=0.0, value=1000.0, format="%.2f")

        c5, c6, c7 = st.columns(3)
        a_ccy    = c5.selectbox("통화", ["KRW", "USD"])
        a_market = c6.selectbox("시장", ["KR-KOSPI", "KR-KOSDAQ", "KR-BOND",
                                         "KR-DERIV", "US-NYSE", "US-NASDAQ", "US-BOND"])
        a_mult   = c7.number_input("승수(mult)", min_value=1, value=1)

        if a_type == "Bond":
            b1, b2, b3, b4 = st.columns(4)
            a_coupon   = b1.number_input("쿠폰율", value=0.035, step=0.001, format="%.3f")
            a_maturity = b2.number_input("잔존만기(년)", value=3.0, step=0.5)
            a_freq     = b3.number_input("이표 횟수/년", value=2, min_value=1)
            a_face     = b4.number_input("액면가", value=10000.0)

        if a_type == "Option":
            o1, o2, o3, o4, o5 = st.columns(5)
            a_strike  = o1.number_input("행사가", value=255.0)
            a_callput = o2.selectbox("콜/풋", ["C", "P"])
            a_iv      = o3.number_input("내재변동성", value=0.18, step=0.01, format="%.2f")
            a_mat     = o4.number_input("잔존만기(년)", value=0.25, step=0.01, format="%.2f")
            a_rf      = o5.number_input("무위험금리", value=0.035, step=0.001, format="%.3f")

        col_add, col_reset = st.columns([1, 1])
        if col_add.button("추가", type="primary"):
            if not a_ticker.strip():
                st.warning("티커를 입력해 주세요.")
            else:
                new_pos = {
                    "ticker": a_ticker.strip().upper(), "type": a_type,
                    "qty": a_qty, "price": a_price,
                    "ccy": a_ccy, "market": a_market, "mult": a_mult,
                }
                if a_type == "Bond":
                    new_pos.update({"coupon": a_coupon, "maturity": a_maturity,
                                    "cf_freq": a_freq, "face_value": a_face})
                if a_type == "Option":
                    new_pos.update({"strike": a_strike, "callput": a_callput,
                                    "iv": a_iv, "mat": a_mat, "rf": a_rf})
                st.session_state.positions.append(new_pos)
                st.session_state.returns = {}
                st.session_state.var_results = None
                st.rerun()

        if col_reset.button("샘플 포트폴리오로 초기화"):
            st.session_state.positions = DEFAULT_POSITIONS.copy()
            st.session_state.returns = {}
            st.session_state.var_results = None
            st.rerun()

    # 실시간 가격 갱신
    if st.button("📡 실시간 주가 갱신 (yfinance)"):
        stock_tickers = tuple(
            p["ticker"] for p in st.session_state.positions if p["type"] == "Stock"
        )
        if stock_tickers:
            with st.spinner("실시간 가격 조회 중..."):
                fetch_prices_bulk.clear()
                live_px = fetch_prices_bulk(stock_tickers)
            updated = 0
            for p in st.session_state.positions:
                if p["type"] == "Stock" and p["ticker"] in live_px:
                    p["price"] = live_px[p["ticker"]]
                    updated += 1
            fetch_returns_cached.clear()
            st.session_state.returns = {}
            st.session_state.var_results = None
            st.success(f"{updated}개 종목 가격 업데이트 완료")
            st.rerun()

    # 포트폴리오 테이블
    positions = st.session_state.positions
    if not positions:
        st.info("포지션을 추가해 주세요.")
    else:
        mv_list  = [p["price"] * p["qty"] * (usdkrw if p["ccy"] == "USD" else 1) * p.get("mult", 1)
                    for p in positions]
        total_mv = sum(mv_list)

        table_data = []
        for i, (p, mv) in enumerate(zip(positions, mv_list)):
            table_data.append({
                "#":        i + 1,
                "티커":     p["ticker"],
                "유형":     p["type"],
                "시장":     p.get("market", ""),
                "수량":     p["qty"],
                "현재가":   p["price"],
                "통화":     p["ccy"],
                "평가금액(원)": round(mv),
                "비중":     round(mv / total_mv * 100, 1),
                "삭제":     False,
            })

        edited = st.data_editor(
            pd.DataFrame(table_data),
            column_config={"삭제": st.column_config.CheckboxColumn("삭제")},
            use_container_width=True,
            hide_index=True,
            disabled=["#", "티커", "유형", "시장", "수량", "현재가", "통화", "평가금액(원)", "비중"],
        )

        to_del = [i for i, row in edited.iterrows() if row["삭제"]]
        if to_del and st.button(f"선택 항목 삭제 ({len(to_del)}개)", type="secondary"):
            st.session_state.positions = [p for i, p in enumerate(positions) if i not in to_del]
            st.session_state.returns = {}
            st.session_state.var_results = None
            st.rerun()

        # 요약 지표
        st.divider()
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("총 평가금액", fmt_n(total_mv) + "원")
        m2.metric("종목 수", f"{len(positions)}개")
        m3.metric("주식", f"{sum(1 for p in positions if p['type']=='Stock')}개")
        m4.metric("채권/파생", f"{sum(1 for p in positions if p['type']!='Stock')}개")

        # 비중 파이 차트
        fig_pie = go.Figure(go.Pie(
            labels=[p["ticker"] for p in positions],
            values=mv_list,
            hole=0.4,
            marker_colors=COLORS * (len(positions) // len(COLORS) + 1),
            textinfo="label+percent",
        ))
        fig_pie.update_layout(title="포트폴리오 비중 (평가금액 기준)", height=380,
                              legend=dict(orientation="h", y=-0.15))
        st.plotly_chart(fig_pie, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — 기초 통계량
# ══════════════════════════════════════════════════════════════════════════════
with tab2:
    st.subheader("수익률 기초 통계량")

    positions = st.session_state.positions
    if not positions:
        st.warning("① 탭에서 포트폴리오를 먼저 구성해 주세요.")
    else:
        tickers = [p["ticker"] for p in positions]

        if st.button("📥 수익률 데이터 수집", type="primary"):
            with st.spinner(f"{len(tickers)}개 종목 수익률 수집 중 ({obs_days}일)..."):
                ret_dict, _, any_live = fetch_returns_multi(tickers, obs_days)
            st.session_state.returns = ret_dict
            st.session_state.var_results = None
            mode = "실시간" if any_live else "샘플(yfinance 미설치)"
            st.success(f"수익률 수집 완료 ({mode})")

        if not st.session_state.returns:
            st.info("'수익률 데이터 수집' 버튼을 눌러주세요.")
        else:
            ret_dict = st.session_state.returns
            from scipy.stats import skew as _skew, kurtosis as _kurt, jarque_bera as _jb

            stat_rows = []
            for tk in tickers:
                if tk not in ret_dict:
                    continue
                r = ret_dict[tk]
                jb_s, jb_p = _jb(r)
                stat_rows.append({
                    "티커":          tk,
                    "관측수(일)":    len(r),
                    "연율수익률(%)": round(np.mean(r) * 252 * 100, 2),
                    "연율변동성(%)": round(np.std(r) * np.sqrt(252) * 100, 2),
                    "왜도":          round(float(_skew(r)), 3),
                    "초과첨도":      round(float(_kurt(r)), 3),
                    "JB p-value":    round(float(jb_p), 4),
                    "정규분포 검정": "기각 ❌" if jb_p < 0.05 else "채택 ✅",
                })

            st.dataframe(pd.DataFrame(stat_rows), use_container_width=True, hide_index=True)

            st.markdown("""<div class="info-box">
            <b>해석 가이드</b><br>
            · <b>왜도</b> &lt; 0 → 좌측 꼬리 두꺼움 (손실 쪽 극단 사건 빈번)<br>
            · <b>초과첨도</b> &gt; 0 → 정규분포 대비 Fat-tail → 정규 VaR 과소 추정 가능<br>
            · <b>JB 기각</b> → 정규분포 가정 부적절 → Historical·MC VaR 병행 권장
            </div>""", unsafe_allow_html=True)

            # 수익률 히스토그램
            st.divider()
            st.subheader("개별 종목 수익률 분포")
            n_col = min(len(tickers), 2)
            cols = st.columns(n_col)
            for i, tk in enumerate(tickers):
                if tk not in ret_dict:
                    continue
                r = ret_dict[tk]
                x_line = np.linspace(r.min(), r.max(), 200)
                mu_r, sig_r = np.mean(r), np.std(r)
                pdf = stats.norm.pdf(x_line, mu_r, sig_r) * len(r) * (r.max() - r.min()) / 50

                fig = go.Figure()
                fig.add_trace(go.Histogram(
                    x=r.tolist(), nbinsx=50,
                    marker_color=COLORS[i % len(COLORS)], opacity=0.7, name="실제"
                ))
                fig.add_trace(go.Scatter(
                    x=x_line.tolist(), y=pdf.tolist(), mode="lines",
                    line=dict(color="#A32D2D", width=2), name="정규분포"
                ))
                fig.update_layout(title=tk, height=260, showlegend=False,
                                  margin=dict(t=35, b=10, l=10, r=10))
                cols[i % n_col].plotly_chart(fig, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — VaR 분석
# ══════════════════════════════════════════════════════════════════════════════
with tab3:
    st.subheader("VaR 분석")

    positions = st.session_state.positions
    if not positions:
        st.warning("① 탭에서 포트폴리오를 먼저 구성해 주세요.")
    elif not st.session_state.returns:
        st.warning("② 탭에서 수익률 데이터를 먼저 수집해 주세요.")
    else:
        if st.button("🧮 VaR 계산 실행", type="primary"):
            ret_dict = st.session_state.returns
            tickers  = [p["ticker"] for p in positions]
            mv_list  = [p["price"] * p["qty"] * (usdkrw if p["ccy"] == "USD" else 1) * p.get("mult", 1)
                        for p in positions]
            total_mv = sum(mv_list)
            weights  = [m / total_mv for m in mv_list]

            with st.spinner("VaR 계산 중..."):
                ind_pvars = []; ind_hvars = []; ind_mcvars = []

                for p, mv in zip(positions, mv_list):
                    tk = p["ticker"]
                    r  = ret_dict.get(tk, np.zeros(obs_days))
                    ind_pvars.append(param_var(r, mv, holding, conf, lam)["var"])
                    ind_hvars.append(hist_var(r, mv, holding, conf)["var"])
                    ind_mcvars.append(mc_var_individual(r, mv, holding, conf, mc_n, use_t, mc_df, lam)["var"])

                valid_tickers = [tk for tk in tickers if tk in ret_dict]
                min_len = min(len(ret_dict[tk]) for tk in valid_tickers)
                ret_mat = np.column_stack([ret_dict[tk][-min_len:] for tk in valid_tickers])
                cov_mat = ewma_cov_mat(ret_mat, lam)

                pf_res  = pf_var_ewma(weights, cov_mat, total_mv, conf, holding)
                h_res   = hist_pf_var(ret_dict, valid_tickers, weights, total_mv, holding, conf)
                pf_mc   = mc_var_portfolio(ret_dict, valid_tickers, weights, total_mv,
                                           holding, conf, mc_n, use_t, mc_df, lam)
                ivars   = incr_var(weights, cov_mat, total_mv, conf, holding)

            st.session_state.var_results = {
                "tickers": tickers, "mv_list": mv_list, "weights": weights,
                "total_mv": total_mv, "ind_pvars": ind_pvars,
                "ind_hvars": ind_hvars, "ind_mcvars": ind_mcvars,
                "pf_res": pf_res, "h_res": h_res, "pf_mc": pf_mc,
                "ivars": ivars, "cov_mat": cov_mat,
                "conf": conf, "holding": holding, "lam": lam,
                "dist_str": dist_str,
            }
            st.success("계산 완료!")

        vr = st.session_state.var_results
        if not vr:
            st.info("'VaR 계산 실행' 버튼을 눌러주세요.")
        else:
            tickers    = vr["tickers"];  mv_list  = vr["mv_list"]
            weights    = vr["weights"];  total_mv = vr["total_mv"]
            pf_res     = vr["pf_res"];   h_res    = vr["h_res"]
            pf_mc      = vr["pf_mc"];    ivars    = vr["ivars"]
            ind_pvars  = vr["ind_pvars"]; ind_hvars = vr["ind_hvars"]
            ind_mcvars = vr["ind_mcvars"]
            conf_      = vr["conf"];     holding_ = vr["holding"]
            dist_str_  = vr["dist_str"]

            # ── 포트폴리오 VaR 요약 ──────────────────────────────────────────
            st.subheader(f"포트폴리오 VaR ({int(conf_*100)}%, {holding_}일)")
            c1, c2, c3 = st.columns(3)
            c1.metric("Parametric VaR (EWMA)",
                      fmt_n(pf_res["pf_var"]) + "원",
                      f"{pf_res['pf_var']/total_mv*100:.3f}%")
            c2.metric("Historical Simulation VaR",
                      fmt_n(h_res["pf_var"]) + "원",
                      f"{h_res['pf_var']/total_mv*100:.3f}%")
            c3.metric(f"Monte Carlo VaR ({dist_str_})",
                      fmt_n(pf_mc["pf_var"]) + "원",
                      f"{pf_mc['pf_var']/total_mv*100:.3f}%")

            undiv_p  = sum(ind_pvars)
            div_eff  = undiv_p - pf_res["pf_var"]
            st.markdown(f"""<div class="info-box">
            포트폴리오 σ({holding_}일): <b>{pf_res['pf_sigma']*100:.4f}%</b> &nbsp;|&nbsp;
            단순합 VaR: <b>{fmt_n(undiv_p)}원</b> &nbsp;|&nbsp;
            분산 효과: <b>{fmt_n(div_eff)}원 ({div_eff/undiv_p*100:.1f}%)</b>
            </div>""", unsafe_allow_html=True)

            # 3종류 VaR 막대
            fig_3var = go.Figure(go.Bar(
                x=["Parametric", "Historical", f"MC({dist_str_})"],
                y=[pf_res["pf_var"]/1e6, h_res["pf_var"]/1e6, pf_mc["pf_var"]/1e6],
                marker_color=["#A32D2D", "#854F0B", "#534AB7"],
                text=[fmt_n(v) + "원" for v in
                      [pf_res["pf_var"], h_res["pf_var"], pf_mc["pf_var"]]],
                textposition="outside",
            ))
            fig_3var.update_layout(
                title=f"3종류 포트폴리오 VaR 비교 (백만원) | {int(conf_*100)}%, {holding_}일",
                height=320, showlegend=False, yaxis_title="VaR (백만원)"
            )
            st.plotly_chart(fig_3var, use_container_width=True)

            # ── 개별 종목 VaR ────────────────────────────────────────────────
            st.divider()
            st.subheader("개별 종목 VaR 상세")

            comp_rows = []
            for i, tk in enumerate(tickers):
                comp    = float(pf_res["comp"][i])
                contrib = comp / pf_res["pf_var"] * 100 if pf_res["pf_var"] else 0
                comp_rows.append({
                    "티커":            tk,
                    "비중(%)":         round(weights[i] * 100, 1),
                    "평가금액(억)":    round(mv_list[i] / 1e8, 3),
                    "P-VaR(M)":        round(ind_pvars[i] / 1e6, 2),
                    "H-VaR(M)":        round(ind_hvars[i] / 1e6, 2),
                    "MC-VaR(M)":       round(ind_mcvars[i] / 1e6, 2),
                    "Marginal VaR":    round(float(pf_res["mvar"][i]), 4),
                    "Component VaR(M)": round(comp / 1e6, 2),
                    "기여도(%)":       round(contrib, 1),
                    "IVaR(M)":         round(ivars[i] / 1e6, 2),
                    "Beta":            round(float(pf_res["beta"][i]), 3),
                })

            st.dataframe(pd.DataFrame(comp_rows), use_container_width=True, hide_index=True)

            # 종목별 3종류 VaR 비교
            fig_ind = go.Figure()
            for vals, name, color in [
                (ind_pvars, "P-VaR", "#A32D2D"),
                (ind_hvars, "H-VaR", "#854F0B"),
                (ind_mcvars, f"MC-VaR", "#534AB7"),
            ]:
                fig_ind.add_trace(go.Bar(
                    x=tickers, y=[v / 1e6 for v in vals], name=name,
                    marker_color=color,
                    text=[f"{v/1e6:.1f}M" for v in vals], textposition="outside"
                ))
            fig_ind.update_layout(
                title=f"종목별 3종류 VaR 비교 (백만원)",
                barmode="group", height=380,
                legend=dict(orientation="h", y=-0.2)
            )
            st.plotly_chart(fig_ind, use_container_width=True)

            # ── Component VaR 기여도 파이 ────────────────────────────────────
            st.divider()
            col_pie, col_mc = st.columns(2)

            with col_pie:
                st.subheader("Component VaR 기여도")
                fig_cpie = go.Figure(go.Pie(
                    labels=tickers,
                    values=[abs(float(pf_res["comp"][i])) for i in range(len(tickers))],
                    hole=0.4,
                    marker_colors=COLORS * (len(tickers) // len(COLORS) + 1),
                ))
                fig_cpie.update_layout(height=380, legend=dict(orientation="h", y=-0.15))
                st.plotly_chart(fig_cpie, use_container_width=True)

            # ── Monte Carlo 손익 분포 ─────────────────────────────────────────
            with col_mc:
                st.subheader("Monte Carlo 손익 분포")
                fig_mc_hist = go.Figure()
                fig_mc_hist.add_trace(go.Histogram(
                    x=(pf_mc["pf_rets"] * total_mv / 1e6).tolist(),
                    nbinsx=60, marker_color="rgba(83,74,183,0.55)", name="MC 손익"
                ))
                for val, color, label in [
                    (pf_mc["pf_var"],   "#534AB7", f"MC VaR {int(conf_*100)}%"),
                    (pf_res["pf_var"],  "#A32D2D", "P-VaR"),
                    (h_res["pf_var"],   "#854F0B", "H-VaR"),
                ]:
                    fig_mc_hist.add_vline(
                        x=-val / 1e6, line_dash="dash", line_color=color,
                        annotation_text=label, annotation_position="top right"
                    )
                fig_mc_hist.update_layout(
                    title=f"MC 손익 분포 ({dist_str_}, {mc_n:,}회)",
                    xaxis_title="손익 (백만원)", height=380, showlegend=False
                )
                st.plotly_chart(fig_mc_hist, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — Back-testing
# ══════════════════════════════════════════════════════════════════════════════
with tab4:
    st.subheader("VaR Back-testing")

    positions = st.session_state.positions
    stock_positions = [p for p in positions if p["type"] == "Stock"]

    if not stock_positions:
        st.warning("Back-testing은 주식 포지션이 필요합니다. ① 탭에서 주식을 추가해 주세요.")
    else:
        bt_col1, bt_col2 = st.columns(2)
        bt_tk   = bt_col1.selectbox("백테스팅 종목", [p["ticker"] for p in stock_positions])
        bt_conf = bt_col2.selectbox("신뢰수준", [0.99, 0.975, 0.95],
                                    format_func=lambda x: f"{x*100:.1f}%", key="bt_conf")

        if st.button("📈 Back-testing 실행", type="primary"):
            n_total = bt_window + 252
            bt_pos  = next((p for p in stock_positions if p["ticker"] == bt_tk), None)
            mv_bt   = bt_pos["price"] * bt_pos["qty"] * (usdkrw if bt_pos["ccy"] == "USD" else 1)

            with st.spinner("데이터 수집 및 Back-testing 계산 중..."):
                bt_ret, _, _ = fetch_returns_cached(bt_tk, n_total)

                pnl_s = bt_ret * mv_bt
                vol_s = ewma_vol(pnl_s / mv_bt, lam) * mv_bt
                z_bt  = stats.norm.ppf(bt_conf)
                var_s = z_bt * vol_s
                var_s = np.roll(var_s, 1); var_s[0] = var_s[1]

                pnl_ = pnl_s[-bt_window:]
                var_ = var_s[-bt_window:]

                exc   = int((-pnl_ > var_).sum())
                exc_r = exc / bt_window * 100
                zone  = ("🟢 녹색 (모형 적합)" if exc <= 4 else
                         "🟡 황색 (경고)"       if exc <= 9 else
                         "🔴 적색 (부적합)")

                p_th = 1 - bt_conf; k = exc; n = bt_window
                lr_k = (2 * (k * mlog(k / n / p_th + 1e-12) +
                             (n - k) * mlog((n - k) / n / (1 - p_th) + 1e-12))
                        if k > 0 else 0.0)
                p_k  = 1 - stats.chi2.cdf(lr_k, df=1)
                kup  = "기각 ❌ (부적합)" if p_k < 0.05 else "채택 ✅ (적합)"

            # 결과 지표
            r1, r2, r3, r4 = st.columns(4)
            r1.metric("VaR 초과 횟수",    f"{exc}회",        f"{exc_r:.2f}%")
            r2.metric("Basel 판정",        zone)
            r3.metric("Kupiec LR 통계량",  f"{lr_k:.3f}")
            r4.metric("Kupiec 검정 결과",  kup,              f"p-value = {p_k:.4f}")

            st.markdown("""<div class="info-box">
            <b>Basel III 기준 (250영업일)</b><br>
            🟢 녹색 (0~4회) → 모형 적합, 자본승수 ×3.0<br>
            🟡 황색 (5~9회) → 경고, 자본승수 ×3.4~3.9<br>
            🔴 적색 (10회+) → 모형 부적합, 자본승수 ×4.0<br><br>
            <b>Kupiec POF 검정</b>: LR = 2[k·ln(p_e/p_0) + (n-k)·ln((1-p_e)/(1-p_0))] ~ χ²(1)
            </div>""", unsafe_allow_html=True)

            # 시각화
            dates_bt = [
                (datetime.today() - timedelta(days=bt_window - i)).strftime("%y/%m/%d")
                for i in range(bt_window)
            ]
            exc_idx = [i for i in range(bt_window) if -pnl_[i] > var_[i]]

            fig_bt = go.Figure()
            fig_bt.add_trace(go.Bar(
                x=dates_bt, y=pnl_.tolist(), name="실제 손익",
                marker_color=[
                    "rgba(162,45,45,0.7)" if v < 0 else "rgba(55,138,221,0.5)"
                    for v in pnl_
                ]
            ))
            fig_bt.add_trace(go.Scatter(
                x=dates_bt, y=(-var_).tolist(), mode="lines",
                name=f"VaR {int(bt_conf*100)}%",
                line=dict(color="#A32D2D", dash="dash", width=2)
            ))
            if exc_idx:
                fig_bt.add_trace(go.Scatter(
                    x=[dates_bt[i] for i in exc_idx],
                    y=[float(pnl_[i]) for i in exc_idx],
                    mode="markers",
                    name=f"VaR 초과 ({len(exc_idx)}회)",
                    marker=dict(color="#A32D2D", size=10, symbol="x")
                ))
            fig_bt.update_layout(
                title=f"Back-testing: {bt_tk} | {int(bt_conf*100)}% | {bt_window}영업일",
                height=440, legend=dict(orientation="h", y=-0.2)
            )
            st.plotly_chart(fig_bt, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 — 리포트 다운로드
# ══════════════════════════════════════════════════════════════════════════════
with tab5:
    st.subheader("리포트 생성 및 다운로드")

    vr = st.session_state.var_results
    if not vr:
        st.warning("③ VaR 분석 탭에서 계산을 먼저 실행해 주세요.")
    else:
        positions  = st.session_state.positions
        tickers    = vr["tickers"];  mv_list  = vr["mv_list"]
        weights    = vr["weights"];  total_mv = vr["total_mv"]
        pf_res     = vr["pf_res"];   h_res    = vr["h_res"]
        pf_mc      = vr["pf_mc"];    ivars    = vr["ivars"]
        ind_pvars  = vr["ind_pvars"]; ind_hvars = vr["ind_hvars"]
        ind_mcvars = vr["ind_mcvars"]
        conf_      = vr["conf"];     holding_ = vr["holding"]
        dist_str_  = vr["dist_str"]
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ts  = datetime.now().strftime("%Y%m%d_%H%M")

        col_xl, col_md = st.columns(2)

        # ── Excel 리포트 ──────────────────────────────────────────────────────
        with col_xl:
            st.subheader("📥 Excel 리포트")
            if not OPX:
                st.error("openpyxl 미설치. `pip install openpyxl`")
            else:
                if st.button("Excel 생성 후 다운로드", type="primary"):
                    wb = Workbook(); wb.remove(wb.active)
                    thin = Side(style="thin", color="BFBFBF")
                    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

                    def hd(ws, r, cols, texts, bg="1F3864", fg="FFFFFF"):
                        for c, t in zip(cols, texts):
                            cell = ws.cell(row=r, column=c, value=t)
                            cell.font      = Font(bold=True, color=fg, size=10, name="Arial")
                            cell.fill      = PatternFill("solid", start_color=bg)
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.border    = bdr

                    def cl(ws, r, c, v, bold=False, fmt=None,
                           fg="000000", bg=None, align="right"):
                        cell = ws.cell(row=r, column=c, value=v)
                        cell.font      = Font(bold=bold, color=fg, name="Arial", size=10)
                        cell.alignment = Alignment(horizontal=align, vertical="center")
                        cell.border    = bdr
                        if bg:  cell.fill          = PatternFill("solid", start_color=bg)
                        if fmt: cell.number_format = fmt

                    # Sheet 1: 포트폴리오
                    ws1 = wb.create_sheet("포트폴리오")
                    ws1.sheet_view.showGridLines = False
                    for col_l, w_v in zip("ABCDEFG", [22, 10, 12, 10, 18, 10, 10]):
                        ws1.column_dimensions[col_l].width = w_v
                    ws1.row_dimensions[1].height = 30
                    ws1.merge_cells("A1:G1")
                    ws1["A1"] = f"포트폴리오 구성  |  {now}"
                    ws1["A1"].font      = Font(bold=True, color="FFFFFF", size=13, name="Arial")
                    ws1["A1"].fill      = PatternFill("solid", start_color="1F3864")
                    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
                    hd(ws1, 2, range(1, 8),
                       ["티커", "유형", "시장", "수량", "평가금액(원)", "통화", "비중"],
                       bg="2F5496")
                    for i, (p, mv) in enumerate(zip(positions, mv_list)):
                        bg = "F2F2F2" if i % 2 == 0 else "FFFFFF"; rr = i + 3
                        cl(ws1, rr, 1, p["ticker"], bold=True, align="left", bg=bg)
                        cl(ws1, rr, 2, p["type"],   align="center", bg=bg)
                        cl(ws1, rr, 3, p.get("market", ""), align="center", bg=bg)
                        cl(ws1, rr, 4, p["qty"],  fmt="#,##0", bg=bg)
                        cl(ws1, rr, 5, mv,        fmt="#,##0", bg=bg)
                        cl(ws1, rr, 6, p["ccy"],  align="center", bg=bg)
                        cl(ws1, rr, 7, mv / total_mv, fmt="0.00%", bg=bg)

                    # Sheet 2: VaR 비교
                    ws2 = wb.create_sheet("VaR 비교")
                    ws2.sheet_view.showGridLines = False
                    for ci, ww in enumerate([28, 18, 18, 10, 20], 1):
                        ws2.column_dimensions[get_column_letter(ci)].width = ww
                    ws2.row_dimensions[1].height = 30
                    ws2.merge_cells("A1:E1")
                    ws2["A1"] = f"VaR 비교  |  {int(conf_*100)}%  |  {holding_}일  |  {now}"
                    ws2["A1"].font      = Font(bold=True, color="FFFFFF", size=13, name="Arial")
                    ws2["A1"].fill      = PatternFill("solid", start_color="1F3864")
                    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
                    hd(ws2, 2, range(1, 6),
                       ["방법", "포트폴리오 VaR(원)", "CVaR(원)", "VaR%", "비고"],
                       bg="2F5496")
                    var_data = [
                        ("Parametric (EWMA)", pf_res["pf_var"], pf_res["pf_var"] * 1.25,
                         pf_res["pf_var"] / total_mv, f"EWMA λ={vr['lam']}"),
                        ("Historical Simulation", h_res["pf_var"], h_res["pf_cvar"],
                         h_res["pf_var"] / total_mv, "분포 가정 없음"),
                        (f"Monte Carlo ({dist_str_})", pf_mc["pf_var"], pf_mc["pf_cvar"],
                         pf_mc["pf_var"] / total_mv, dist_str_),
                    ]
                    for i, (nm, v, cv, vp, note) in enumerate(var_data):
                        bg = "F2F2F2" if i % 2 == 0 else "FFFFFF"
                        cl(ws2, 3 + i, 1, nm,   bold=True, align="left", bg=bg)
                        cl(ws2, 3 + i, 2, v,    fmt="#,##0", fg="C00000", bg=bg)
                        cl(ws2, 3 + i, 3, cv,   fmt="#,##0", fg="7F6000", bg=bg)
                        cl(ws2, 3 + i, 4, vp,   fmt="0.000%", bg=bg)
                        cl(ws2, 3 + i, 5, note, align="left", bg=bg)

                    # Sheet 3: 종목별 VaR
                    ws3 = wb.create_sheet("종목별 VaR")
                    ws3.sheet_view.showGridLines = False
                    for ci, ww in enumerate([20, 8, 14, 12, 14, 12, 14, 14, 10, 12, 8], 1):
                        ws3.column_dimensions[get_column_letter(ci)].width = ww
                    ws3.row_dimensions[1].height = 30
                    ws3.merge_cells("A1:K1")
                    ws3["A1"] = f"종목별 VaR 상세  |  {int(conf_*100)}%  |  {holding_}일"
                    ws3["A1"].font      = Font(bold=True, color="FFFFFF", size=13, name="Arial")
                    ws3["A1"].fill      = PatternFill("solid", start_color="1F3864")
                    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
                    hd(ws3, 2, range(1, 12),
                       ["티커", "비중", "P-VaR", "H-VaR", "H-VaR%",
                        "MC-VaR", "Component VaR", "기여도", "IVaR", "Beta", "평가금액"],
                       bg="2F5496")
                    for i, tk in enumerate(tickers):
                        bg   = "F2F2F2" if i % 2 == 0 else "FFFFFF"; rr = i + 3
                        comp = float(pf_res["comp"][i])
                        ct   = comp / pf_res["pf_var"] * 100 if pf_res["pf_var"] else 0
                        cl(ws3, rr,  1, tk,              bold=True, align="left", bg=bg)
                        cl(ws3, rr,  2, weights[i],      fmt="0.0%",  bg=bg)
                        cl(ws3, rr,  3, ind_pvars[i],    fmt="#,##0", fg="C00000", bg=bg)
                        cl(ws3, rr,  4, ind_hvars[i],    fmt="#,##0", fg="C00000", bg=bg)
                        cl(ws3, rr,  5, ind_hvars[i]/mv_list[i], fmt="0.000%", bg=bg)
                        cl(ws3, rr,  6, ind_mcvars[i],   fmt="#,##0", fg="C00000", bg=bg)
                        cl(ws3, rr,  7, comp,             fmt="#,##0", bg=bg)
                        cl(ws3, rr,  8, ct / 100,         fmt="0.0%",  bg=bg)
                        cl(ws3, rr,  9, ivars[i],         fmt="#,##0", bg=bg)
                        cl(ws3, rr, 10, float(pf_res["beta"][i]), fmt="0.000", bg=bg)
                        cl(ws3, rr, 11, mv_list[i],       fmt="#,##0", bg=bg)

                    buf = io.BytesIO()
                    wb.save(buf); buf.seek(0)
                    st.download_button(
                        "⬇️ Excel 다운로드",
                        data=buf.getvalue(),
                        file_name=f"VaR_Report_{ts}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

        # ── Markdown 리포트 ───────────────────────────────────────────────────
        with col_md:
            st.subheader("📄 Markdown 리포트")
            L = []; a = L.append
            a("# 포트폴리오 리스크 분석 리포트"); a("")
            a(f"| 작성일시 | {now} | 신뢰수준 | {int(conf_*100)}% | 보유기간 | {holding_}일 |")
            a("|------|------|------|------|------|------|"); a("")
            a("## 1. 포트폴리오 구성"); a("")
            a(f"- 총 평가금액: **{fmt_n(total_mv)}원**  |  종목 수: **{len(positions)}개**"); a("")
            a("| 티커 | 유형 | 수량 | 평가금액 | 비중 |")
            a("|------|------|-----:|--------:|-----:|")
            for p, mv in zip(positions, mv_list):
                a(f"| {p['ticker']} | {p['type']} | {p['qty']:,.0f} "
                  f"| {fmt_n(mv)}원 | {mv/total_mv*100:.1f}% |")
            a(""); a("---"); a("## 2. VaR 비교 (3종류)"); a("")
            a("| 방법 | 포트폴리오 VaR | CVaR | VaR% |")
            a("|------|-------------:|-----:|-----:|")
            a(f"| Parametric (EWMA) | {fmt_n(pf_res['pf_var'])}원 "
              f"| {fmt_n(pf_res['pf_var']*1.25)}원 | {pf_res['pf_var']/total_mv*100:.3f}% |")
            a(f"| Historical Simulation | {fmt_n(h_res['pf_var'])}원 "
              f"| {fmt_n(h_res['pf_cvar'])}원 | {h_res['pf_var']/total_mv*100:.3f}% |")
            a(f"| Monte Carlo ({dist_str_}) | {fmt_n(pf_mc['pf_var'])}원 "
              f"| {fmt_n(pf_mc['pf_cvar'])}원 | {pf_mc['pf_var']/total_mv*100:.3f}% |")
            undiv_p = sum(ind_pvars); div_eff = undiv_p - pf_res["pf_var"]
            a("")
            a(f"- **분산 효과**: {fmt_n(div_eff)}원 ({div_eff/undiv_p*100:.1f}%)")
            a(f"- **포트폴리오 σ({holding_}일)**: {pf_res['pf_sigma']*100:.4f}%")
            a(""); a("---"); a("## 3. 리스크 기여도 (Component VaR)"); a("")
            a("| 티커 | 비중 | P-VaR | H-VaR | MC-VaR | Component | 기여도 | IVaR | Beta |")
            a("|------|-----:|------:|------:|-------:|----------:|-------:|-----:|-----:|")
            for i, tk in enumerate(tickers):
                comp = float(pf_res["comp"][i])
                ct   = comp / pf_res["pf_var"] * 100 if pf_res["pf_var"] else 0
                a(f"| {tk} | {weights[i]*100:.1f}% "
                  f"| {ind_pvars[i]/1e6:.1f}M | {ind_hvars[i]/1e6:.1f}M "
                  f"| {ind_mcvars[i]/1e6:.1f}M | {comp/1e6:.1f}M "
                  f"| {ct:.1f}% | {ivars[i]/1e6:.1f}M | {float(pf_res['beta'][i]):.3f} |")
            a(""); a("---")
            a(f"*Generated by VaR_Cal_app.py  |  {now}*")

            md_text = "\n".join(L)
            st.download_button(
                "⬇️ Markdown 다운로드",
                data=md_text.encode("utf-8"),
                file_name=f"VaR_Report_{ts}.md",
                mime="text/markdown",
            )
            with st.expander("미리보기"):
                st.markdown(md_text)
